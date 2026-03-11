import sys
import socket
import paramiko
import re
import json
import time
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QComboBox, QTextEdit,
    QGroupBox, QTabWidget, QScrollArea, QProgressBar,
    QTextBrowser, QFrame, QSizePolicy, QMessageBox, QSpacerItem,
    QTableWidget, QTableWidgetItem, QHeaderView, QSplitter,
    QPlainTextEdit, QInputDialog, QFileDialog,
    QDialog, QListWidget, QListWidgetItem, QDialogButtonBox
)
from PyQt5.QtCore import Qt, pyqtSignal, QObject, QDate, QThread, QTimer, QPropertyAnimation, QEasingCurve
from PyQt5.QtGui import QFont, QTextCursor, QColor
from PyQt5.QtWidgets import QDateEdit
import os
import stat
import shutil
import logging
from logging.handlers import RotatingFileHandler
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from functools import wraps

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")


def setup_logging():
    path = "C:/tmp"

    # Check if the directory exists
    if not os.path.exists(path):
        # Create the directory
        os.makedirs(path)
        print(f"Directory {path} created.")
    else:
        print(f"Directory {path} already exists.")

    LOG_FILE = os.path.join(path, f'test_station_interface_{timestamp}.log')

    # Custom formatter that includes function name
    class ContextFormatter(logging.Formatter):
        def format(self, record):
            # Only add function name if it's not already there
            if not hasattr(record, 'func_name'):
                record.func_name = "Internal_Function_driver"
            return super().format(record)

    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)

    # File handler
    file_handler = RotatingFileHandler(
        LOG_FILE,
        maxBytes=1024 * 1024,
        backupCount=5,
        encoding='utf-8'
    )
    file_handler.setLevel(logging.DEBUG)

    # Console handler
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)

    # Formatter with function name
    formatter = ContextFormatter(
        '%(asctime)s - %(name)s - %(levelname)s - [%(func_name)s] - %(message)s'
    )
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    # Decorator to add function name to log records
    def log_function(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            logger = logging.getLogger(func.__module__)
            logger.debug(f"Entering {func.__name__}", extra={'func_name': func.__name__})
            try:
                result = func(*args, **kwargs)
                logger.debug(f"Exiting {func.__name__}", extra={'func_name': func.__name__})
                return result
            except Exception as e:
                logger.error(f"Error in {func.__name__}: {str(e)}",
                             exc_info=True,
                             extra={'func_name': func.__name__})
                raise

        return wrapper

    return logger, log_function, LOG_FILE


# Initialize logging
logger, log_function, log_file_path = setup_logging()


class ExcelLogger:
    SHEET_PASSWORD = os.environ.get("EXCEL_SHEET_PASSWORD", "Admin@1234")

    def __init__(self, file_path=os.path.join('C:/tmp', f'test_station_records_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')):
        self.file_path = file_path
        self.workbook = None
        self.unit_sheet = None
        self.interlock_sheet = None
        self.self_test_sheet = None
        self.BNC_sheet = None
        self.resistance_sheet = None
        self.Imp_sheet = None
        self.logger1 = logger.getChild('ExcelLogger')
        self.pn=''
        self.sn=''
        self.excel_time=datetime.now().strftime("%Y%m%d_%H%M%S")
        self._is_finalized = False        # Set True after update_overall_result makes file read-only
        self._current_result = None       # Tracks worst result seen; FAIL is sticky (never downgraded to PASS)
        self._log_file_path = log_file_path  # Path to the session log file

        # Create the directory if it doesn't exist
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        # Initialize or load the workbook
        self._init_workbook()

    @log_function
    def _init_workbook(self):
        """Initialize or load the workbook"""
        if os.path.exists(self.file_path):
            try:
                self.workbook = load_workbook(self.file_path)
            except Exception as e:
                self.logger1.error(f"Error loading existing workbook:", exc_info=True,
                                   extra={'func_name': 'init_workbook'})
                self._create_new_workbook()
        else:
            self._create_new_workbook()

    @log_function
    def _create_new_workbook(self):
        """Create a new empty workbook"""
        self.workbook = Workbook()

        # Instead of removing the default sheet, just rename it to one of your expected sheets
        default_sheet = self.workbook.active
        default_sheet.title = "Unit Setup"  # Or whichever sheet you expect to use first

        # Create headers for the default sheet
        self._create_unit_headers(default_sheet)

        try:
            self.workbook.save(self.file_path)
            self.logger1.info(f"Created new Excel file at {self.file_path}",
                              extra={'func_name': 'create_new_workbook'})
        except Exception as e:
            self.logger1.error(f"Error saving new workbook: {e}", exc_info=True,
                               extra={'func_name': 'create_new_workbook'})

    @staticmethod
    def _freq_to_sheet_suffix(freq_text):
        """Convert frequency text to a safe sheet name suffix e.g. '60 MHz' -> '60MHz'"""
        return freq_text.replace(' ', '').replace('.', '_')

    def _protect_all_sheets(self):
        """Apply password protection to every worksheet in the workbook.

        After protection is enabled users can still navigate and read cells,
        but any attempt to edit cell content, insert/delete rows or columns,
        or change formatting will prompt for the password before the action
        is allowed.  The password is taken from the ``SHEET_PASSWORD`` class
        attribute (defaults to the ``EXCEL_SHEET_PASSWORD`` environment
        variable or "Admin@1234" if that variable is not set).
        """
        for sheet in self.workbook.worksheets:
            sheet.protection.sheet = True
            sheet.protection.password = self.SHEET_PASSWORD
            sheet.protection.enable()

    def _save_workbook(self):
        """Save the workbook to disk, unless the file has already been finalized.

        After ``update_overall_result`` password-protects the sheets and marks
        the file read-only, any further save attempt would raise a
        ``PermissionError``.  This helper checks the ``_is_finalized`` flag and
        silently skips the save when the file is already locked.
        """
        if self._is_finalized:
            self.logger1.debug(
                "Skipping save: workbook is already finalized (read-only).",
                extra={'func_name': '_save_workbook'}
            )
            return
        self.workbook.save(self.file_path)

    @log_function
    def reset_sheet(self, sheet_name):
        """Clear all data from a specific sheet (except headers) and recreate headers if needed"""
        try:
            if sheet_name not in self.workbook.sheetnames:
                self.logger1.warning(f"Sheet '{sheet_name}' does not exist",
                                     extra={'func_name': 'reset_sheet'})
                return False

            sheet = self.workbook[sheet_name]

            # "Unit Setup" has no header row, so clear from row 1;
            # all other sheets keep the header in row 1.
            if sheet_name == "Unit Setup":
                if sheet.max_row >= 1:
                    sheet.delete_rows(1, sheet.max_row)
            elif sheet.max_row > 1:
                sheet.delete_rows(2, sheet.max_row)  # Delete from row 2 to end

            # Reapply headers based on sheet type
            if sheet_name == "Interlock Test":
                self._create_interlock_headers(sheet)
            elif sheet_name == "Self Test":
                self._create_self_test_headers(sheet)
            elif sheet_name == "Zone1-Inner_Res_scan":
                self._create_resistance_headers(sheet)
            elif sheet_name == "Zone2-Mid_Inner_Res_scan":
                self._create_resistance_headers(sheet)
            elif sheet_name == "Zone3-Mid_Edge_Res_scan":
                self._create_resistance_headers(sheet)
            elif sheet_name == "Zone4-Edge_Res_scan":
                self._create_resistance_headers(sheet)
            elif sheet_name == "Zone5-Outer_Res_scan":
                self._create_resistance_headers(sheet)
            elif "_Imp_scan" in sheet_name:
                self._create_impedance_headers(sheet)
            elif sheet_name == "BNC Port Verification":
                self._create_BNC_headers(sheet)
            elif sheet_name == "Unit Setup":
                self._create_unit_headers(sheet)

            self._save_workbook()
            self.logger1.info(f"Reset sheet '{sheet_name}' successfully",
                              extra={'func_name': 'reset_sheet'})
            return True
        except Exception as e:
            self.logger1.error(f"Error resetting sheet '{sheet_name}': {e}", exc_info=True,
                               extra={'func_name': 'reset_sheet'})
            return False

    def _ensure_sheet_exists(self, sheet_name, create_headers_func):
        """Ensure a sheet exists, creating it if necessary"""
        if sheet_name not in self.workbook.sheetnames:
            sheet = self.workbook.create_sheet(sheet_name)
            create_headers_func(sheet)
            self._save_workbook()
            return sheet
        return self.workbook[sheet_name]

    def _create_unit_headers(self, sheet):
        """Create headers for the unit setup sheet"""
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 30

    def _create_interlock_headers(self, sheet):
        """Create headers for the interlock test sheet"""
        sheet.column_dimensions['A'].width = 20  # Timestamp
        sheet.column_dimensions['B'].width = 20  # Test Name
        sheet.column_dimensions['C'].width = 15  # Open Count
        sheet.column_dimensions['D'].width = 10  # Closed Count
        sheet.column_dimensions['E'].width = 15  # Test Result
        sheet.column_dimensions['F'].width = 30  # Notes

        headers = [
            "Timestamp", "Test Name", "Open Count", "Closed Count",
            "Test Result", "Notes"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        sheet.freeze_panes = "A2"

    def _create_self_test_headers(self, sheet):
        """Create headers for the self test sheet"""
        sheet.column_dimensions['A'].width = 20  # Timestamp
        sheet.column_dimensions['B'].width = 20  # Unit Identifier
        sheet.column_dimensions['C'].width = 15  # Test Result
        sheet.column_dimensions['D'].width = 30  # Test Details
        sheet.column_dimensions['E'].width = 30  # Notes

        headers = [
            "Timestamp", "Unit Identifier", "Test Result",
            "Test Details", "Notes"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        sheet.freeze_panes = "A2"

    def _create_resistance_headers(self, sheet):
        """Create headers for the resistance measurements sheet"""
        sheet.column_dimensions['A'].width = 20  # Timestamp
        sheet.column_dimensions['B'].width = 20  # Zone
        sheet.column_dimensions['C'].width = 15  # Setpoint
        sheet.column_dimensions['D'].width = 15  # Resistance
        sheet.column_dimensions['E'].width = 10  # Status
        sheet.column_dimensions['F'].width = 10  # Table Row

        headers = [
            "Timestamp", "Zone", "Setpoint", "Resistance (Ω)",
            "Status", "Table Row"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        sheet.freeze_panes = "A2"

    def _create_impedance_headers(self, sheet):
        """Create headers for the impedance measurements sheet"""
        sheet.column_dimensions['A'].width = 20  # Timestamp
        sheet.column_dimensions['B'].width = 20  # Zone
        sheet.column_dimensions['C'].width = 20  # Frequency
        sheet.column_dimensions['D'].width = 15  # Setpoint
        sheet.column_dimensions['E'].width = 15  # Real
        sheet.column_dimensions['F'].width = 15  # Image
        sheet.column_dimensions['G'].width = 15  # Impedance
        sheet.column_dimensions['H'].width = 10  # Status

        headers = [
            "Timestamp", "Zone", "Frequency", "Setpoint",
            "Real(Ω)", "Imaginary", "Impedance(Z)", "Status"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        sheet.freeze_panes = "A2"

    def _create_BNC_headers(self, sheet):
        """Create headers for the BNC Port Verification sheet"""
        sheet.column_dimensions['A'].width = 20  # Timestamp
        sheet.column_dimensions['B'].width = 20  # Zone NAME
        sheet.column_dimensions['C'].width = 20  # TEST_VALUE
        sheet.column_dimensions['D'].width = 15  # STATUS

        headers = [
            "Timestamp", "Zone", "Value(db)", "Status"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        sheet.freeze_panes = "A2"

    def _create_summary_headers(self, sheet):
            """Create headers for the Summary sheet with additional metadata fields"""
            # Set up metadata fields in rows 1-9
            metadata_fields = [
                ("EID", ""),
                ("SERIAL NUMBER", ""),
                ("MODEL NUMBER", ""),
                ("VERSION", ""),
                ("TESTER NAME", ""),
                ("COMMENT", ""),
                ("START TIME", ""),
                ("END TIME", ""),
                ("OVERALL RESULT", ""),
                ("TEST FIXTURE SN", ""),
                ("VNA SN", ""),
                ("ECAL SN", "")
            ]

            # Write metadata fields with formatting
            for row_num, (field, _) in enumerate(metadata_fields, start=1):
                # Field name cell
                sheet.cell(row=row_num, column=1, value=field)
                sheet.cell(row=row_num, column=1).font = Font(bold=True)

                # Value cell (empty initially)
                sheet.cell(row=row_num, column=2, value="")

                # Format for OVERALL RESULT
                if field == "OVERALL RESULT":
                    sheet.cell(row=row_num, column=2).font = Font(bold=True)
                    sheet.cell(row=row_num, column=2).alignment = Alignment(horizontal='center')

            # Add space between metadata and TESTSTEP section
            sheet.row_dimensions[13].height = 15

            # TESTSTEP section (now starting at row 14)
            sheet.cell(row=14, column=1, value="TESTSTEP").font = Font(bold=True, color="FFFFFF")
            sheet.cell(row=14, column=1).fill = PatternFill(
                start_color="404040", end_color="404040", fill_type="solid")

            sheet.cell(row=14, column=2, value="STATUS").font = Font(bold=True, color="FFFFFF")
            sheet.cell(row=14, column=2).fill = PatternFill(
                start_color="404040", end_color="404040", fill_type="solid")

            # Add space between sections
            sheet.row_dimensions[21].height = 15

            # STEP section headers (now starting at row 22 from column A)
            headers_row22 = ["Step", "Unit", "Low_Limit", "Measure", "High_Limit",
                             "TestStep", "TestPoints", "Status"]
            for col_num, header in enumerate(headers_row22, start=1):  # Starting at column A (1)
                cell = sheet.cell(row=22, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Set column widths
            sheet.column_dimensions['A'].width = 20  # Step
            sheet.column_dimensions['B'].width = 15  # Unit
            sheet.column_dimensions['C'].width = 15  # Low_Limit
            sheet.column_dimensions['D'].width = 15  # Measure
            sheet.column_dimensions['E'].width = 15  # High_Limit
            sheet.column_dimensions['F'].width = 15  # TestStep
            sheet.column_dimensions['G'].width = 15  # TestPoints
            sheet.column_dimensions['H'].width = 15  # Status

            sheet.freeze_panes = "A23"  # Freeze above the STEP section

    @log_function
    def update_overall_result(self, result, PN='NA', SN='NA', finalize=True):
        """Update the overall result and rename file accordingly.

        Args:
            result:   'PASS' or 'FAIL' (case-insensitive).
            PN:       Product Number – stored on the logger when provided.
            SN:       Serial Number  – stored on the logger when provided.
            finalize: When *True* (the default) the workbook is password-
                      protected, saved, and the file is made read-only so it
                      cannot be modified after the test run.  Pass
                      ``finalize=False`` for intermediate per-step calls that
                      should only rename the file without locking it, so that
                      subsequent logging calls can still write data.
        """
        try:
            result = result.upper()
            if result not in ['PASS', 'FAIL']:
                self.logger1.warning(f"Invalid result: {result}. Must be 'PASS' or 'FAIL'")
                return False

            # FAIL is sticky: once any test section has produced a FAIL the
            # overall result must remain FAIL, regardless of later PASS results
            # from other test sections (e.g. BNC passing after resistance fails).
            if self._current_result == 'FAIL':
                result = 'FAIL'
            else:
                self._current_result = result  # promote: None → value, PASS → FAIL

            if PN != 'NA' and SN != 'NA':
                self.pn = PN
                self.sn = SN

            # Create new filename based on result
            new_filename = f"{self.pn}_{self.sn}_{self.excel_time}_{result}.xlsx"
            # The containing folder uses only PN_SN_TIME (no PASS/FAIL suffix)
            # so the folder name is stable regardless of test outcome, e.g.:
            #   C:\tmp\PN_SN_TIME\PN_SN_TIME_PASS.xlsx
            folder_name = f"{self.pn}_{self.sn}_{self.excel_time}"
            new_file_path = os.path.join("C:\\tmp", folder_name, new_filename)
            self.logger1.info(f"Target file path: {new_file_path}",
                              extra={'func_name': 'update_overall_result'})

            # If file already exists with different name, rename/move it
            if self.file_path != new_file_path:
                if os.path.exists(self.file_path):
                    # If the file was already finalized as read-only (e.g. it
                    # was previously named _PASS.xlsx), we must temporarily
                    # restore write access so the rename succeeds.
                    if self._is_finalized:
                        os.chmod(self.file_path,
                                 stat.S_IRUSR | stat.S_IWUSR | stat.S_IRGRP | stat.S_IROTH)
                        self._is_finalized = False
                        self.logger1.info(
                            "Temporarily restored write access to allow rename.",
                            extra={'func_name': 'update_overall_result'}
                        )

                    # Capture old directory before we update self.file_path
                    old_dir = os.path.dirname(os.path.abspath(self.file_path))

                    # Ensure the destination subfolder exists before moving
                    os.makedirs(os.path.dirname(new_file_path), exist_ok=True)

                    # Close the workbook before renaming
                    self.workbook.close()

                    # Move/rename the file into its new subfolder
                    os.rename(self.file_path, new_file_path)
                    self.file_path = new_file_path

                    # Reopen the workbook
                    self.workbook = load_workbook(self.file_path)
                    self.logger1.info(f"Renamed file to: {new_file_path}",
                                      extra={'func_name': 'update_overall_result'})

                    # Clean up the old directory if it is now empty, is
                    # different from the new file's parent, and is NOT the
                    # base C:\tmp folder (to avoid accidentally removing it).
                    new_dir = os.path.dirname(os.path.abspath(new_file_path))
                    _tmp_root = os.path.abspath("C:\\tmp")
                    if (old_dir != new_dir
                            and os.path.normcase(old_dir) != os.path.normcase(_tmp_root)
                            and os.path.isdir(old_dir)):
                        try:
                            os.rmdir(old_dir)  # only succeeds if the directory is empty
                        except OSError:
                            pass  # not empty or permission issue — leave it

            # Apply protection and read-only only when finalize=True AND the
            # workbook has not already been finalized in this state.
            if finalize and not self._is_finalized and os.path.exists(self.file_path):
                # Apply password-protection to all sheets before locking the file
                self._protect_all_sheets()
                self.workbook.save(self.file_path)
                self.logger1.info(f"Applied sheet protection to: {self.file_path}",
                                  extra={'func_name': 'update_overall_result'})
                os.chmod(self.file_path,
                         stat.S_IRUSR | stat.S_IRGRP | stat.S_IROTH)
                self.logger1.info(f"Set file as read-only: {self.file_path}",
                                  extra={'func_name': 'update_overall_result'})
                self._is_finalized = True
                self.logger1.info("Workbook finalized; further saves are disabled.",
                                  extra={'func_name': 'update_overall_result'})

                # Copy the session log file into the same subfolder as the xlsx
                # so that both artefacts are co-located in C:\tmp\<name_no_ext>\
                dest_dir = os.path.dirname(self.file_path)
                if self._log_file_path and os.path.exists(self._log_file_path) and os.path.isdir(dest_dir):
                    try:
                        dest_log = os.path.join(dest_dir, os.path.basename(self._log_file_path))
                        shutil.copy2(self._log_file_path, dest_log)
                        self.logger1.info(f"Copied log file to: {dest_log}",
                                          extra={'func_name': 'update_overall_result'})
                    except Exception:
                        self.logger1.warning("Could not copy log file into result folder.",
                                             exc_info=True,
                                             extra={'func_name': 'update_overall_result'})

            return True

        except Exception as e:
            self.logger1.error(f"Error updating overall result: {e}", exc_info=True,
                               extra={'func_name': 'update_overall_result'})
            return False

    @log_function
    def log_unit_setup(self, unit_data):
        """Log unit setup data to the Excel file"""
        try:
            # Ensure sheet exists
            self.unit_sheet = self._ensure_sheet_exists(
                "Unit Setup",
                self._create_unit_headers
            )

            # Get the next available row
            row_num = 1 if self.unit_sheet.max_row == 1 and all(
                cell.value is None for cell in self.unit_sheet[1]) else self.unit_sheet.max_row + 1

            fields = [
                ("Vendor Name", unit_data.get('Vendor_name', '')),
                ("Fixture Number", unit_data.get('Fixture_number', '')),
                ("Test Operator Name", unit_data.get('test_operator_name', '')),
                ("Test Date", unit_data.get('test_date', '')),
                ("VNA Calibration Date", unit_data.get('vna_calibration_date', '')),
                ("VNA SN", unit_data.get('vna_sn', '')),
                ("Ecal SN", unit_data.get('ecal_sn', '')),
                ("PCB Control Part Number", unit_data.get('pcb_part_number', '')),
                ("PCB Control Revision", unit_data.get('pcb_revision', '')),
                ("PCB Control Serial Number", unit_data.get('pcb_serial_number', '')),
                ("Assembly Part Number", unit_data.get('assembly_part_number', '')),
                ("Assembly Revision", unit_data.get('assembly_revision', '')),
                ("Assembly Serial Number", unit_data.get('assembly_serial_number', '')),
                ("Product ID", unit_data.get('product_id', '')),
                ("ESI Revision", unit_data.get('esi_revision', '')),
                ("Configuration ID", unit_data.get('configuration_id', '')),
                ("EtherCAT Address", unit_data.get('ethercat_address', '')),
                ("Firmware Version", unit_data.get('firmware_version', ''))
            ]

            for i, (field_name, field_value) in enumerate(fields, start=0):
                field_row = row_num + i
                self.unit_sheet.cell(row=field_row, column=1, value=field_name)
                self.unit_sheet.cell(row=field_row, column=1).font = Font(bold=True)
                self.unit_sheet.cell(row=field_row, column=2, value=field_value)
                for col in [1, 2]:
                    self.unit_sheet.cell(row=field_row, column=col).alignment = Alignment(
                        horizontal='left', vertical='center'
                    )

            self.unit_sheet.append([])
            self._save_workbook()
            self.logger1.info(f"Logged unit setup data to {self.file_path}",
                              extra={'func_name': 'log_unit_setup'})
            return True
        except Exception as e:
            self.logger1.error(f"Error logging unit setup data: {e}", exc_info=True,
                               extra={'func_name': 'log_unit_setup'})
            return False

    @log_function
    def log_interlock_test(self, test_name, test_passed, open_count, closed_count, notes=""):
        """Log interlock test results to the Excel file"""
        try:
            # Ensure sheet exists
            self.interlock_sheet = self._ensure_sheet_exists(
                "Interlock Test",
                self._create_interlock_headers
            )

            # Get the next available row
            row_num = self.interlock_sheet.max_row + 1

            # Write data with timestamp
            self.interlock_sheet.cell(
                row=row_num, column=1,
                value=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
            self.interlock_sheet.cell(row=row_num, column=2, value=test_name)

            # Test Result with color coding
            result_cell = self.interlock_sheet.cell(
                row=row_num, column=5,
                value="PASS" if test_passed else "FAIL"
            )
            result_cell.font = Font(bold=True)
            result_cell.fill = PatternFill(
                start_color="00AA00" if test_passed else "FF0000",
                end_color="00AA00" if test_passed else "FF0000",
                fill_type="solid"
            )
            result_cell.alignment = Alignment(horizontal='center')

            self.interlock_sheet.cell(row=row_num, column=3, value=open_count)
            self.interlock_sheet.cell(row=row_num, column=4, value=closed_count)
            self.interlock_sheet.cell(row=row_num, column=6, value=notes)

            # Save the workbook
            self._save_workbook()
            self.logger1.info(f"Logged interlock test result to {self.file_path}",
                              extra={'func_name': 'Log_interlock_test'})
            return True
        except Exception as e:
            self.logger1.error(f"Error logging interlock test result: {e}", exc_info=True,
                               extra={'func_name': 'Log_interlock_test'})
            return False

    @log_function
    def log_self_test(self, unit_identifier, test_passed, test_details="", notes=""):
        """Log self test results to the Excel file"""
        try:
            # Ensure sheet exists
            self.self_test_sheet = self._ensure_sheet_exists(
                "Self Test",
                self._create_self_test_headers
            )

            # Get the next available row
            row_num = self.self_test_sheet.max_row + 1

            # Write data with timestamp
            self.self_test_sheet.cell(
                row=row_num, column=1,
                value=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
            self.self_test_sheet.cell(row=row_num, column=2, value=unit_identifier)

            # Test Result with color coding
            result_cell = self.self_test_sheet.cell(
                row=row_num, column=3,
                value="PASS" if test_passed else "FAIL"
            )
            result_cell.font = Font(bold=True)
            result_cell.fill = PatternFill(
                start_color="00AA00" if test_passed else "FF0000",
                end_color="00AA00" if test_passed else "FF0000",
                fill_type="solid"
            )
            result_cell.alignment = Alignment(horizontal='center')

            self.self_test_sheet.cell(row=row_num, column=4, value=test_details)
            self.self_test_sheet.cell(row=row_num, column=5, value=notes)

            # Save the workbook
            self._save_workbook()
            self.logger1.info(f"Logged self test result to {self.file_path}",
                              extra={'func_name': 'Log_self_test'})
            return True
        except Exception as e:
            self.logger1.error(f"Error logging self test result: {e}", exc_info=True,
                               extra={'func_name': 'Log_self_test'})
            return False

    @log_function
    def log_resistance_measurement(self, measurement_data, sheet_name):
        """Log resistance measurement to combined sheet"""
        try:
            # Ensure sheet exists
            self.resistance_sheet = self._ensure_sheet_exists(
                sheet_name,
                self._create_resistance_headers
            )

            # Write measurement data
            row_num = self.resistance_sheet.max_row + 1
            self.resistance_sheet.cell(row=row_num, column=1, value=measurement_data['timestamp'])
            self.resistance_sheet.cell(row=row_num, column=2, value=measurement_data['zone_title'])
            self.resistance_sheet.cell(row=row_num, column=3, value=measurement_data['setpoint'])
            self.resistance_sheet.cell(row=row_num, column=4, value=measurement_data['resistance'])

            # Status with color coding
            status_cell = self.resistance_sheet.cell(
                row=row_num, column=5,
                value=measurement_data['status']
            )
            if measurement_data['status'] == "PASS":
                status_cell.fill = PatternFill(
                    start_color="00AA00",  # Green
                    fill_type="solid"
                )
            else:
                status_cell.fill = PatternFill(
                    start_color="FF0000",  # Red
                    fill_type="solid"
                )

            self.resistance_sheet.cell(row=row_num, column=6, value=measurement_data['table_row'])

            # Save workbook
            self._save_workbook()
            self.logger1.info(
                f"Logged resistance data to combined worksheet",
                extra={'func_name': 'log_resistance_measurement'}
            )
            return True
        except Exception as e:
            self.logger1.error(
                f"Error logging resistance data: {str(e)}",
                exc_info=True,
                extra={'func_name': 'log_resistance_measurement'}
            )
            return False


    @log_function
    def log_summary(self, metadata=None, teststep_data=None, step_data=None, update_existing=True):
        """Log data to the Summary sheet with option to update existing entries"""
        try:
            # Ensure sheet exists
            summary_sheet = self._ensure_sheet_exists(
                "Summary",
                self._create_summary_headers
            )

            # Update metadata if provided
            if metadata:
                metadata_mapping = {
                    'eid': 1,
                    'serial_number': 2,
                    'model_number': 3,
                    'version': 4,
                    'tester_name': 5,
                    'comment': 6,
                    'start_time': 7,
                    'end_time': 8,
                    'overall_result': 9,
                    'test_fixture_sn': 10,
                    'vna_sn': 11,
                    'ecal_sn': 12
                }

                for key, value in metadata.items():
                    if key.lower() in metadata_mapping:
                        row_num = metadata_mapping[key.lower()]
                        summary_sheet.cell(row=row_num, column=2, value=value)

                        if key.lower() == 'overall_result':
                            result_cell = summary_sheet.cell(row=row_num, column=2)
                            result_cell.font = Font(bold=True)
                            result_cell.alignment = Alignment(horizontal='center')
                            if str(value).upper() == "PASS":
                                result_cell.fill = PatternFill(
                                    start_color="00AA00", end_color="00AA00", fill_type="solid")
                                result_cell.font = Font(color="FFFFFF", bold=True)
                            elif str(value).upper() == "FAIL":
                                result_cell.fill = PatternFill(
                                    start_color="FF0000", end_color="FF0000", fill_type="solid")
                                result_cell.font = Font(color="FFFFFF", bold=True)

            # Update TESTSTEP data directly if provided
            if teststep_data:
                teststep_name = teststep_data.get('teststep', '')
                teststep_status = teststep_data.get('status', '')

                if teststep_name:
                    # Find existing teststep in TESTSTEP section (rows 15-20, column 1)
                    teststep_updated = False
                    for row in range(15, 21):  # TESTSTEP section rows
                        existing_teststep = summary_sheet.cell(row=row, column=1).value
                        if existing_teststep == teststep_name:
                            # Update existing teststep
                            status_cell = summary_sheet.cell(row=row, column=2, value=teststep_status)
                            self._apply_status_formatting(status_cell, teststep_status)
                            teststep_updated = True
                            break

                    # If not found and we have space, add to first empty row
                    if not teststep_updated:
                        for row in range(15, 21):
                            existing_teststep = summary_sheet.cell(row=row, column=1).value
                            if not existing_teststep or existing_teststep == "":
                                # Add new teststep
                                summary_sheet.cell(row=row, column=1, value=teststep_name)
                                status_cell = summary_sheet.cell(row=row, column=2, value=teststep_status)
                                self._apply_status_formatting(status_cell, teststep_status)
                                teststep_updated = True
                                break

            # Log STEP data if provided
            if step_data:
                testpoints = step_data.get('testpoints', '')

                # Find existing row if update_existing is True - search in TestPoints column (column G, index 7)
                step_row = None
                if update_existing and testpoints:
                    # Search from row 23 onwards for matching testpoints
                    for row in range(23, summary_sheet.max_row + 1):
                        cell_value = summary_sheet.cell(row=row, column=7).value  # Column G (7) is TestPoints
                        if cell_value == testpoints:
                            step_row = row
                            break

                # If not found or not updating, use next available row
                if step_row is None:
                    step_row = 23
                    # Find first empty row in the STEP section
                    while (summary_sheet.cell(row=step_row, column=7).value is not None and
                           summary_sheet.cell(row=step_row, column=7).value != ""):
                        step_row += 1

                    # For new rows, write step data including the step column
                    summary_sheet.cell(row=step_row, column=1, value=step_data.get('step', ''))  # Column A: Step
                # For existing rows, DO NOT update the Step column (Column A) - keep it as is

                # Update columns B to H (Unit to Status) - preserving Step column (A)
                summary_sheet.cell(row=step_row, column=2, value=step_data.get('unit', ''))  # Column B: Unit
                summary_sheet.cell(row=step_row, column=3, value=step_data.get('low_limit', ''))  # Column C: Low_Limit
                summary_sheet.cell(row=step_row, column=4, value=step_data.get('measure', ''))  # Column D: Measure
                summary_sheet.cell(row=step_row, column=5,
                                   value=step_data.get('high_limit', ''))  # Column E: High_Limit
                summary_sheet.cell(row=step_row, column=6, value=step_data.get('teststep', ''))  # Column F: TestStep
                summary_sheet.cell(row=step_row, column=7, value=testpoints)  # Column G: TestPoints



                # Write STATUS with color coding in column H (index 8)
                status = step_data.get('status', '')
                status_cell = summary_sheet.cell(row=step_row, column=8, value=status)  # Column H: Status
                self._apply_status_formatting(status_cell, status)

            # If step_data was provided, update TESTSTEP section automatically based on STEP section data
            # BUT only update teststeps that are managed by STEP data (not manually set ones)
            if step_data:
                self._update_teststep_from_step_data_preserve_manual(summary_sheet)

            # AUTOMATICALLY UPDATE OVERALL RESULT BASED ON TESTSTEP STATUS (ROWS 15-20, COLUMN B)
            self._update_overall_result_based_on_teststep_status(summary_sheet)

            # Save the workbook
            self._save_workbook()
            self.logger1.info("Logged summary data successfully",
                              extra={'func_name': 'log_summary'})
            return True

        except Exception as e:
            self.logger1.error(f"Error logging summary data: {e}", exc_info=True,
                               extra={'func_name': 'log_summary'})
            return False

    @log_function
    def _update_teststep_from_step_data_preserve_manual(self, summary_sheet):
        """Update TESTSTEP section based on STEP section data but preserve manually set teststeps"""
        try:
            # Get current manually set teststeps (rows 15-20)
            manual_teststeps = {}
            for row in range(15, 21):
                teststep_name = summary_sheet.cell(row=row, column=1).value
                if teststep_name:
                    manual_teststeps[teststep_name] = {
                        'row': row,
                        'status': summary_sheet.cell(row=row, column=2).value
                    }

            # Dictionary to store teststep statuses from STEP data
            step_teststep_status_map = {}

            # Collect all teststeps and their statuses from STEP section (column F and H)
            for row in range(23, summary_sheet.max_row + 1):
                teststep_cell = summary_sheet.cell(row=row, column=6)  # Column F: TestStep
                status_cell = summary_sheet.cell(row=row, column=8)  # Column H: Status

                teststep_name = teststep_cell.value
                status_value = str(status_cell.value).upper().strip() if status_cell.value else ""

                # Only process teststeps that are NOT manually managed
                if teststep_name and teststep_name not in manual_teststeps:
                    if teststep_name not in step_teststep_status_map:
                        step_teststep_status_map[teststep_name] = "PASS"  # Start with PASS assumption

                    # If any step in a teststep fails, the entire teststep fails
                    if status_value == "FAIL":
                        step_teststep_status_map[teststep_name] = "FAIL"

            # Update TESTSTEP section (rows 15-20) - only for step-managed teststeps
            current_teststep_row = 15

            # First, keep existing manual teststeps
            for row in range(15, 21):
                teststep_name = summary_sheet.cell(row=row, column=1).value
                if teststep_name and teststep_name in manual_teststeps:
                    # Keep manual teststep as is
                    current_teststep_row += 1

            # Then add step-managed teststeps to remaining rows
            for teststep_name, status in step_teststep_status_map.items():
                if current_teststep_row > 20:  # Don't exceed the TESTSTEP section
                    break

                # Write teststep name
                summary_sheet.cell(row=current_teststep_row, column=1, value=teststep_name)

                # Write status with color coding
                status_cell = summary_sheet.cell(row=current_teststep_row, column=2, value=status)
                self._apply_status_formatting(status_cell, status)

                current_teststep_row += 1

            # Clear any remaining rows in TESTSTEP section
            for row in range(current_teststep_row, 21):
                # Only clear if not a manual teststep
                existing_teststep = summary_sheet.cell(row=row, column=1).value
                if existing_teststep not in manual_teststeps:
                    summary_sheet.cell(row=row, column=1, value="")
                    summary_sheet.cell(row=row, column=2, value="")

            self.logger1.info(
                f"Updated TESTSTEP section - Manual: {len(manual_teststeps)}, Step-managed: {len(step_teststep_status_map)}",
                extra={'func_name': '_update_teststep_from_step_data_preserve_manual'})

        except Exception as e:
            self.logger1.error(f"Error updating TESTSTEP from STEP data: {e}", exc_info=True,
                               extra={'func_name': '_update_teststep_from_step_data_preserve_manual'})

    @log_function
    def _apply_status_formatting(self, cell, status):
        """Apply consistent status formatting to a cell"""
        status = str(status).upper().strip() if status else ""
        cell.font = Font(bold=True)

        if status == "PASS":
            cell.fill = PatternFill(start_color="00AA00", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        elif status == "FAIL":
            cell.fill = PatternFill(start_color="FF0000", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

    @log_function
    def _update_overall_result_based_on_teststep_status(self, summary_sheet):
        """Update overall result based only on TESTSTEP status (rows 15-20, column B)"""
        try:
            overall_result = "PASS"  # Start with PASS assumption

            # Check TESTSTEP status from rows 15-20, column B
            for row in range(15, 21):  # TESTSTEP section rows 15-20
                status_cell = summary_sheet.cell(row=row, column=2)  # Column B: Status
                status_value = str(status_cell.value).upper().strip() if status_cell.value else ""

                # If any teststep status is FAIL or empty, overall result becomes FAIL
                if status_value == "FAIL" or status_value == "":
                    overall_result = "FAIL"
                    break  # No need to check further if we found one FAIL

            # Update overall result in row 9, column 2
            result_cell = summary_sheet.cell(row=9, column=2, value=overall_result)
            result_cell.font = Font(bold=True)
            result_cell.alignment = Alignment(horizontal='center')
            self._apply_status_formatting(result_cell, overall_result)

            self.logger1.info(f"Updated overall result to: {overall_result} (based on TESTSTEP status)",
                              extra={'func_name': '_update_overall_result_based_on_teststep_status'})

        except Exception as e:
            self.logger1.error(f"Error updating overall result from teststep status: {e}", exc_info=True,
                               extra={'func_name': '_update_overall_result_based_on_teststep_status'})

    @log_function
    def log_Imp_measurement(self, measurement_data, sheet_name):
        """Log impedance measurement to combined sheet"""
        try:
            # Ensure sheet exists
            self.Imp_sheet = self._ensure_sheet_exists(
                sheet_name,
                self._create_impedance_headers
            )

            row_num = self.Imp_sheet.max_row + 1
            self.Imp_sheet.cell(row=row_num, column=1, value=measurement_data['timestamp'])
            self.Imp_sheet.cell(row=row_num, column=2, value=measurement_data['zone_title'])
            self.Imp_sheet.cell(row=row_num, column=3, value=measurement_data['Frequency'])
            self.Imp_sheet.cell(row=row_num, column=4, value=measurement_data['setpoint'])
            self.Imp_sheet.cell(row=row_num, column=5, value=measurement_data['Real'])
            self.Imp_sheet.cell(row=row_num, column=6, value=measurement_data['Imag'])
            self.Imp_sheet.cell(row=row_num, column=7, value=measurement_data['Z'])
            self.Imp_sheet.cell(row=row_num, column=8, value=measurement_data['status'])

            # Status with color coding
            status_cell = self.Imp_sheet.cell(
                row=row_num, column=8,
                value=measurement_data['status']
            )
            if measurement_data['status'] == "PASS":
                status_cell.fill = PatternFill(
                    start_color="00AA00",  # Green
                    fill_type="solid"
                )
            else:
                status_cell.fill = PatternFill(
                    start_color="FF0000",  # Red
                    fill_type="solid"
                )

            # Save workbook
            self._save_workbook()
            self.logger1.info(
                f"Logged impedance data to combined worksheet",
                extra={'func_name': 'log_impedance_measurement'}
            )
            return True
        except Exception as e:
            self.logger1.error(
                f"Error logging impedance data: {str(e)}",
                exc_info=True,
                extra={'func_name': 'log_impedance_measurement'}
            )
            return False



    @log_function
    def log_BNC_measurement(self, test_zone, test_details, test_passed):
        """Log BNC measurement to combined sheet"""
        try:
            # Ensure sheet exists
            self.BNC_sheet = self._ensure_sheet_exists(
                "BNC Port Verification",
                self._create_BNC_headers
            )

            # Get the next available row
            row_num = self.BNC_sheet.max_row + 1

            # Write data with timestamp
            self.BNC_sheet.cell(
                row=row_num, column=1,
                value=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
            self.BNC_sheet.cell(row=row_num, column=2, value=test_zone)
            self.BNC_sheet.cell(row=row_num, column=3, value=test_details)
            # Test Result with color coding
            result_cell = self.BNC_sheet.cell(
                row=row_num, column=4,
                value="PASS" if test_passed else "FAIL"
            )
            result_cell.font = Font(bold=True)
            result_cell.fill = PatternFill(
                start_color="00AA00" if test_passed else "FF0000",
                end_color="00AA00" if test_passed else "FF0000",
                fill_type="solid"
            )
            result_cell.alignment = Alignment(horizontal='center')

            # Save the workbook
            self._save_workbook()
            self.logger1.info(f"Logged BNC test result to {self.file_path}",
                              extra={'func_name': 'Log_BNC_test'})
            return True
        except Exception as e:
            self.logger1.error(f"Error logging BNC test result: {e}", exc_info=True,
                               extra={'func_name': 'Log_BNC_test'})
            return False


# Initialize the Excel logger
excel_logger = ExcelLogger()


class SSH_setup:
    def __init__(self):
        self.is_connect = False
        self.ssh = None
        #self.host = "192.168.1.2"
        self.host = "10.119.9.225"
        self.port = 22
        self.username = "robot"
        self.password = "robot"
        self.script_path = "/home/robot/Manufacturing_test/aipc_beta/test.py ecat"
        self.timeout = 10  # seconds
        #self.config = self.load_config()
        self.logger2 = logger.getChild('SSH_setup')


    @log_function
    def Connect_RPI(self, host=None, port=None, username=None, password=None):
        try:
            # Use provided parameters or fall back to instance variables
            host = host or self.host
            port = port or self.port
            username = username or self.username
            password = password or self.password

            self.ssh = paramiko.SSHClient()
            self.ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            self.ssh.connect(
                host,
                port,
                username,
                password,
                timeout=self.timeout
            )
            self.is_connect = True

            # logger.info("SSH connection established successfully")
            self.logger2.info(f"SSH connection established successfully",
                              extra={'func_name': 'Connect_RPI'})

            return True, "Connected successfully"
        except paramiko.AuthenticationException:
            error_msg = "Authentication failed, please verify your credentials"
            self.logger2.error(f"{error_msg}", exc_info=True,
                               extra={'func_name': 'Connect_RPI'})
            return False, error_msg
        except paramiko.SSHException as e:
            error_msg = f"SSH error: {str(e)}"
            self.logger2.error(f"{error_msg}", exc_info=True,
                               extra={'func_name': 'Connect_RPI'})
            return False, error_msg
        except Exception as e:
            error_msg = f"Connection error: {str(e)}"
            self.logger2.error(f"{error_msg}", exc_info=True,
                               extra={'func_name': 'Connect_RPI'})
            return False, error_msg

    def SSH_com(self, command, script_path=None):
        if not self.is_connect or not self.ssh:
            return "", "Not connected to SSH"


        script_path = script_path or self.script_path

        try:
            stdin, stdout, stderr = self.ssh.exec_command(f'sudo python3 {script_path} {command}')
            stdout_data = stdout.read().decode()
            stderr_data = stderr.read().decode()
            return stdout_data, stderr_data
        except Exception as e:
            return "", f"Command execution failed: {str(e)}"

    def SSH_com_stream(self, script_path, command):
        if not self.is_connect:
            raise Exception("SSH connection not established")

        # Run the Python script
        stdin, stdout, stderr = self.ssh.exec_command(f'sudo python3 {script_path} {command}', get_pty=True)

        # Continuously read the output and error
        while True:
            line = stdout.readline()
            if not line:
                break
            yield line.strip()

    @log_function
    def SSH_disconnect(self):
        try:
            if self.ssh:
                self.ssh.close()
                # logger.info("SSH connection closed")
                self.logger2.info(f"SSH connection closed",
                                  extra={'func_name': 'SSH_disconnect'})

        except Exception as e:
            # logger.error(f"Error disconnecting SSH: {str(e)}")
            self.logger2.error(f"Error disconnecting SSH: {str(e)}", exc_info=True,
                               extra={'func_name': 'SSH_disconnect'})
        finally:
            self.is_connect = False


class Worker(QThread):
    output_ready = pyqtSignal(str)
    finished_signal = pyqtSignal()
    error_occurred = pyqtSignal(str)

    def __init__(self, ssh_handler, script_path, command, timeout=30):
        super().__init__()
        self.ssh_handler = ssh_handler
        self.script_path = script_path
        self.command = command
        self._is_running = True
        self.work_timeout = 30
        self.t = timeout
        self.logger3 = logger.getChild('SSH_setup')

    def run(self):
        try:
            if not self._is_running:
                return

            # Connect SSH
            success, message = self.ssh_handler.Connect_RPI()
            if not success:
                self.error_occurred.emit(f"SSH Connection Failed: {message}")
                return

            # Execute command
            stdin, stdout, stderr = self.ssh_handler.ssh.exec_command(
                f'sudo python3 {self.script_path} {self.command}',
                get_pty=True,
                timeout=self.t
            )

            self.logger3.info(f"Command output received",
                             extra={'func_name': f'sudo python3 {self.script_path} {self.command}'})

            if stdout:
                self.logger3.debug(f"stdout:\n{stdout}",
                                  extra={'func_name': f'sudo python3 {self.script_path} {self.command}'})
            if stderr:
                self.logger3.error(f"stderr:\n{stderr}",
                                  extra={'func_name': f'sudo python3 {self.script_path} {self.command}'})

            while self._is_running:
                line = stdout.readline()
                self.logger3.info(f"{self.script_path} {self.command} {line}")
                if not line:
                    break
                if self._is_running == False:
                    break
                self.output_ready.emit(line.strip())


        except Exception as e:
            self.error_occurred.emit(f"Error during execution: {str(e)}")
        finally:
            self.cleanup()
            self.finished_signal.emit()

    def stop(self):
        self._is_running = False
        self.cleanup()

        # if self.isRunning():
        #    self.terminate()
        #    self.wait(2000)

    def cleanup(self):
        if hasattr(self.ssh_handler, 'SSH_disconnect'):
            self.ssh_handler.SSH_disconnect()


class SoemCompileWorker(QThread):
    """Background thread that runs the SOEM compile command over an *already-
    established* SSH connection, streams each output line in real time via
    ``output_ready``, and emits the complete stdout/stderr via ``compile_done``
    when the remote process finishes.  The SSH connection is intentionally
    *not* closed here so the caller can continue with subsequent commands."""

    # Each output line as it arrives
    output_ready = pyqtSignal(str)
    # Full accumulated stdout and stderr when the process ends
    compile_done = pyqtSignal(str, str)
    # Emitted when an unexpected exception occurs
    error_occurred = pyqtSignal(str)

    def __init__(self, ssh_handler):
        super().__init__()
        self.ssh_handler = ssh_handler
        self._is_running = True
        self._logger = logger.getChild('SoemCompileWorker')

    def run(self):
        try:
            script_path = self.ssh_handler.script_path
            stdin, stdout_ch, stderr_ch = self.ssh_handler.ssh.exec_command(
                f'sudo python3 {script_path} soemcompile',
                get_pty=True,
            )
            self._logger.info("soemcompile started in background thread",
                              extra={'func_name': 'soemcompile'})

            stdout_lines = []
            while self._is_running:
                line = stdout_ch.readline()
                if not line:
                    break
                stripped = line.strip()
                stdout_lines.append(stripped)
                self._logger.debug(stripped, extra={'func_name': 'soemcompile'})
                self.output_ready.emit(stripped)

            stderr_data = ""
            try:
                stderr_data = stderr_ch.read().decode(errors='replace')
            except Exception:
                pass
            if stderr_data:
                self._logger.error(f"soemcompile stderr:\n{stderr_data}",
                                   extra={'func_name': 'soemcompile'})

            self.compile_done.emit('\n'.join(stdout_lines), stderr_data)

        except Exception as exc:
            self._logger.error(f"soemcompile thread error: {str(exc)}", exc_info=True,
                               extra={'func_name': 'soemcompile'})
            self.error_occurred.emit(f"SOEM compile error: {str(exc)}")

    def stop(self):
        self._is_running = False


class SshConsoleWorker(QThread):
    """Worker thread that opens an interactive SSH shell and streams output."""
    output_ready = pyqtSignal(str)
    connected = pyqtSignal()
    disconnected = pyqtSignal()
    error_occurred = pyqtSignal(str)

    def __init__(self, host, port, username, password):
        super().__init__()
        self.host = host
        self.port = port
        self.username = username
        self.password = password
        self._is_running = False
        self._channel = None
        self._channel_lock = __import__('threading').Lock()
        self._ssh = None

    def run(self):
        self._is_running = True
        try:
            self._ssh = paramiko.SSHClient()
            self._ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            self._ssh.connect(
                self.host, self.port, self.username, self.password, timeout=10
            )
            with self._channel_lock:
                self._channel = self._ssh.invoke_shell(width=220, height=50)
                self._channel.settimeout(0.2)
            self.connected.emit()
            while self._is_running:
                try:
                    with self._channel_lock:
                        if self._channel is None or self._channel.closed:
                            break
                        data = self._channel.recv(4096)
                    if data:
                        self.output_ready.emit(data.decode('utf-8', errors='replace'))
                    elif not data:
                        break
                except socket.timeout:
                    pass
                except Exception as exc:
                    logger.debug(f"SshConsoleWorker recv: {exc}")
        except Exception as e:
            self.error_occurred.emit(str(e))
        finally:
            self._cleanup()
            self.disconnected.emit()

    def send_command(self, cmd):
        with self._channel_lock:
            if self._channel and not self._channel.closed:
                self._channel.send(cmd)

    def stop(self):
        self._is_running = False

    def _cleanup(self):
        try:
            with self._channel_lock:
                if self._channel:
                    self._channel.close()
                    self._channel = None
            if self._ssh:
                self._ssh.close()
                self._ssh = None
        except Exception as exc:
            logger.debug(f"SshConsoleWorker cleanup: {exc}")


class ScpWorker(QThread):
    """Worker thread for SFTP file upload / download (SCP-like)."""
    progress = pyqtSignal(str)   # status / progress messages
    finished = pyqtSignal(bool, str)  # success, message

    def __init__(self, host, port, username, password,
                 direction, local_path, remote_path):
        super().__init__()
        self.host = host
        self.port = port
        self.username = username
        self.password = password
        # direction: 'upload'  → local → remote
        #            'download' → remote → local
        self.direction = direction
        self.local_path = local_path
        self.remote_path = remote_path

    def run(self):
        transport = None
        try:
            transport = paramiko.Transport((self.host, self.port))
            transport.connect(username=self.username, password=self.password)
            sftp = paramiko.SFTPClient.from_transport(transport)

            if self.direction == 'upload':
                self.progress.emit(
                    f"[SCP] Uploading  {self.local_path}  →  {self.remote_path} …"
                )
                sftp.put(self.local_path, self.remote_path,
                         callback=self._sftp_callback)
                sftp.close()
                self.finished.emit(True, f"[SCP] Upload complete: {self.remote_path}")
            else:
                self.progress.emit(
                    f"[SCP] Downloading  {self.remote_path}  →  {self.local_path} …"
                )
                sftp.get(self.remote_path, self.local_path,
                         callback=self._sftp_callback)
                sftp.close()
                self.finished.emit(True, f"[SCP] Download complete: {self.local_path}")
        except Exception as exc:
            self.finished.emit(False, f"[SCP] Error: {exc}")
        finally:
            if transport:
                transport.close()

    def _sftp_callback(self, transferred, total):
        if total > 0:
            pct = int(transferred * 100 / total)
            self.progress.emit(f"\r[SCP] {pct}%  ({transferred}/{total} bytes)")


class RemoteFileBrowserDialog(QDialog):
    """
    A modal dialog that browses the RPI filesystem over SFTP.

    mode='file'  – user must select a remote *file*   (used for Download)
    mode='dir'   – user selects a remote *directory*  (used for Upload destination)

    After exec_() == QDialog.Accepted  →  self.selected_path holds the chosen path.
    """

    def __init__(self, host, port, username, password,
                 mode='file', start_path=None, parent=None):
        super().__init__(parent)
        self.host = host
        self.port = port
        self.username = username
        self.password = password
        self.mode = mode          # 'file' or 'dir'
        self.selected_path = ''
        self._sftp = None
        self._transport = None
        # Use a safe default start path, falling back to root if no username
        safe_user = username.strip('/') if username else 'home'
        self._current_path = start_path or f'/home/{safe_user}'

        self.setWindowTitle(
            "Browse RPI – Select File" if mode == 'file'
            else "Browse RPI – Select Destination Folder"
        )
        self.resize(600, 420)
        self._build_ui()
        self._connect_sftp()
        self._load_dir(self._current_path)

    # ------------------------------------------------------------------
    # UI
    # ------------------------------------------------------------------
    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(6)

        # ── path bar ──────────────────────────────────────────────────
        path_row = QHBoxLayout()
        up_btn = QPushButton("⬆ Up")
        up_btn.setFixedWidth(60)
        up_btn.clicked.connect(self._go_up)
        path_row.addWidget(up_btn)

        self._path_edit = QLineEdit(self._current_path)
        self._path_edit.returnPressed.connect(self._go_to_typed_path)
        path_row.addWidget(self._path_edit)

        go_btn = QPushButton("Go")
        go_btn.setFixedWidth(40)
        go_btn.clicked.connect(self._go_to_typed_path)
        path_row.addWidget(go_btn)

        layout.addLayout(path_row)

        # ── status label (shows connection state / errors) ────────────
        self._status_lbl = QLabel("Connecting…")
        self._status_lbl.setStyleSheet("color: #888; font-size: 8pt;")
        layout.addWidget(self._status_lbl)

        # ── file list ─────────────────────────────────────────────────
        self._list = QListWidget()
        self._list.setAlternatingRowColors(True)
        self._list.itemDoubleClicked.connect(self._on_double_click)
        self._list.itemClicked.connect(self._on_single_click)
        layout.addWidget(self._list, stretch=1)

        # ── selection label ───────────────────────────────────────────
        self._sel_lbl = QLabel("Nothing selected")
        self._sel_lbl.setStyleSheet("color: #555; font-style: italic;")
        layout.addWidget(self._sel_lbl)

        # ── buttons ───────────────────────────────────────────────────
        btn_row = QHBoxLayout()

        if self.mode == 'dir':
            self._select_folder_btn = QPushButton("📂  Select This Folder")
            self._select_folder_btn.setStyleSheet(
                "background-color: #17a2b8; color: white; padding: 4px 10px;"
            )
            self._select_folder_btn.clicked.connect(self._select_current_folder)
            btn_row.addWidget(self._select_folder_btn)

        btn_row.addStretch()
        box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self._ok_btn = box.button(QDialogButtonBox.Ok)
        self._ok_btn.setText("Select")
        self._ok_btn.setEnabled(False)
        box.accepted.connect(self._on_accept)
        box.rejected.connect(self.reject)
        btn_row.addWidget(box)
        layout.addLayout(btn_row)

    # ------------------------------------------------------------------
    # SFTP connection
    # ------------------------------------------------------------------
    def _connect_sftp(self):
        try:
            self._transport = paramiko.Transport((self.host, self.port))
            self._transport.connect(username=self.username, password=self.password)
            self._sftp = paramiko.SFTPClient.from_transport(self._transport)
            self._status_lbl.setText(f"Connected  |  {self.host}")
        except Exception as exc:
            self._status_lbl.setText(f"SFTP error: {exc}")
            self._list.addItem(QListWidgetItem(f"⚠  Cannot connect: {exc}"))

    def closeEvent(self, event):
        self._close_sftp()
        super().closeEvent(event)

    def _close_sftp(self):
        try:
            if self._sftp:
                self._sftp.close()
                self._sftp = None
        except paramiko.SSHException as exc:
            logger.debug(f"RemoteFileBrowserDialog SFTP close: {exc}")
        try:
            if self._transport:
                self._transport.close()
                self._transport = None
        except paramiko.SSHException as exc:
            logger.debug(f"RemoteFileBrowserDialog transport close: {exc}")

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    @staticmethod
    def _remote_join(directory, name):
        """Join a remote (POSIX) directory path with an entry name."""
        import posixpath
        return posixpath.join(directory, name)

    @staticmethod
    def _remote_parent(path):
        """Return the parent directory of a remote (POSIX) path."""
        import posixpath
        parent = posixpath.dirname(path.rstrip('/'))
        return parent or '/'

    # ------------------------------------------------------------------
    # Directory loading
    # ------------------------------------------------------------------
    def _load_dir(self, path):
        if not self._sftp:
            return
        try:
            entries = self._sftp.listdir_attr(path)
        except Exception as exc:
            self._status_lbl.setText(f"Cannot list {path}: {exc}")
            return

        self._current_path = path
        self._path_edit.setText(path)
        self._list.clear()
        self._sel_lbl.setText("Nothing selected")
        self._ok_btn.setEnabled(False)

        # Sort: folders first, then files, both alphabetically
        dirs = sorted(
            [e for e in entries if stat.S_ISDIR(e.st_mode)],
            key=lambda e: e.filename.lower()
        )
        files = sorted(
            [e for e in entries if not stat.S_ISDIR(e.st_mode)],
            key=lambda e: e.filename.lower()
        )

        for entry in dirs:
            item = QListWidgetItem(f"📁  {entry.filename}")
            item.setData(Qt.UserRole,
                         ('dir', self._remote_join(path, entry.filename)))
            self._list.addItem(item)

        for entry in files:
            item = QListWidgetItem(f"📄  {entry.filename}")
            item.setData(Qt.UserRole,
                         ('file', self._remote_join(path, entry.filename)))
            self._list.addItem(item)

        self._status_lbl.setText(
            f"{path}  –  {len(dirs)} folder(s), {len(files)} file(s)"
        )

        # In 'dir' mode the current folder is always a valid selection
        if self.mode == 'dir':
            self._ok_btn.setEnabled(True)
            self._sel_lbl.setText(f"Destination: {path}")
            self.selected_path = path

    # ------------------------------------------------------------------
    # Navigation slots
    # ------------------------------------------------------------------
    def _go_up(self):
        self._load_dir(self._remote_parent(self._current_path))

    def _go_to_typed_path(self):
        path = self._path_edit.text().strip()
        if path:
            self._load_dir(path)

    def _on_double_click(self, item):
        kind, path = item.data(Qt.UserRole)
        if kind == 'dir':
            self._load_dir(path)

    def _on_single_click(self, item):
        kind, path = item.data(Qt.UserRole)
        if self.mode == 'file':
            if kind == 'file':
                self.selected_path = path
                self._sel_lbl.setText(f"Selected: {path}")
                self._ok_btn.setEnabled(True)
            else:
                self._ok_btn.setEnabled(False)
                self._sel_lbl.setText("Double-click a folder to navigate into it")
        else:  # dir mode – single-click on a subfolder selects it
            if kind == 'dir':
                self.selected_path = path
                self._sel_lbl.setText(f"Destination: {path}")
                self._ok_btn.setEnabled(True)

    def _select_current_folder(self):
        """'Select This Folder' button – confirm the current directory."""
        self.selected_path = self._current_path
        self._close_sftp()
        self.accept()

    # ------------------------------------------------------------------
    # Accept / reject
    # ------------------------------------------------------------------
    def _on_accept(self):
        if self.selected_path:
            self._close_sftp()
            self.accept()


class TerminalWidget(QPlainTextEdit):
    """
    A terminal-emulator widget that behaves like PuTTY / MobaXterm:
    - Every keystroke is forwarded immediately to the SSH channel.
    - Arrow keys, Ctrl+C/D/Z, Tab, Backspace, F-keys all work.
    - Received text is displayed with basic control-character handling
      (\r overwrite mode, \n newline, backspace echo).
    - Right-click → "Paste to terminal" to send clipboard text.
    """

    # Maps Qt key codes to ANSI/VT100 escape sequences
    _KEY_MAP = {
        Qt.Key_Up:       '\x1b[A',
        Qt.Key_Down:     '\x1b[B',
        Qt.Key_Right:    '\x1b[C',
        Qt.Key_Left:     '\x1b[D',
        Qt.Key_Home:     '\x1b[H',
        Qt.Key_End:      '\x1b[F',
        Qt.Key_Delete:   '\x1b[3~',
        Qt.Key_PageUp:   '\x1b[5~',
        Qt.Key_PageDown: '\x1b[6~',
        Qt.Key_F1:       '\x1bOP',
        Qt.Key_F2:       '\x1bOQ',
        Qt.Key_F3:       '\x1bOR',
        Qt.Key_F4:       '\x1bOS',
        Qt.Key_F5:       '\x1b[15~',
        Qt.Key_F6:       '\x1b[17~',
        Qt.Key_F7:       '\x1b[18~',
        Qt.Key_F8:       '\x1b[19~',
        Qt.Key_F9:       '\x1b[20~',
        Qt.Key_F10:      '\x1b[21~',
        Qt.Key_F11:      '\x1b[23~',
        Qt.Key_F12:      '\x1b[24~',
    }

    # Strip ANSI colour/style/cursor-movement escape sequences from output
    _ANSI_STRIP = re.compile(
        r'\x1b(?:'
        r'[@-Z\\-_]'            # two-byte ESC sequences
        r'|\[[0-?]*[ -/]*[@-~]' # CSI sequences  e.g. \x1b[1;32m
        r'|\][^\x07\x1b]*(?:\x07|\x1b\\)'  # OSC sequences
        r'|[\(\)][A-Z0-9=]'     # character-set designators
        r')'
    )

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setReadOnly(True)
        self._send_fn = None
        self.setFocusPolicy(Qt.StrongFocus)
        self.setLineWrapMode(QPlainTextEdit.NoWrap)
        self.setStyleSheet("""
            QPlainTextEdit {
                background-color: #1e1e1e;
                color: #d4d4d4;
                font-family: 'Courier New', monospace;
                font-size: 11pt;
                border: 1px solid #444;
                border-radius: 4px;
                padding: 4px;
            }
        """)

    def set_send_fn(self, fn):
        """Set (or clear) the function used to send keystrokes to the SSH channel."""
        self._send_fn = fn

    # ------------------------------------------------------------------
    # Keyboard handling
    # ------------------------------------------------------------------
    def keyPressEvent(self, event):
        if self._send_fn is None:
            # Not connected – only allow scrolling shortcuts
            super().keyPressEvent(event)
            return

        key = event.key()
        mods = event.modifiers()
        text = event.text()

        # --- Ctrl+<letter> ---
        if mods & Qt.ControlModifier and not (mods & Qt.ShiftModifier):
            ctrl_map = {
                Qt.Key_C: '\x03',  # interrupt
                Qt.Key_D: '\x04',  # EOF
                Qt.Key_Z: '\x1a',  # suspend
                Qt.Key_L: '\x0c',  # clear screen
                Qt.Key_A: '\x01',  # beginning of line
                Qt.Key_E: '\x05',  # end of line
                Qt.Key_U: '\x15',  # kill to start of line
                Qt.Key_K: '\x0b',  # kill to end of line (VT / readline ^K)
                Qt.Key_W: '\x17',  # delete word back
                Qt.Key_R: '\x12',  # reverse history search
            }
            if key in ctrl_map:
                self._send_fn(ctrl_map[key])
                return
            # Ctrl+C with selection → copy to clipboard (allow default)
            if key == Qt.Key_C and self.textCursor().hasSelection():
                super().keyPressEvent(event)
                return
            return  # absorb other Ctrl combos

        # --- Special / navigation keys ---
        if key in self._KEY_MAP:
            self._send_fn(self._KEY_MAP[key])
            return

        if key in (Qt.Key_Return, Qt.Key_Enter):
            self._send_fn('\r')
            return

        if key == Qt.Key_Backspace:
            self._send_fn('\x7f')
            return

        if key == Qt.Key_Tab:
            self._send_fn('\t')
            return

        if key == Qt.Key_Escape:
            self._send_fn('\x1b')
            return

        # --- Printable character ---
        if text:
            self._send_fn(text)
            return

        # --- Modifier-only keys (Shift, Alt, …) – let Qt handle (no text change) ---
        super().keyPressEvent(event)

    # ------------------------------------------------------------------
    # Output display
    # ------------------------------------------------------------------
    def write(self, text):
        """
        Process and display text received from the SSH channel.
        Handles carriage return (overwrite current line), newline, backspace echo,
        and strips ANSI escape sequences.
        """
        text = self._ANSI_STRIP.sub('', text)

        doc = self.document()
        cur = QTextCursor(doc)
        cur.movePosition(QTextCursor.End)

        i = 0
        while i < len(text):
            ch = text[i]
            if ch == '\r':
                cur.movePosition(QTextCursor.StartOfBlock)
            elif ch == '\n':
                cur.movePosition(QTextCursor.End)
                cur.insertText('\n')
            elif ch in ('\x08', '\x7f'):
                # Backspace / DEL echo from remote
                if not cur.atBlockStart():
                    cur.deletePreviousChar()
            elif ch == '\x07':
                pass  # bell – ignore
            elif ord(ch) >= 32 or ch == '\t':
                # Overwrite mode: replace character under cursor if not at line end
                if not cur.atBlockEnd():
                    cur.deleteChar()
                cur.insertText(ch)
            i += 1

        self.setTextCursor(cur)
        sb = self.verticalScrollBar()
        sb.setValue(sb.maximum())

    # ------------------------------------------------------------------
    # Context menu: paste to terminal
    # ------------------------------------------------------------------
    def contextMenuEvent(self, event):
        menu = self.createStandardContextMenu()
        menu.addSeparator()
        paste_action = menu.addAction("Paste to terminal")
        paste_action.setEnabled(self._send_fn is not None)
        paste_action.triggered.connect(self._paste_to_terminal)
        menu.exec_(event.globalPos())

    def _paste_to_terminal(self):
        text = QApplication.clipboard().text()
        if text and self._send_fn:
            self._send_fn(text)


class AssemblySuffixErrorDialog(QDialog):
    """
    Animated error dialog shown when the assembly suffix is not set.
    The dialog border pulses red to draw attention, then stays solid red until
    the user dismisses it.  The test is expected to be stopped by the caller
    before this dialog is displayed.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Configuration Error — Test Stopped")
        self.setModal(True)
        self.setFixedWidth(480)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)

        # ── Base stylesheet (border colour is overridden by animation) ──────
        self._base_style = (
            "AssemblySuffixErrorDialog {"
            "  background-color: #1a0a0a;"
            "  border: 3px solid {border};"
            "  border-radius: 10px;"
            "}"
        )
        self._apply_border("#ff4444")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 20, 24, 20)
        layout.setSpacing(14)

        # ── Icon row ─────────────────────────────────────────────────────────
        icon_label = QLabel("⛔")
        icon_font = QFont("Segoe UI Emoji", 36)
        icon_label.setFont(icon_font)
        icon_label.setAlignment(Qt.AlignHCenter)
        icon_label.setStyleSheet("color: #ff4444; background: transparent;")
        layout.addWidget(icon_label)

        # ── Title ─────────────────────────────────────────────────────────────
        title = QLabel("Assembly Suffix Not Set — Test Stopped")
        title.setFont(QFont("Arial", 12, QFont.Bold))
        title.setAlignment(Qt.AlignHCenter)
        title.setWordWrap(True)
        title.setStyleSheet("color: #ff6666; background: transparent;")
        layout.addWidget(title)

        # ── Body text ─────────────────────────────────────────────────────────
        body = QLabel(
            "The application cannot determine the configuration path because the "
            "<b>Assembly Part Number</b> has not been entered or is not recognised.<br><br>"
            "The current test has been <b>stopped automatically</b>.<br><br>"
            "Please go to the <b>Unit&nbsp;Setup</b> tab, fill in all required fields "
            "(Assembly Part Number, Serial Number, Revision, etc.), "
            "and restart the test."
        )
        body.setFont(QFont("Arial", 9))
        body.setWordWrap(True)
        body.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        body.setStyleSheet("color: #dddddd; background: transparent;")
        layout.addWidget(body)

        # ── OK button ─────────────────────────────────────────────────────────
        ok_btn = QPushButton("OK — Go to Unit Setup")
        ok_btn.setFont(QFont("Arial", 9, QFont.Bold))
        ok_btn.setFixedHeight(34)
        ok_btn.setStyleSheet(
            "QPushButton {"
            "  background-color: #8b0000;"
            "  color: white;"
            "  border: 1px solid #ff4444;"
            "  border-radius: 5px;"
            "  padding: 4px 16px;"
            "}"
            "QPushButton:hover { background-color: #b00000; }"
            "QPushButton:pressed { background-color: #600000; }"
        )
        ok_btn.clicked.connect(self.accept)
        layout.addWidget(ok_btn, alignment=Qt.AlignHCenter)

        # ── Pulse animation: cycle border between bright-red and dark-red ─────
        self._pulse_timer = QTimer(self)
        self._pulse_timer.setInterval(500)   # 0.5 s per half-cycle
        self._pulse_state = True
        self._pulse_timer.timeout.connect(self._tick_pulse)
        self._pulse_timer.start()

        # Stop animation after 3 s (6 ticks) so dialog settles on solid red
        QTimer.singleShot(3000, self._pulse_timer.stop)

    # ─────────────────────────────────────────────────────────────────────────
    def _apply_border(self, colour: str):
        self.setStyleSheet(self._base_style.replace("{border}", colour))

    def _tick_pulse(self):
        self._pulse_state = not self._pulse_state
        self._apply_border("#ff0000" if self._pulse_state else "#3a0000")


class TestStationInterface(QMainWindow):
    # No-output idle watchdog timeout (milliseconds).  Change only here; the
    # error dialog text is derived from this value automatically.
    _IDLE_TIMEOUT_MS: int = 120_000  # 2 minutes

    def __init__(self):
        super().__init__()
        self.excel_logger = excel_logger
        self.setWindowTitle("Test Station Interface")
        self.config = ''
        self.assembly_suffix = None

        # Set minimum size instead of fixed geometry
        self.setMinimumSize(1000, 700)

        # Central widget with layout
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.main_layout = QVBoxLayout(self.central_widget)
        self.main_layout.setContentsMargins(8, 8, 8, 8)
        self.main_layout.setSpacing(8)
        self.Zone1_Inner_res = 0
        self.Zone2_Mid_Inner_res = 0
        self.Zone3_Mid_Edge_res = 0
        self.Zone4_Edge_res = 0
        self.Zone5_Outer_res = 0
        self.Zone1_Inner_imp = 0
        self.Zone2_Mid_Inner_imp = 0
        self.Zone3_Mid_Edge_imp = 0
        self.Zone4_Edge_imp = 0
        self.Zone5_Outer_imp = 0
        #self.config_transfer()

        # Initialize other components
        self.ssh_handler = SSH_setup()
        self.console_output = None
        self.check = False
        self.handling_flag = 0
        self.Firmware_check = True
        self.worker = None
        self.worker_thread = None
        self.open_count = 0
        self.closed_count = 0
        self.logger = logger.getChild('TestStationInterface')
        self.check_true = 0
        self.vna_timer = QTimer()
        self.vna_timer.timeout.connect(self.update_vna_progress)
        self.bnc_idle_timer = QTimer()
        self.bnc_idle_timer.setSingleShot(True)
        self.bnc_idle_timer.timeout.connect(self._on_bnc_idle_timeout)
        self.vna_idle_timer = QTimer()
        self.vna_idle_timer.setSingleShot(True)
        self.vna_idle_timer.timeout.connect(self._on_vna_idle_timeout)
        self.imp_idle_timer = QTimer()
        self.imp_idle_timer.setSingleShot(True)
        self.imp_idle_timer.timeout.connect(self._on_imp_idle_timeout)
        self.res_idle_timer = QTimer()
        self.res_idle_timer.setSingleShot(True)
        self.res_idle_timer.timeout.connect(self._on_res_idle_timeout)
        self.interlock_idle_timer = QTimer()
        self.interlock_idle_timer.setSingleShot(True)
        self.interlock_idle_timer.timeout.connect(self._on_interlock_idle_timeout)
        self.vna_progress_value = 0
        self.names = ''
        self.unit_test = 0
        self.impedance_scan = 0
        self.self_t = 0
        self.Res_scan = 0
        self.bnc_t = 0
        self.VNA_c = 0
        self.interlock_t = 0
        self.step_no = 0
        self.resistance_test = 'PASS'
        self.Impedance_test = 'PASS'
        self.over_all_result = 'PASS'
        self.test_result = 'PASS'
        self.PN = ''
        self.SN = ''
        self.init_ui()
        self.stop_increment = False
        self._soem_thread_started = False  # guards finally in auto_load_connect
        self._soem_loading_timer = QTimer()
        self._soem_loading_timer.timeout.connect(self._update_soem_loading_line)
        self._soem_dot_count = 0
        self._self_test_worker = None   # Worker thread for the Self Test tab
        self._self_test_lines = []      # accumulates streamed lines until done

    def closeEvent(self, event):
        self.cleanup_resources()
        event.accept()

    """      
    def config_transfer(self):
        #host = "192.168.1.2"  # Replace with your Pi's IP
        host = "10.119.28.136"
        username = "robot"
        password = "robot"  # Default, change if needed

        # File paths
        local_file = "C:\\Config\\config.json"  # Windows path
        remote_file = "/home/robot/Manufacturing_test/aipc_beta/config.json"  # Destination on Pi
        local_file1 = "C:\\Config\\zone_config.cfg"  # Windows path
        remote_file1 = "/home/robot/Manufacturing_test/aipc_beta/zone_config.cfg"  # Destination on Pi

        # Initialize SSH client
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(host, username=username, password=password)

        # Transfer file using SFTP
        sftp = ssh.open_sftp()
        sftp.put(local_file, remote_file)
        sftp.put(local_file1, remote_file1)
        sftp.close()
        ssh.close()
    """

    def config_transfer(self, suffix):
        """
        Transfer config files to RPI based on assembly part number suffix.
        suffix: '003' or '004' - determines which config directory to use
        """
        # host = "192.168.1.2"  # Replace with your Pi's IP
        host = "10.119.9.225"
        username = "robot"
        password = "robot"  # Default, change if needed

        # Determine config directory based on suffix
        if suffix is None:
            config_dir = "C:\\Config"
        else:
            config_dir = f"C:\\Config\\{suffix}"

        # File paths
        local_file = f"{config_dir}\\config.json"  # Windows path
        remote_file = "/home/robot/Manufacturing_test/aipc_beta/config.json"  # Destination on Pi
        local_file1 = f"{config_dir}\\zone_config.cfg"  # Windows path
        remote_file1 = "/home/robot/Manufacturing_test/aipc_beta/zone_config.cfg"  # Destination on Pi

        if not os.path.exists(local_file):
            error_msg = f"Configuration file not found: {local_file}"
            self.append_console_message(error_msg + "\n", is_error=True)
            self.logger.error(error_msg, extra={'func_name': 'config_transfer'})
            QMessageBox.critical(self, "ERROR : ", error_msg)
            return False


        if not os.path.exists(local_file1):
            error_msg = f"Zone configuration file not found: {local_file1}"
            self.append_console_message(error_msg + "\n", is_error=True)
            self.logger.error(error_msg, extra={'func_name': 'config_transfer'})
            QMessageBox.critical(self, "ERROR : ", error_msg)
            return False


        # Initialize SSH client
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(host, username=username, password=password)

        # Transfer file using SFTP
        sftp = ssh.open_sftp()
        sftp.put(local_file, remote_file)
        sftp.put(local_file1, remote_file1)
        sftp.close()
        ssh.close()
        return True

    def _extract_assembly_suffix(self, assembly_part_number):
        """
        Extract the suffix (003 or 004) from assembly part number.
        Format: XXX-AXXXXX-003 or XXX-AXXXXX-004
        Returns: '003', '004', or None if invalid format
        """
        parts = assembly_part_number.split('-')
        if len(parts) == 3:
            suffix = parts[2]
            if suffix in ['003', '004', '005']:
                return suffix
        return None

    def load_config(self, suffix):
        try:
            if suffix is None:
                raise ValueError("Assembly suffix is not set; cannot determine config path.")
            with open(rf'C:\Config\{suffix}\config.json') as f:
                return json.load(f)
        except ValueError as e:
            self.logger.error(f"Error loading config: {str(e)}", exc_info=True,
                               extra={'func_name': 'load_config'})
            # Show animated error dialog and stop the test
            dlg = AssemblySuffixErrorDialog(self)
            dlg.exec_()
            # Navigate the user directly to the Unit Setup tab (index 0)
            self.tab_widget.setCurrentIndex(0)
            # Return None so callers can detect the failure and abort
            return None
        except Exception as e:
            self.logger.error(f"Error loading config: {str(e)}", exc_info=True,
                               extra={'func_name': 'load_config'})

            return {"expected_firmware_version": "0.0.0"}

    def create_otp_file(self, fname: str,
                        app_pn: str, app_sn: str,
                        assy_pn: str, assy_sn: str) -> None:
        """
        Creates an OTP (One-Time Program) file with board and assembly information.

        Args:
            fname: File name/path to create
            app_pn: Application board part number
            app_sn: Application board serial number
            assy_pn: Assembly part number
            assy_sn: Assembly serial number

        Raises:
            IOError: If file cannot be written
        """
        try:
            header = "# LAM Research App Board and Assy Part and Serial Number"
            content = [
                header,
                f"Board PN: {app_pn}",
                f"Board SN: {app_sn}",
                f"Assy PN: {assy_pn}",
                f"Assy SN: {assy_sn}",
                ""  # Adds final newline
            ]
            self.append_console_message("Write APPOTP FILE")
            with open(fname, 'w', encoding='utf-8') as f:
                f.write('\n'.join(content))

        except IOError as e:
            self.append_console_message("Failed to create OTP file ",is_error= True)
            raise IOError(f"Failed to create OTP file {fname}") from e

    def start_vna_progress(self):
        """Start the 12-second progress timer"""
        self.vna_progress_value = 0
        self.vna_progress.setValue(0)
        self.vna_timer.start(1800)  # 1800ms interval for 180 seconds (100*1800ms=180s)

    def update_vna_progress(self):
        """Update progress bar incrementally"""
        self.vna_progress_value += 1
        self.vna_progress.setValue(self.vna_progress_value)

        if self.vna_progress_value >= 100:
            self.vna_timer.stop()
            self.vna_progress.setValue(100)

    def validate_part_number(self, part_number, part_type):
        pattern = r"^\d{3}-[A-Z]\d{5}-\d{3}$"
        if not re.match(pattern, part_number):
            QMessageBox.critical(
                self,
                "Validation Error",
                f"Invalid {part_type} number format!\n\n"
                f"Entered: {part_number}\n"
                "Expected format: XXX-AXXXXX-XXX\n"
                "(3 digits, hyphen, uppercase letter, 5 digits, hyphen, 3 digits)"
            )
            return False
        return True

    def validate_revision_number(self, serial_number, part_type):
        pattern = r"^[a-zA-Z]$"
        if not re.match(pattern, serial_number):
            QMessageBox.critical(
                self,
                "Validation Error",
                f"Invalid {serial_number} number format!\n\n"
                f"Entered: {serial_number}\n"
                "Expected format: [A-Z]or[a-z]\n"
            )
            return False
        return True

    def parse_ssh_output(self, output):
        results = {
            'product_id': 'Not found',
            'esi_revision': 'Not found',
            'firmware_version': 'Not found',
            'ethercat_address': 'Not found',
            'pcb_part_number': 'Not found',
            'pcb_serial_number': 'Not found',
            'assembly_part_number': 'Not found',
            'assembly_serial_number': 'Not found'
        }

        # Extract standard fields
        product_code_match = re.search(r'Product Code: ([^\n]+)', output, re.DOTALL)
        if product_code_match:
            results['product_id'] = product_code_match.group(1).strip()

        revision_match = re.search(r'Revision: ([^\n]+)', output, re.DOTALL)
        if revision_match:
            results['esi_revision'] = revision_match.group(1).strip()

        ECAT_add = re.search(r'ECAT Address: ([^\n]+)', output, re.DOTALL)
        if ECAT_add:
            results['ethercat_address'] = ECAT_add.group(1).strip()

        version_match = re.search(r'Software version: ([^\n]+)', output)
        if version_match:
            results['firmware_version'] = version_match.group(1).strip()

        # Extract OTP programmed values
        pcb_pn_match = re.search(r'PCB_Part_Number:([^\s]+)', output)
        if pcb_pn_match:
            results['pcb_part_number'] = pcb_pn_match.group(1).strip()

        pcb_sn_match = re.search(r'PCB_Serial_Number:([^\s]+)', output)
        if pcb_sn_match:
            results['pcb_serial_number'] = pcb_sn_match.group(1).strip()

        assy_pn_match = re.search(r'Assembly_Part_Number:([^\s]+)', output)
        if assy_pn_match:
            results['assembly_part_number'] = assy_pn_match.group(1).strip()

        assy_sn_match = re.search(r'Assembly_Serial_Number:([^\s]+)', output)
        if assy_sn_match:
            results['assembly_serial_number'] = assy_sn_match.group(1).strip()

        return results

    def append_console_message(self, message, is_error=False):
        """Helper method to append colored messages to console"""
        if hasattr(self, 'console_output') and self.console_output is not None:
            if is_error:
                self.console_output.append(f'<span style="color:#f85149; font-weight:bold;">{message}</span>')
            else:
                self.console_output.append(f'<span style="color:#3fb950; font-weight:bold;">{message}</span>')
            # Auto-scroll to bottom
            self.console_output.verticalScrollBar().setValue(
                self.console_output.verticalScrollBar().maximum()
            )

    def handle_ssh_error(self, error_msg):
        if hasattr(self, 'console_output') and self.console_output is not None:
            self.append_console_message(f"!!! SSH ERROR !!!\n{error_msg}\n", is_error=True)

        QMessageBox.critical(self, "SSH Error", error_msg)

    def VNA_cal_test(self):
        try:

            if hasattr(self, 'worker') and self.worker:
                self.worker.stop()

            self.vna_t = time.time()
            # Reset UI state
            self.VNAtest_console.clear()
            self.VNA_status_label_start.setText("● Running…")
            self.VNA_status_label_start.setStyleSheet(self._PILL_RUN_SS)
            self.start_vna_progress()
            self.VNA_start_button.setEnabled(False)

            self.append_vna_message("\n================== VNA CAL Test Started =======================\n")

            self.worker = Worker(
                self.ssh_handler,
                '/home/robot/Manufacturing_test/aipc_beta/vnacalibration.py',
                f'dimm 50000 70000',
                350
            )
            self.worker.output_ready.connect(self.handle_vna_output)
            self.worker.finished_signal.connect(self.on_vna_test_finished)
            self.worker.error_occurred.connect(self.handle_vna_error)

            # Start the thread
            self.worker.start()
            self._set_tabs_locked(True)
            self.vna_idle_timer.start(self._IDLE_TIMEOUT_MS)

        except Exception as e:
            self.append_vna_message(f"Failed to start test: {str(e)}", is_error=True)
            self.cleanup_resources()

    def handle_vna_output(self, line):
        try:
            # Reset the 5-minute idle watchdog on every received line
            self.vna_idle_timer.start(self._IDLE_TIMEOUT_MS)
            self.append_vna_message(f"{line}\n")
            if "no ping" in line:
                self.append_vna_message(f"VNA not connected to the Network", is_error=True)
                self.vna_idle_timer.stop()
                self.worker.stop()
                self.VNA_start_button.setEnabled(True)

            if "mailbox error" in line.lower():
                self.append_vna_message(
                    "Mailbox Error on Ethercat please check the Ethercat Data or Contact Support Team",
                    is_error=True
                )
                self.vna_idle_timer.stop()
                self.worker.stop()
                self.VNA_start_button.setEnabled(True)
                return

            if "Calibration PASS" in line:
                self.VNA_status_label_start.setText('● Completed — PASS')
                self.VNA_status_label_start.setStyleSheet(self._PILL_PASS_SS)
                self.append_vna_message("\n=== VNA Calibration PASSED ===")
                self.vna_timer.stop()
                self.vna_idle_timer.stop()
                self.worker.stop()
                self.VNA_start_button.setEnabled(True)
                # self.work_timeout = 30

            elif "Calibration FAIL" in line :
                self.VNA_status_label_start.setText('● Completed — FAIL')
                self.VNA_status_label_start.setStyleSheet(self._PILL_FAIL_SS)
                self.append_vna_message("\n!!! VNA Calibration FAILED !!!", is_error=True)

                self.vna_timer.stop()
                self.vna_idle_timer.stop()
                self.worker.stop()
                self.VNA_start_button.setEnabled(True)
                # self.work_timeout = 30
            elif "ERROR: Connect ECal module" in line :
                self.VNA_status_label_start.setText('● Completed — FAIL')
                self.VNA_status_label_start.setStyleSheet(self._PILL_FAIL_SS)
                self.append_vna_message("\n!!! VNA Calibration FAILED : Please connect Ecal Module... !!!", is_error=True)
                self.vna_timer.stop()
                self.vna_idle_timer.stop()
                self.worker.stop()
                self.VNA_start_button.setEnabled(True)

        except Exception as e:
            # self.work_timeout = 30
            self.append_vna_message(f"Error processing output: {str(e)}", is_error=True)
            self.VNA_start_button.setEnabled(True)

    def _on_vna_idle_timeout(self):
        """Called when no output line has been received from the VNA script for 2 minutes."""
        self.append_vna_message(
            "=== ERROR: No data from Raspberry Pi — EtherCAT data broken. Please contact support team. ===",
            is_error=True,
        )
        self._show_idle_timeout_error("VNA Calibration")
        self.vna_timer.stop()
        self.worker.stop()
        self.VNA_start_button.setEnabled(True)
        self._set_tabs_locked(False)

    def on_vna_test_finished(self):
        self.vna_timer.stop()
        self.vna_idle_timer.stop()
        self.VNA_start_button.setEnabled(True)
        self._set_tabs_locked(False)
        if hasattr(self, 'worker') and self.worker:
            self.worker.stop()

        # If test didn't explicitly pass or fail, mark it as incomplete
        if "Passed" not in self.VNA_status_label_start.text() and "Failed" not in self.VNA_status_label_start.text():
            self.VNA_status_label_start.setText('● Incomplete')
            self.VNA_status_label_start.setStyleSheet(self._PILL_GRAY_SS)
            self.append_vna_message("\n!!! Test did not complete properly !!!", is_error=True)
        # self.work_timeout = 30

    def handle_vna_error(self, error_msg):
        self.vna_timer.stop()
        self.vna_idle_timer.stop()
        # self.work_timeout = 30
        self.cleanup_resources()
        self.append_vna_message(f"\n!!!  ERROR: {error_msg} !!!", is_error=True)
        self.VNA_start_button.setEnabled(True)

    def program_otp(self):
        if hasattr(self, 'console_output') and self.console_output is not None:
            self.console_output.clear()
        pcb_part_number = self.pcb_pn_input.text().strip()
        assembly_part_number = self.assembly_pn_input.text().strip()
        assembly_ser = self.assembly_sn_input.text().strip()
        assembly_rev = self.assembly_rev_input.text().strip()
        PCB_rev = self.pcb_rev_input.text().strip()
        PCB_ser = self.pcb_sn_input.text().strip()
        if len(pcb_part_number) == 0:
            self.append_console_message(f"ERROR: PCB part number not Entered\n", is_error=True)
            QMessageBox.critical(self, "Error", f"PCB part number not Entered")
            self.logger.error(f"PCB part number not Entered",
                              extra={'func_name': 'auto_load_connect'})
            return
        if len(PCB_rev) == 0:
            self.append_console_message(f"ERROR: PCB Revision not Entered\n", is_error=True)
            QMessageBox.critical(self, "Error", f"PCB Revision not Entered")
            self.logger.error(f"PCB Revision not Entered",
                              extra={'func_name': 'auto_load_connect'})
            return
        if len(assembly_ser) == 0:
            self.append_console_message(f"ERROR: Assembly Serial number not Entered\n", is_error=True)
            QMessageBox.critical(self, "Error", f"Assembly Serial number not Entered")
            # logger.error("Assembly Serial number not Entered")
            self.logger.error(f"Assembly Serial number not Entered",
                              extra={'func_name': 'auto_load_connect'})
            return
        if len(assembly_rev) == 0:
            self.append_console_message(f"ERROR: Assembly Revision not Entered\n", is_error=True)
            QMessageBox.critical(self, "Error", f"Assembly Revision not Entered")
            # logger.error("Assembly Revision not Entered")
            self.logger.error(f"Assembly Revision not Entered",
                              extra={'func_name': 'auto_load_connect'})
            return

        if len(PCB_ser) == 0:
            self.append_console_message(f"ERROR: PCB Serial not Entered\n", is_error=True)
            QMessageBox.critical(self, "Error", f"PCB Serial not Entered")
            self.logger.error(f"PCB Serial not Entered",
                              extra={'func_name': 'auto_load_connect'})
            return

        if len(assembly_part_number) == 0:
            self.append_console_message(f"ERROR :Assembly part number  not Entered \n", is_error=True)
            QMessageBox.critical(self, "Error", f"Assembly part number  not Entered")
            self.logger.error(f"Assembly part number  not Entered",
                              extra={'func_name': 'auto_load_connect'})
            return

        if not pcb_part_number or not self.validate_part_number(pcb_part_number, "PCB"):
            self.append_console_message(f"ERROR: Invalid PCB Part Number: {pcb_part_number}\n", is_error=True)
            self.logger.error(f"ERROR: Invalid PCB Part Number: {pcb_part_number}",
                              extra={'func_name': 'auto_load_connect'})
            return

        if not PCB_rev or not self.validate_revision_number(PCB_rev, "PCB"):
            self.append_console_message(f"ERROR: Invalid PCB Revision Number: {PCB_rev}\n", is_error=True)
            self.logger.error(f"ERROR: Invalid PCB Revision : {PCB_rev}",
                              extra={'func_name': 'auto_load_connect'})
            return

        if not assembly_part_number or not self.validate_part_number(assembly_part_number, "Assembly"):
            self.append_console_message(f"ERROR: Invalid Assembly Part Number: {assembly_part_number}\n",
                                        is_error=True)
            self.logger.error(f"ERROR: Invalid Assembly Part Number: {assembly_part_number}",
                              extra={'func_name': 'auto_load_connect'})
            return

        if not assembly_rev or not self.validate_revision_number(assembly_rev, "Assembly"):
            self.append_console_message(f"ERROR: Invalid Assembly Revision Number: {assembly_rev}\n", is_error=True)
            self.logger.error(f"ERROR: Invalid Assembly Revision : {assembly_rev}",
                              extra={'func_name': 'auto_load_connect'})
            return

        #self.append_console_message("Part numbers validated successfully. Connecting...\n\n")
        QApplication.processEvents()  # Update UI


        self.create_otp_file("C:\\tmp\\APPOTP", pcb_part_number, PCB_ser, assembly_part_number,assembly_ser)
        self.file_transer("C:\\tmp\\APPOTP","/home/robot/Manufacturing_test/aipc_beta/APPOTP")
        time.sleep(1)
        success, message = self.ssh_handler.Connect_RPI()
        if not success:
            self.handle_ssh_error(f"Connection failed: {message}")
            return

        self.execute_command("programotp", self.handle_otp_test_output, 0)

    def handle_otp_test_output(self,stdout, stderr):
        if "mailbox error" in (stdout or "").lower() or "mailbox error" in (stderr or "").lower():
            self.append_console_message(
                "Mailbox Error on Ethercat please check the Ethercat Data or Contact Support Team",
                is_error=True
            )
            return

        test_passed = "UPDATE_PASS" in stdout
        test_details = stdout.strip()

        if "Error in slave initialization" in test_details:
            QMessageBox.critical(
                self,
                "Critical Error",
                "Slave initialization failed! Please check the EtherCAT connection and restart the test."
            )
            self.test_status_label_start.setText("Failed")

        if test_passed:
            self.append_console_message("OTP Update PASS")
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText(f"Insert the SD card and update the Firmware")
            msg.setWindowTitle("Firmware Update")
            msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
            retval = msg.exec_()

        else:
            self.append_console_message("OTP Update FAIL", is_error=True)



    def auto_load_connect(self):
        self.test_result = "PASS"
        if hasattr(self, 'console_output') and self.console_output is not None:
            self.console_output.clear()

        try:
            # self.config = self.load_config()
            
            if self.unit_test >= 1:
                excel_logger.reset_sheet("Unit Setup")
            self.unit_test += 1
            # Validate PCB Part Number
            # logger.info("Unit SETUP Test")
            self.auto_load_btn.setEnabled(False)
            self._set_tabs_locked(True)
            self.append_console_message("==========TEST Started=============\n")
            Fixture = self.Fixture.text().strip()
            Testing_name = self.Test_name.text().strip()
            Ecal = self.Ecal_SN.text().strip()
            VNA = self.VNA_SN.text().strip()
            Vendor_name = self.Vendor_name.text().strip()
            pcb_part_number = self.pcb_pn_input.text().strip()
            assembly_part_number = self.assembly_pn_input.text().strip()
            assembly_ser = self.assembly_sn_input.text().strip()
            assembly_rev = self.assembly_rev_input.text().strip()
            PCB_rev = self.pcb_rev_input.text().strip()
            PCB_ser = self.pcb_sn_input.text().strip()
            self.test_result = 'PASS'
            self.PN = assembly_part_number
            self.SN = assembly_ser

            self.append_console_message("1. Validate Entry check\n")

            if len(pcb_part_number) == 0:
                self.append_console_message(f"ERROR: PCB part number not Entered\n", is_error=True)
                QMessageBox.critical(self, "Error", f"PCB part number not Entered")
                self.logger.error(f"PCB part number not Entered",
                                  extra={'func_name': 'auto_load_connect'})
                return
            if len(PCB_rev) == 0:
                self.append_console_message(f"ERROR: PCB Revision not Entered\n", is_error=True)
                QMessageBox.critical(self, "Error", f"PCB Revision not Entered")
                self.logger.error(f"PCB Revision not Entered",
                                  extra={'func_name': 'auto_load_connect'})
                return
            if len(assembly_ser) == 0:
                self.append_console_message(f"ERROR: Assembly Serial number not Entered\n", is_error=True)
                QMessageBox.critical(self, "Error", f"Assembly Serial number not Entered")
                # logger.error("Assembly Serial number not Entered")
                self.logger.error(f"Assembly Serial number not Entered",
                                  extra={'func_name': 'auto_load_connect'})
                return
            if len(assembly_rev) == 0:
                self.append_console_message(f"ERROR: Assembly Revision not Entered\n", is_error=True)
                QMessageBox.critical(self, "Error", f"Assembly Revision not Entered")
                # logger.error("Assembly Revision not Entered")
                self.logger.error(f"Assembly Revision not Entered",
                                  extra={'func_name': 'auto_load_connect'})
                return

            if len(PCB_ser) == 0:
                self.append_console_message(f"ERROR: PCB Serial not Entered\n", is_error=True)
                QMessageBox.critical(self, "Error", f"PCB Serial not Entered")
                self.logger.error(f"PCB Serial not Entered",
                                  extra={'func_name': 'auto_load_connect'})
                return

            if len(assembly_part_number) == 0:
                self.append_console_message(f"ERROR :Assembly part number  not Entered \n", is_error=True)
                QMessageBox.critical(self, "Error", f"Assembly part number  not Entered")
                self.logger.error(f"Assembly part number  not Entered",
                                  extra={'func_name': 'auto_load_connect'})
                return

            if len(Vendor_name) == 0:
                self.append_console_message(f"ERROR :Vendor Name not Entered \n", is_error=True)
                QMessageBox.critical(self, "Error", f"Vendor Name not Entered")
                self.logger.error(f"Vendor Name  not Entered",
                                  extra={'func_name': 'auto_load_connect'})
                return

            if len(Fixture) == 0:
                self.append_console_message(f"ERROR :Fixture Number not Entered \n", is_error=True)
                QMessageBox.critical(self, "Error", f"Fixture Number not Entered")
                self.logger.error(f"Fixture Number not Entered",
                                  extra={'func_name': 'auto_load_connect'})
                return

            if len(Testing_name) == 0:
                self.append_console_message(f"ERROR: Test Name not Entered\n", is_error=True)
                QMessageBox.critical(self, "Error", f"Test Operator Name not Entered")
                self.logger.error(f"Test Operator Name not Entered",
                                  extra={'func_name': 'auto_load_connect'})
                return
            if len(Ecal) == 0:
                self.append_console_message(f"ERROR : ECAL SN not Entered \n", is_error=True)
                QMessageBox.critical(self, "Error", f"ECAL SN not Entered")
                self.logger.error(f"ECAL SN not Entered",
                                  extra={'func_name': 'auto_load_connect'})
                return
            if len(VNA) == 0:
                self.append_console_message(f"ERROR: VNA SN not Entered\n", is_error=True)
                QMessageBox.critical(self, "Error", f"VNA SN not Entered")
                self.logger.error(f"VNA SN not Entered",
                                  extra={'func_name': 'auto_load_connect'})
                return

            if not pcb_part_number or not self.validate_part_number(pcb_part_number, "PCB"):
                self.append_console_message(f"ERROR: Invalid PCB Part Number: {pcb_part_number}\n", is_error=True)
                self.logger.error(f"ERROR: Invalid PCB Part Number: {pcb_part_number}",
                                  extra={'func_name': 'auto_load_connect'})
                return

            if not PCB_rev or not self.validate_revision_number(PCB_rev, "PCB"):
                self.append_console_message(f"ERROR: Invalid PCB Revision Number: {PCB_rev}\n", is_error=True)
                self.logger.error(f"ERROR: Invalid PCB Revision : {PCB_rev}",
                                  extra={'func_name': 'auto_load_connect'})
                return

            if not assembly_part_number or not self.validate_part_number(assembly_part_number, "Assembly"):
                self.append_console_message(f"ERROR: Invalid Assembly Part Number: {assembly_part_number}\n",
                                            is_error=True)
                self.logger.error(f"ERROR: Invalid Assembly Part Number: {assembly_part_number}",
                                  extra={'func_name': 'auto_load_connect'})
                return

            if not assembly_rev or not self.validate_revision_number(assembly_rev, "Assembly"):
                self.append_console_message(f"ERROR: Invalid Assembly Revision Number: {assembly_rev}\n", is_error=True)
                self.logger.error(f"ERROR: Invalid Assembly Revision : {assembly_rev}",
                                  extra={'func_name': 'auto_load_connect'})
                return



            self.append_console_message("Part numbers validated successfully. Connecting...\n\n")
            QApplication.processEvents()  # Update UI

            assembly_suffix = self._extract_assembly_suffix(assembly_part_number)
            self.assembly_suffix = assembly_suffix
            if assembly_suffix:
                self.append_console_message(f"Using configuration for assembly suffix: {assembly_suffix}\n")
                if self.config_transfer(assembly_suffix) == False:
                    return
                self.config = self.load_config(assembly_suffix)
                

                

            else:
                self.append_console_message(
                    f"WARNING: Assembly part number suffix not recognized.Please check the part no and  update the config file\n", is_error=True)
                return
                #self.config_transfer()  # Use default config


            # Connect to SSH
            success, message = self.ssh_handler.Connect_RPI()
            if not success:
                self.handle_ssh_error(f"Connection failed: {message}")
                return

            # Run SOEM compile in a background thread so the UI stays
            # responsive during the (potentially long) compile step.
            self.append_console_message("\n2. soemcompile\n")
            QApplication.processEvents()
            self._soem_thread_started = True   # tell finally not to clean up
            self.soem_compile_worker = SoemCompileWorker(self.ssh_handler)
            # Do NOT connect output_ready to the console – individual compile
            # lines are suppressed.  A QTimer animates a "SOEM compiling…" dot
            # indicator instead.
            self.soem_compile_worker.compile_done.connect(self._on_soemcompile_done)
            self.soem_compile_worker.error_occurred.connect(self._on_soemcompile_error)
            self.soem_compile_worker.start()
            # Kick off the animated loading indicator
            self._soem_dot_count = 0
            self.append_console_message("SOEM compiling")
            self._soem_loading_timer.start(500)  # tick every 500 ms
            return

        except Exception as e:
            # self.logger.error(f"Error in auto_load_connect: {str(e)}",exc_info=True,extra={'func_name': 'auto_load_connect'} )
            self.auto_load_btn.setEnabled(True)
            self._set_tabs_locked(False)
            QMessageBox.critical(self, "Error", f"Auto load failed: {str(e)}")
        finally:
            # Only clean up here if the soemcompile thread was NOT started.
            # When the thread IS running, _on_soemcompile_done / _on_soemcompile_error
            # are responsible for re-enabling the button and disconnecting SSH.
            # NOTE: _soem_thread_started is only ever written on the main GUI thread
            # (here and in the two continuation slots), so no additional locking
            # is required.
            if not self._soem_thread_started:
                self.auto_load_btn.setEnabled(True)
                self._set_tabs_locked(False)
                self.ssh_handler.SSH_disconnect()

    # ------------------------------------------------------------------ #
    # Continuation slots called by SoemCompileWorker signals              #
    # ------------------------------------------------------------------ #

    def _replace_last_console_line(self, html):
        """Replace the text content of the last line in console_output with *html*.

        Uses StartOfBlock / EndOfBlock anchors so the block separator (newline)
        is preserved and adjacent lines are not merged.
        """
        if not (hasattr(self, 'console_output') and self.console_output is not None):
            return
        cursor = self.console_output.textCursor()
        cursor.movePosition(QTextCursor.End)
        cursor.movePosition(QTextCursor.StartOfBlock)
        cursor.movePosition(QTextCursor.EndOfBlock, QTextCursor.KeepAnchor)
        cursor.removeSelectedText()
        cursor.insertHtml(html)
        self.console_output.verticalScrollBar().setValue(
            self.console_output.verticalScrollBar().maximum()
        )

    def _update_soem_loading_line(self):
        """Timer slot: cycles the trailing dots on the 'SOEM compiling…' line."""
        self._soem_dot_count = (self._soem_dot_count % 3) + 1
        dots = "." * self._soem_dot_count
        self._replace_last_console_line(
            f'<span style="color:#3fb950; font-weight:bold;">SOEM compiling{dots}</span>'
        )

    def _stop_soem_loading(self):
        """Stop the loading animation and finalise the indicator line."""
        self._soem_loading_timer.stop()

    def _on_soemcompile_error(self, error_msg):
        """Called when SoemCompileWorker emits error_occurred."""
        self._stop_soem_loading()
        self._replace_last_console_line(
            '<span style="color:#f85149; font-weight:bold;">SOEM compile FAILED</span>'
        )
        self.append_console_message(f"!!! ERROR !!!\n{error_msg}\n", is_error=True)
        self.logger.error(error_msg, extra={'func_name': 'soemcompile'})
        self.auto_load_btn.setEnabled(True)
        self._set_tabs_locked(False)
        self.ssh_handler.SSH_disconnect()
        self._soem_thread_started = False

    def _on_soemcompile_done(self, stdout, stderr):
        """Called when SoemCompileWorker emits compile_done.
        Handles the compile output, then runs the remaining unit-setup
        commands (firmwarecheck, otpcheck, slaveinfo) and finalizes logging.
        """
        self._stop_soem_loading()
        self._replace_last_console_line(
            '<span style="color:#3fb950; font-weight:bold;">SOEM compile done</span>'
        )
        try:
            if stdout:
                self.logger.debug(f"soemcompile stdout:\n{stdout}",
                                  extra={'func_name': 'soemcompile'})
            if stderr:
                self.logger.error(f"soemcompile stderr:\n{stderr}",
                                  extra={'func_name': 'soemcompile'})

            if not self.handle_soemcompile_output(stdout, stderr):
                return  # early exit; finally block handles cleanup

            # Continue with remaining commands sequentially
            commands = [
                ("firmwarecheck", self.handle_firmare_check_output),
                ("otpcheck", self.handle_otpcheck_output),
                ("slaveinfo", self.handle_slaveinfo_output)
            ]
            val = 3
            for cmd, handler in commands:
                if not self.execute_command(cmd, handler, val):
                    if cmd != 'firmwarecheck':
                        return  # early exit; finally block handles cleanup
                    if self.Firmware_check == False:
                        return  # early exit; finally block handles cleanup
                val = val + 1

            self.append_console_message(
                "======================== Unit Setup completed====================================\n")
            # Log metadata along with test data
            self.excel_logger.log_summary(
                metadata={
                    'eid': f'{self.assembly_pn_input.text().strip()}_{self.assembly_rev_input.text().strip()}_{self.assembly_sn_input.text().strip()[-3:]}',
                    'serial_number': self.assembly_sn_input.text().strip(),
                    'model_number': 'AIPC I-BOX',
                    'version': 'V1.0',
                    'tester_name': self.Test_name.text().strip(),
                    'comment': 'Control Limit',
                    'start_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    'end_time': '',  # Will be filled when test completes
                    'overall_result': '',  # Will be updated to PASS/FAIL
                    'test_fixture_sn': self.Fixture.text().strip(),
                    'vna_sn': self.VNA_SN.text().strip(),
                    'ecal_sn': self.Ecal_SN.text().strip()
                }
            )

            unit_data = {
                'Vendor_name': self.Vendor_name.text().strip(),
                'Fixture_number': self.Fixture.text().strip(),
                'test_operator_name': self.Test_name.text().strip(),
                'test_date': self.Test_Date.date().toString("yyyy-MM-dd"),
                'vna_calibration_date': self.VNA_calibration.date().toString("yyyy-MM-dd"),
                'vna_sn': self.VNA_SN.text().strip(),
                'ecal_sn': self.Ecal_SN.text().strip(),
                'pcb_part_number': self.pcb_pn_input.text().strip(),
                'pcb_revision': self.pcb_rev_input.text().strip(),
                'pcb_serial_number': self.pcb_sn_input.text().strip(),
                'assembly_part_number': self.assembly_pn_input.text().strip(),
                'assembly_revision': self.assembly_rev_input.text().strip(),
                'assembly_serial_number': self.assembly_sn_input.text().strip(),
                'product_id': self.product_id.text().strip(),
                'esi_revision': self.esi_revision.text().strip(),
                'configuration_id': self.configuration_id.text().strip(),
                'ethercat_address': self.ethercat_address.text().strip(),
                'firmware_version': self.firmware_version.text().strip()
            }
            self.excel_logger.log_unit_setup(unit_data)
            self.excel_logger.update_overall_result(self.test_result, PN=self.PN, SN=self.SN, finalize=False)

        except Exception as e:
            self.logger.error(f"Error after soemcompile: {str(e)}", exc_info=True,
                              extra={'func_name': '_on_soemcompile_done'})
            QMessageBox.critical(self, "Error", f"Auto load failed: {str(e)}")
        finally:
            self.auto_load_btn.setEnabled(True)
            self._set_tabs_locked(False)
            self.ssh_handler.SSH_disconnect()
            self._soem_thread_started = False

    def append_vna_message(self, message, is_error=False):
        """Helper method to append colored messages to VNA console"""
        if hasattr(self, 'VNAtest_console') and self.VNAtest_console is not None:
            if is_error:
                self.VNAtest_console.append(f'<span style="color:#f85149; font-weight:bold;">{message}</span>')
            else:
                self.VNAtest_console.append(f'<span style="color:#3fb950; font-weight:bold;">{message}</span>')

            # Auto-scroll to bottom
            self.VNAtest_console.verticalScrollBar().setValue(
                self.VNAtest_console.verticalScrollBar().maximum()
            )

    def append_BNC_message(self, message, is_error=False):
        """Append a colour-coded HTML message to the BNC test console.

        Args:
            message (str): The text to display.
            is_error (bool): When True the text is rendered in red; otherwise green.
        """
        if hasattr(self, 'BNCtest_console') and self.BNCtest_console is not None:
            # Vibrant colours chosen for the dark terminal background
            colour = "#f85149" if is_error else "#3fb950"
            self.BNCtest_console.append(
                f'<span style="color:{colour}; font-weight:bold;">{message}</span>'
            )
            # Auto-scroll to the latest output
            self.BNCtest_console.verticalScrollBar().setValue(
                self.BNCtest_console.verticalScrollBar().maximum()
            )

    def BNC_test(self):
        """Start the BNC Port Verification test sequence.

        Resets state from any previous run, clears the console, and presents
        the first zone-connection prompt (Zone 2) to the operator.
        """
        try:
            if self.bnc_t >= 1:
                self.excel_logger.reset_sheet("BNC Port Verification")
            self.bnc_t += 1
            # Do NOT reset over_all_result here — failures from prior test
            # sections (resistance, impedance, etc.) must persist in the
            # overall result so the final filename reflects them.

            # Stop any previously running worker
            if hasattr(self, 'worker') and self.worker:
                self.worker.stop()

            # Reset zone indicator labels to "pending" state
            if hasattr(self, 'BNC_zone_labels'):
                for znum, lbl in self.BNC_zone_labels.items():
                    subtitle = self._BNC_ZONE_SUBTITLES.get(znum, "")
                    lbl.setText(f"⏳  Zone {znum}\n{subtitle}")
                    lbl.setStyleSheet("""
                        QLabel {
                            background-color: #fff3cd;
                            color: #856404;
                            border: 2px solid #ffc107;
                            border-radius: 6px;
                            font-size: 9pt;
                            font-weight: bold;
                            padding: 4px 6px;
                        }
                    """)

            # Reset progress bar
            if hasattr(self, 'bnc_progress_bar'):
                self.bnc_progress_bar.setValue(0)

            # Reset UI to "running" state
            self.BNCtest_console.clear()
            self.BNC_status_label_start.setText("● Running…")
            self.BNC_status_label_start.setStyleSheet("""
                QLabel {
                    background-color: #fd7e14;
                    color: white;
                    padding: 6px 14px;
                    border-radius: 14px;
                    font-weight: bold;
                    font-size: 10pt;
                }
            """)
            self.BNC_start_button.setEnabled(False)

            self.append_BNC_message(
                "\n================== BNC Test Started =======================\n"
            )

            # Kick off with the first zone prompt
            self._set_tabs_locked(True)
            self.show_zone_prompt(2)

        except Exception as e:
            self.append_BNC_message(f"Failed to start test: {str(e)}", is_error=True)
            self.cleanup_resources()

    def show_zone_prompt(self, zone_number):
        """Prompt the operator to connect a zone, then launch the remote test worker.

        Args:
            zone_number (int): The zone number to display in the prompt dialog.
        """
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText(f"Please connect Zone {zone_number} and click OK to continue")
        msg.setWindowTitle("Zone Connection")
        msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        command = f'{zone_number} dimm'

        retval = msg.exec_()

        if retval == QMessageBox.Ok:
            self.append_BNC_message(f"\nTesting Zone {zone_number}...\n")

            # Start a new worker for the remote BNC test script
            self.worker = Worker(
                self.ssh_handler,
                '/home/robot/Manufacturing_test/aipc_beta/BNC.py',
                command
            )
            self.worker.output_ready.connect(self.handle_BNC_output)
            self.worker.error_occurred.connect(self.handle_BNC_error)
            self.worker.finished_signal.connect(lambda: self._set_tabs_locked(False))
            self.worker.start()
            self._set_tabs_locked(True)
            # Start (or restart) the 5-minute idle watchdog
            self.bnc_idle_timer.start(self._IDLE_TIMEOUT_MS)
        else:
            self.bnc_idle_timer.stop()
            self.append_BNC_message("Test cancelled by user", is_error=True)
            self.BNC_start_button.setEnabled(True)
            self._set_tabs_locked(False)
            self.BNC_status_label_start.setText("● Cancelled")
            self.BNC_status_label_start.setStyleSheet("""
                QLabel {
                    background-color: #dc3545;
                    color: white;
                    padding: 6px 14px;
                    border-radius: 14px;
                    font-weight: bold;
                    font-size: 10pt;
                }
            """)

    # ------------------------------------------------------------------ #
    # BNC zone result helpers                                              #
    # ------------------------------------------------------------------ #

    # Maps the zone number to its human-readable subtitle used in the UI pill
    _BNC_ZONE_SUBTITLES = {
        2: "Mid-Inner",
        3: "Mid-Edge",
        4: "Edge",
        5: "Outer",
    }

    def _update_bnc_zone_label(self, zone_num, passed):
        """Update the visual indicator for a completed BNC zone.

        Args:
            zone_num (int): Zone number (2–5).
            passed (bool): Whether the zone measurement passed.
        """
        if not hasattr(self, 'BNC_zone_labels'):
            return
        lbl = self.BNC_zone_labels.get(zone_num)
        if lbl is None:
            return
        subtitle = self._BNC_ZONE_SUBTITLES.get(zone_num, "")
        if passed:
            icon = "✅"
            style = """
                QLabel {
                    background-color: #d4edda;
                    color: #155724;
                    border: 2px solid #28a745;
                    border-radius: 6px;
                    font-size: 9pt;
                    font-weight: bold;
                    padding: 4px 6px;
                }
            """
        else:
            icon = "❌"
            style = """
                QLabel {
                    background-color: #f8d7da;
                    color: #721c24;
                    border: 2px solid #dc3545;
                    border-radius: 6px;
                    font-size: 9pt;
                    font-weight: bold;
                    padding: 4px 6px;
                }
            """
        lbl.setText(f"{icon}  Zone {zone_num}\n{subtitle}")
        lbl.setStyleSheet(style)

        # Advance the progress bar
        if hasattr(self, 'bnc_progress_bar'):
            current = self.bnc_progress_bar.value()
            self.bnc_progress_bar.setValue(current + 1)

    def _handle_bnc_zone_result(self, zone_label, testpoint_label, line, next_zone_number=None):
        """Parse a CSV result line for a BNC zone and log the outcome.

        Expected line format: ``<zone_name>,<value_dB>,<PASS|FAIL>``

        After logging, the worker for the current zone is stopped.  If
        *next_zone_number* is given the operator is prompted to connect that
        zone; otherwise the full test sequence is marked as complete.

        Args:
            zone_label (str): Zone identifier expected in the output line
                (e.g. ``"Zone2-Mid_Inner"``).
            testpoint_label (str): Short label used in the summary log
                (e.g. ``"Zone2"``).
            line (str): Raw output line received from the remote script.
            next_zone_number (int or None): Zone number to prompt next, or
                ``None`` when this is the final zone.
        """
        parts = line.split(",")
        if len(parts) < 3:
            self.append_BNC_message(
                f"Invalid data format for {zone_label}: {line}", is_error=True
            )
            return

        test_name = parts[0]
        value = parts[1]
        passed = parts[2].strip().upper() == "PASS"

        # Extract zone number from testpoint_label (e.g. "Zone2" → 2)
        try:
            zone_num = int(testpoint_label.replace("Zone", ""))
        except ValueError:
            zone_num = None

        if passed:
            self.append_BNC_message(f"\nBNC Test {zone_label} PASS\n")
        else:
            self.append_BNC_message(f"\nBNC Test {zone_label} FAIL\n", is_error=True)
            self.over_all_result = 'FAIL'

        # Update the visual zone indicator
        if zone_num is not None:
            self._update_bnc_zone_label(zone_num, passed)

        self.excel_logger.log_BNC_measurement(
            test_zone=test_name,
            test_details=value,
            test_passed=passed,
        )

        self.step_no += 1
        self.excel_logger.log_summary(
            step_data={
                'step': str(self.step_no),
                'unit': 'dB',
                'low_limit': '-1',
                'measure': value,
                'high_limit': '0',
                'teststep': 'Verify BNC port',
                'testpoints': testpoint_label,
                'status': "PASS" if passed else "FAIL",
            }
        )

        self.worker.stop()

        if next_zone_number is not None:
            # Advance to the next zone
            self.show_zone_prompt(next_zone_number)
        else:
            # All zones complete – finalise the test run
            self.excel_logger.log_summary(
                metadata={
                    'end_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    'overall_result': self.over_all_result,
                }
            )
            self.excel_logger.update_overall_result(self.over_all_result)
            self.append_BNC_message("\nBNC Test completed successfully\n")

            # Update status pill to Completed / Failed depending on overall
            if self.over_all_result == 'PASS':
                status_text = "● Completed — PASS"
                status_style = """
                    QLabel {
                        background-color: #28a745;
                        color: white;
                        padding: 6px 14px;
                        border-radius: 14px;
                        font-weight: bold;
                        font-size: 10pt;
                    }
                """
            else:
                status_text = "● Completed — FAIL"
                status_style = """
                    QLabel {
                        background-color: #dc3545;
                        color: white;
                        padding: 6px 14px;
                        border-radius: 14px;
                        font-weight: bold;
                        font-size: 10pt;
                    }
                """
            self.BNC_status_label_start.setText(status_text)
            self.BNC_status_label_start.setStyleSheet(status_style)
            self.bnc_idle_timer.stop()
            self.BNC_start_button.setEnabled(True)
            self._set_tabs_locked(False)

    def handle_BNC_output(self, line):
        """Handle a single line of output from the remote BNC test script.

        Infrastructure errors (EtherCAT slave init failure, PyVISA errors) are
        shown as blocking modal dialogs and abort the worker immediately.

        Recognised zone-result lines are dispatched to
        :meth:`_handle_bnc_zone_result`, which parses the CSV payload, logs
        the result, and advances to the next zone prompt.

        A 5-minute idle watchdog timer is reset on every received line.  If
        no line arrives within that window the timer fires, displays an error,
        and restores the UI to its idle state.

        Args:
            line (str): A single line of text received from the remote process.
        """
        # Reset the idle watchdog on every received line
        self.bnc_idle_timer.start(self._IDLE_TIMEOUT_MS)

        # --- Infrastructure error checks ------------------------------------
        if "Error in slave initialization" in line:
            QMessageBox.critical(
                self,
                "Critical Error",
                "Slave initialization failed!\n"
                "Please check the EtherCAT connection and restart the test.",
            )
            self.worker.stop()
            self._bnc_restore_failed_state()
            return

        if "pyvisa.errors" in line:
            QMessageBox.critical(
                self,
                "PyVISA Error",
                f"A PyVISA error occurred:\n{line.strip()}",
            )
            self.worker.stop()
            self._bnc_restore_failed_state()
            return

        if "mailbox error" in line.lower():
            self.append_BNC_message(
                "Mailbox Error on Ethercat please check the Ethercat Data or Contact Support Team",
                is_error=True
            )
            self.worker.stop()
            self._bnc_restore_failed_state()
            return

        # --- Zone result dispatch -------------------------------------------
        if "Zone2-Mid_Inner" in line:
            self._handle_bnc_zone_result("Zone2-Mid_Inner", "Zone2", line, next_zone_number=3)

        elif "Zone3-Mid_Edge" in line:
            self._handle_bnc_zone_result("Zone3-Mid_Edge", "Zone3", line, next_zone_number=4)

        elif "Zone4-Edge" in line:
            self._handle_bnc_zone_result("Zone4-Edge", "Zone4", line, next_zone_number=5)

        elif "Zone5-Outer" in line:
            self._handle_bnc_zone_result("Zone5-Outer", "Zone5", line, next_zone_number=None)

    def _on_bnc_idle_timeout(self):
        """Called when no output line has been received from the BNC script for 2 minutes."""
        self.append_BNC_message(
            "=== ERROR: No data from Raspberry Pi — EtherCAT data broken. Please contact support team. ===",
            is_error=True,
        )
        self._show_idle_timeout_error("BNC Test")
        self.worker.stop()
        self._bnc_restore_failed_state()

    def _bnc_restore_failed_state(self):
        """Re-enable the BNC Start button, set status to Failed, and unlock tabs."""
        self.bnc_idle_timer.stop()
        self.BNC_start_button.setEnabled(True)
        self.BNC_status_label_start.setText("● Failed")
        self.BNC_status_label_start.setStyleSheet("""
            QLabel {
                background-color: #dc3545;
                color: white;
                padding: 6px 14px;
                border-radius: 14px;
                font-weight: bold;
                font-size: 10pt;
            }
        """)
        self._set_tabs_locked(False)

    def handle_BNC_error(self, error_msg):
        """Handle an error signal emitted by the BNC test worker.

        Displays the error in the console, re-enables the Start button, and
        cleans up any active resources.

        Args:
            error_msg (str): The error message reported by the worker.
        """
        self.append_BNC_message(f"ERROR: {error_msg}", is_error=True)
        self.cleanup_resources()
        self._bnc_restore_failed_state()


    @log_function
    def execute_command(self, command, output_handler, val):
        """Execute a single command and handle its output"""

        if command != "selftest":
            self.append_console_message(f"\n{str(val)}. {command}\n")
        QApplication.processEvents()  # Update UI
        self.logger.info(f"Executing command: {command}",
                         extra={'func_name': command})
        try:
            stdout, stderr = self.ssh_handler.SSH_com(command)
            self.logger.info(f"Command output received",
                             extra={'func_name': command})

            if stdout:
                self.logger.debug(f"stdout:\n{stdout}",
                                  extra={'func_name': command})
            if stderr:
                self.logger.error(f"stderr:\n{stderr}",
                                  extra={'func_name': command})
            # Process the output
            if stderr and command != 'soemcompile':
                self.append_console_message(f"!!! ERROR !!!\n{stderr}\n", is_error=True)

            return output_handler(stdout, stderr)

        except Exception as e:
            self.logger.error(f"Command failed: {str(e)}",
                              exc_info=True,
                              extra={'func_name': command})
            self.handle_ssh_error(f"Command '{command}' failed: {str(e)}")
            return False

    def handle_otpcheck_output(self, stdout, stderr):
        self.handling_flag = 1

        if "mailbox error" in (stdout or "").lower() or "mailbox error" in (stderr or "").lower():
            self.append_console_message(
                "Mailbox Error on Ethercat please check the Ethercat Data or Contact Support Team",
                is_error=True
            )
            return False

        if "No such file or directory" in stderr:
            error_msg = "ERROR: No OTP file found on device\n"
            self.console_output.append("!!! CRITICAL ERROR !!!\n")
            self.console_output.append(error_msg)
            QMessageBox.critical(self, "OTP Error",
                                 "No OTP file found on device!\n\n"
                                 "Please check if the device has proper OTP configuration.\n"
                                 "Expected file: /dev/otp/APPOTP")
            return False

        parsed_data = self.parse_ssh_output(stdout)

        # Compare OTP programmed values with GUI entered values
        gui_pcb_pn = self.pcb_pn_input.text().strip()
        gui_pcb_sn = self.pcb_sn_input.text().strip()
        gui_assy_pn = self.assembly_pn_input.text().strip()
        gui_assy_sn = self.assembly_sn_input.text().strip()

        otp_pcb_pn = parsed_data.get('pcb_part_number', '')
        otp_pcb_pn = otp_pcb_pn.split('_')[0] if otp_pcb_pn else ''
        otp_pcb_sn = parsed_data.get('pcb_serial_number', '')
        otp_assy_pn = parsed_data.get('assembly_part_number', '')
        otp_assy_sn = parsed_data.get('assembly_serial_number', '')

        match_failures = []

        if gui_pcb_pn and otp_pcb_pn and gui_pcb_pn != otp_pcb_pn:
            match_failures.append(f"PCB Part Number mismatch (GUI: {gui_pcb_pn} vs OTP: {otp_pcb_pn})")

        if gui_pcb_sn and otp_pcb_sn and gui_pcb_sn != otp_pcb_sn:
            match_failures.append(f"PCB Serial Number mismatch (GUI: {gui_pcb_sn} vs OTP: {otp_pcb_sn})")

        if gui_assy_pn and otp_assy_pn and gui_assy_pn != otp_assy_pn:
            match_failures.append(f"Assembly Part Number mismatch (GUI: {gui_assy_pn} vs OTP: {otp_assy_pn})")

        if gui_assy_sn and otp_assy_sn and gui_assy_sn != otp_assy_sn:
            match_failures.append(f"Assembly Serial Number mismatch (GUI: {gui_assy_sn} vs OTP: {otp_assy_sn})")

        if match_failures:
            error_msg = "OTP Programming Verification Failed:\n" + "\n".join(match_failures)
            self.append_console_message("!!! OTP PROGRAMMING ERROR !!!\n", is_error=True)
            self.append_console_message(error_msg + "\n", is_error=True)
            self.test_result = 'FAIL'
            return False
        else:
            success_msg = "OTP Programming Verified Successfully!\n"
            success_msg += f"PCB Part Number: {otp_pcb_pn}\n"
            success_msg += f"PCB Serial Number: {otp_pcb_sn}\n"
            success_msg += f"Assembly Part Number: {otp_assy_pn}\n"
            success_msg += f"Assembly Serial Number: {otp_assy_sn}\n"

            # self.append_console_message("=== OTP VERIFICATION SUCCESS ===\n")
            self.append_console_message(success_msg)
            # QMessageBox.information(self, "OTP Verified", success_msg)
            return True

    def handle_firmare_check_output(self, stdout, sterr):
        if "mailbox error" in (stdout or "").lower() or "mailbox error" in (sterr or "").lower():
            self.append_console_message(
                "Mailbox Error on Ethercat please check the Ethercat Data or Contact Support Team",
                is_error=True
            )
            return

        parsed_data = self.parse_ssh_output(stdout)
        Actual_version = parsed_data['firmware_version']
        Expected_version = self.config["expected_firmware_version"]
        if Actual_version == Expected_version:
            self.Firmware_check = True
            self.append_console_message(
                f"✔ Correct version: {Expected_version}")
        else:
            error_msg = "Firmware Mismatch Test cannot be proceed further please update the latest Firmware"
            self.Firmware_check = False
            self.append_console_message(
                f"✖ Incorrect version (expected: {Expected_version} Actual: {Actual_version})",
                is_error=True
            )
            QMessageBox.critical(self, "Firmware Version Mismatch", error_msg)
            self.test_result = 'FAIL'

    def handle_self_test_output(self, stdout, sterr):
        if "mailbox error" in (stdout or "").lower() or "mailbox error" in (sterr or "").lower():
            self.append_self_message(
                "Mailbox Error on Ethercat please check the Ethercat Data or Contact Support Team",
                is_error=True
            )
            return False

        test_passed = "Self Test PASS" in stdout
        test_details = stdout.strip()

        if "Error in slave initialization" in  test_details:
            QMessageBox.critical(
                self,
                "Critical Error",
                "Slave initialization failed! Please check the EtherCAT connection and restart the test."
            )
            self.test_status_label_start.setText("Failed")


        if test_passed:
            self.append_self_message("SELF TEST PASS")
            self.test_status_label_start.setText("● Completed — PASS")
            self.test_status_label_start.setStyleSheet(self._PILL_PASS_SS)
        else:
            self.append_self_message("SELF TEST FAIL", is_error=True)
            self.over_all_result = "FAIL"
            self.test_status_label_start.setText("● Completed — FAIL")
            self.test_status_label_start.setStyleSheet(self._PILL_FAIL_SS)

        # Log to Excel
        #unit_identifier = f"{self.assembly_pn_input.text().strip()} ({self.assembly_sn_input.text().strip()})"
        self.excel_logger.log_self_test(
            unit_identifier="self_test",
            test_passed=test_passed,
            test_details=test_details,
            notes="Self Test completed"
        )

        self.excel_logger.log_summary(
            metadata={
                'end_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'overall_result': self.over_all_result
            }
        )

        return test_passed

    def check_output_for_strings(self, stdout):
        """Check the output for specific strings and return findings"""
        results = {
            'compilation_success': False,
            'pdo_map_success': False,
        }

        # Check for successful compilation
        if "Ethercat compiled Sucessfully" in stdout:
            results['compilation_success'] = True

        # Check for PDO map completion
        if "pdo map successfully reached end" in stdout:
            results['pdo_map_success'] = True

        return results

    # Example usage in your handle_soemcompile_output method:
    def handle_soemcompile_output(self, stdout, stderr):
        self.handling_flag = 2

        if "mailbox error" in (stdout or "").lower() or "mailbox error" in (stderr or "").lower():
            self.append_console_message(
                "Mailbox Error on Ethercat please check the Ethercat Data or Contact Support Team",
                is_error=True
            )
            return False

        # Analyze the output
        analysis = self.check_output_for_strings(stdout)

        # Check for specific conditions
        if not analysis['compilation_success']:
            error_msg = "ERROR: Ethercat compilation failed\n"
            self.append_console_message("!!! CRITICAL ERROR !!!\n", is_error=True)
            self.append_console_message(error_msg, is_error=True)
            QMessageBox.critical(self, "Compilation Error", error_msg)
            self.test_result = 'FAIL'
            return False

        if not analysis['pdo_map_success']:
            error_msg = "ERROR: PDO Map generation failed\n"
            self.append_console_message("!!! CRITICAL ERROR !!!\n", is_error=True)
            self.append_console_message(error_msg, is_error=True)
            QMessageBox.critical(self, "PDO Map Error", error_msg)
            self.test_result = 'FAIL'
            return False

        if "Ethercat compiled Sucessfully" in stdout and "pdo map successfully reached end" in stdout:
            success_msg = "Ethercat compiled Successfully with PDO mapping"
            self.append_console_message(success_msg)
            self.check = True
            return True

        return False

    def handle_slaveinfo_output(self, stdout, stderr):
        Counter = 0

        if "mailbox error" in (stdout or "").lower() or "mailbox error" in (stderr or "").lower():
            self.append_console_message(
                "Mailbox Error on Ethercat please check the Ethercat Data or Contact Support Team",
                is_error=True
            )
            return False

        if not self.check:
            return False

        self.handling_flag = 3
        parsed_data = self.parse_ssh_output(stdout)

        if len(parsed_data['product_id']) == 0 or parsed_data['product_id'] is None:
            self.append_console_message(" Error : Product ID is None", is_error=True)
            Counter = 1
        if len(parsed_data['esi_revision']) == 0 or parsed_data['esi_revision'] is None:
            self.append_console_message(" Error : ESI Revision is None", is_error=True)
            Counter = 1
        if len(parsed_data['ethercat_address']) == 0 or parsed_data['ethercat_address'] is None:
            self.append_console_message(" Error : Ethercat Address is None", is_error=True)
            Counter = 1

        if int(parsed_data['ethercat_address'],16) !=  int('0x444',16) :
            Address = parsed_data['ethercat_address']
            self.append_console_message(f" Error : Ethercat Address Failed Actual_Value = {Address} Expected_Value = 0X444  ", is_error=True)
            self.append_console_message(f"Set hex switch to 0x444 and Power Cycle and restart the test", is_error=True)
            Counter = 1

        self.product_id.setText(parsed_data['product_id'])
        self.esi_revision.setText(parsed_data['esi_revision'])
        self.firmware_version.setText(parsed_data['firmware_version'])
        self.ethercat_address.setText(parsed_data['ethercat_address'])

        if Counter == 0:
            self.append_console_message("Slave Information Test Passed")
        else:
            self.append_console_message("Slave Information Test Fail", is_error=True)
            self.test_result = 'FAIL'

        return True

    def create_form_row(self, label, widget, layout, readonly=False):
        """Create a responsive form row"""
        row = QWidget()
        row_layout = QVBoxLayout(row)
        row_layout.setContentsMargins(0, 0, 0, 0)
        row_layout.setSpacing(2)

        lbl = QLabel(label)
        lbl.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        row_layout.addWidget(lbl)

        if isinstance(widget, (QLineEdit, QDateEdit)):
            widget.setMinimumHeight(30)  # Reduced from 38 for better scaling
            if readonly:
                widget.setReadOnly(True)
                if isinstance(widget, QLineEdit):
                    widget.setStyleSheet("""
                              background-color: #e0e0e0;
                              color: #495057;
                              border: 1px solid #000000;
                              font-weight: bold;
                          """)

        widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        row_layout.addWidget(widget)
        layout.addWidget(row)
        return widget

    def _create_Impedance_zone_panel(self, zone_name):
        """Create a responsive impedance zone panel"""
        panel = QGroupBox(zone_name)

        panel.setStyleSheet("""
              QGroupBox {
                  font-weight: bold; font-size: 9pt; color: white;
                  border: 2px solid rgba(255,255,255,0.35); border-radius: 6px;
                  margin-top: 20px; padding-top: 20px;
                  background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                      stop:0 #004D40, stop:1 #00352c);
              }
              QGroupBox::title {
                  subcontrol-origin: margin; subcontrol-position: top center;
                  padding: 2px 8px; color: white; background-color: #004D40;
              }
          """)

        panel_layout = QVBoxLayout(panel)
        panel_layout.setSpacing(10)
        panel_layout.setContentsMargins(8, 15, 8, 8)

        # Create measurement table with scroll area
        table_scroll = QScrollArea()
        table_scroll.setWidgetResizable(True)
        table_scroll.setMinimumHeight(200)

        measurement_table = QTableWidget(1 if zone_name == "Zone1-Inner" else len(self._relay_values_imp), 5)
        measurement_table.setHorizontalHeaderLabels(["Setpt", "R", "I", "Z", "P/F"])
        measurement_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        measurement_table.verticalHeader().setVisible(False)
        measurement_table.setFont(QFont('Arial', 9))

        # Style the table
        measurement_table.setStyleSheet("""
              QTableWidget {
                  background-color: #0d1117;
                  color: #c9d1d9;
                  gridline-color: #30363d;
                  border: none;
              }
              QHeaderView::section {
                  background-color: #1f2937;
                  color: #7dd3fc;
                  padding: 4px;
                  border: 1px solid #30363d;
                  font-weight: bold;
              }
              QTableWidget::item {
                  background-color: #161b22;
              }
          """)

        # Populate setpoint values
        if zone_name == "Zone1-Inner":
            measurement_table.setItem(0, 0, QTableWidgetItem(str(self._relay_values_imp[0])))
        else:
            for row, relay in enumerate(self._relay_values_imp):
                measurement_table.setItem(row, 0, QTableWidgetItem(str(relay)))

        table_scroll.setWidget(measurement_table)
        panel_layout.addWidget(table_scroll)

        # Add test button
        test_button = QPushButton(f"▶  Test {zone_name}")
        test_button.setFont(QFont('Arial', 9, QFont.Bold))
        test_button.setFixedHeight(32)
        test_button.setCursor(Qt.PointingHandCursor)
        test_button.setStyleSheet("""
              QPushButton {
                  background-color: #00695C; color: white; border: none;
                  border-radius: 4px; padding: 4px 8px;
              }
              QPushButton:hover   { background-color: #004D40; }
              QPushButton:pressed { background-color: #003d31; }
          """)
        test_button.clicked.connect(lambda _, z=zone_name: self._start_impedance_zone_measurement(z))
        panel_layout.addWidget(test_button, alignment=Qt.AlignCenter)

        self._measurement_tables_imp[zone_name] = measurement_table

        return panel

    def _create_resistance_zone_panel(self, zone_name):
        """Create a responsive resistance zone panel"""
        panel = QGroupBox(zone_name)

        panel.setStyleSheet("""
               QGroupBox {
                   font-weight: bold; font-size: 9pt; color: white;
                   border: 2px solid rgba(255,255,255,0.35); border-radius: 6px;
                   margin-top: 20px; padding-top: 20px;
                   background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                       stop:0 #311B92, stop:1 #1a0f5c);
               }
               QGroupBox::title {
                   subcontrol-origin: margin; subcontrol-position: top center;
                   padding: 2px 8px; color: white; background-color: #311B92;
               }
           """)

        panel_layout = QVBoxLayout(panel)
        panel_layout.setSpacing(10)
        panel_layout.setContentsMargins(8, 15, 8, 8)

        # Create measurement table with scroll area
        table_scroll = QScrollArea()
        table_scroll.setWidgetResizable(True)
        table_scroll.setMinimumHeight(200)

        measurement_table = QTableWidget(1 if zone_name == "Zone1-Inner" else len(self._relay_values), 3)
        measurement_table.setHorizontalHeaderLabels(["Setpt", "R", "P/F"])
        measurement_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        measurement_table.verticalHeader().setVisible(False)
        measurement_table.setFont(QFont('Arial', 9))

        # Style the table
        measurement_table.setStyleSheet("""
               QTableWidget {
                   background-color: #0d1117;
                   color: #c9d1d9;
                   gridline-color: #30363d;
                   border: none;
               }
               QHeaderView::section {
                   background-color: #1f2937;
                   color: #c4b5fd;
                   padding: 4px;
                   border: 1px solid #30363d;
                   font-weight: bold;
               }
               QTableWidget::item {
                   background-color: #161b22;
               }
           """)

        # Populate setpoint values
        if zone_name == "Zone1-Inner":
            measurement_table.setItem(0, 0, QTableWidgetItem(str(self._relay_values[0])))
        else:
            for row, relay in enumerate(self._relay_values):
                measurement_table.setItem(row, 0, QTableWidgetItem(str(relay)))

        table_scroll.setWidget(measurement_table)
        panel_layout.addWidget(table_scroll)

        # Add test button
        test_button = QPushButton(f"▶  Test {zone_name}")
        test_button.setFont(QFont('Arial', 9, QFont.Bold))
        test_button.setFixedHeight(32)
        test_button.setCursor(Qt.PointingHandCursor)
        test_button.setStyleSheet("""
               QPushButton {
                   background-color: #4527A0; color: white; border: none;
                   border-radius: 4px; padding: 4px 8px;
               }
               QPushButton:hover   { background-color: #311B92; }
               QPushButton:pressed { background-color: #200d6b; }
           """)
        test_button.clicked.connect(lambda _, z=zone_name: self._start_resistance_zone_measurement(z))
        panel_layout.addWidget(test_button, alignment=Qt.AlignCenter)

        self._measurement_tables[zone_name] = measurement_table

        return panel

    def _clear_resistance_log_display(self):
        self._log_output.clear()

    def _clear_impedance_log_display(self):
        self._log_output_imp.clear()

    def get_table_item_safe(table, row, column):
        item = table.item(row, column)
        return item.text() if item is not None else ""

    def _get_zone_title(self, zone_name):

        zone_titles = {
            "Zone1-Inner": "Zone 1 - Inner",
            "Zone2-Mid_Inner": "Zone 2 - Mid Inner",
            "Zone3-Mid_Edge": "Zone 3 - Mid Edge",
            "Zone4-Edge": "Zone 4 - Edge",
            "Zone5-Outer": "Zone 5 - Outer"
        }
        return zone_titles.get(zone_name, zone_name)

    def process_single_measurement(self, zone_name, measurement_line):
        """
        Process single measurement line from RPi in format: setpoint,resistance,status
        Example: "0,2.5,True"
        """
        try:
            # Remove any whitespace and split the components
            # setpoint, resistance, status = measurement_line.strip().split(',')
            val = measurement_line.strip().split(',')
            if len(val) < 4:
                self._log_resistance_message(f"Invalid measurement format (expected 4+ fields): {measurement_line}", is_error=True)
                return False
            setpoint = val[1]
            resistance = val[2]
            status = val[3]
            # Find the table for this zone
            table = self._measurement_tables.get(zone_name)
            self.config1 = self.load_config(self.assembly_suffix)

            # If config could not be loaded (e.g. suffix missing), stop the test
            if self.config1 is None:
                self._abort_scan_missing_suffix(self._log_resistance_message, self.res_idle_timer)
                return False

            if not table:
                self._log_resistance_message(f"No table found for zone {zone_name}", is_error=True)
                return False

            # Find the row with matching setpoint
            for row in range(table.rowCount()):
                #self._log_resistance_message(f"Inside block1 ", is_error=True)
                if table.item(row, 0).text() == setpoint:

                    #self._log_resistance_message(f"Inside block2 {table.item(row, 0).text()}", is_error=True)
                    # Convert status to display format
                    status_text = "PASS" if status.lower() == "true" else "FAIL"
                    measurement_data = {
                        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'zone_title': self._get_zone_title(zone_name),
                        'setpoint': float(setpoint),
                        'resistance': float(resistance),
                        'status': "PASS" if status.lower() == "true" else "FAIL",
                        'table_row': row + 1
                    }

                    if status == 'False':
                        self.resistance_test = 'FAIL'
                        self.over_all_result = 'FAIL'

                    self.step_no = self.step_no + 1
                    if float(setpoint) != 0.0:
                        limit_per = float(self.config1[setpoint])
                        prod_val = float(self.config1[f'Res{setpoint}'])
                        higher_limit = prod_val + (prod_val * limit_per / 100)
                        lower_limit = prod_val - (prod_val * limit_per / 100)
                        #self._log_resistance_message(f"Inside block3", is_error=True)
                        # Log metadata along with test data
                        self.excel_logger.log_summary(
                            step_data={
                                'step': str(self.step_no),
                                'unit': 'ohm',
                                'low_limit': f'{lower_limit}',
                                'measure': f'{float(resistance)}',
                                'high_limit': f'{higher_limit}',
                                'teststep': 'Resistance Test',
                                'testpoints': f'{self._get_zone_title(zone_name)}_{setpoint}',
                                'status': "PASS" if status.lower() == "true" else "FAIL"
                            }
                        )
                    else:
                        self.excel_logger.log_summary(
                            step_data={
                                'step': str(self.step_no),
                                'unit': 'ohm',
                                'low_limit': '0.0',
                                'measure': f'{float(resistance)}',
                                'high_limit': '0.7',
                                'teststep': 'Resistance Test',
                                'testpoints': f'{self._get_zone_title(zone_name)}_{setpoint}',
                                'status': "PASS" if status.lower() == "true" else "FAIL"
                            }
                        )

                    #if self._get_zone_title(zone_name) == "Zone 5 - Outer" and float(setpoint) == 127.0:

                    #self._log_resistance_message(f"Inside block4", is_error=True)
                    self.excel_logger.log_summary(
                            metadata={
                                'end_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                'overall_result': self.over_all_result
                            }
                        )
                    self.excel_logger.update_overall_result(self.over_all_result, finalize=False)

                    #self._log_resistance_message(f"Inside block5", is_error=True)
                    # Update the table
                    self.update_resistance_measurement(
                        zone_name,
                        row,
                        float(resistance),
                        status_text
                    )

                    #self._log_resistance_message(f"Inside block6", is_error=True)
                    # Log to Excel (combined sheet)
                    if not self.excel_logger.log_resistance_measurement(measurement_data, f'{zone_name}_Res_scan'):
                        raise Exception("Failed to log to Excel")
                    return True

            self._log_resistance_message(f"Setpoint {setpoint} not found in {zone_name}", is_error=True)
            return False

        except ValueError:
            self._log_resistance_message(f"Invalid measurement format: {measurement_line}", is_error=True)
            return False
        except Exception as e:
            self._log_resistance_message(f"Error processing measurement: {str(e)}", is_error=True)
            return False

    def _start_impedance_zone_measurement(self, zone_name):

        try:

            # Guard: abort immediately if assembly suffix is not set
            if not self._check_assembly_suffix_or_abort():
                return

            self.names1 = zone_name
            self.stop_increment = False
            selected_freq = self.freq_combo.currentText()
            frequency = selected_freq.split()[0]
            freq_suffix = ExcelLogger._freq_to_sheet_suffix(selected_freq)
            if selected_freq == "60 MHz":
                command = f"{zone_name} 50000 70000"
            else:
                command = f"{zone_name} {frequency}"


            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText(f"Please connect Zone {zone_name} and click OK to continue")
            msg.setWindowTitle("Impedance Scan")
            msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
            retval = msg.exec_()
            if retval == QMessageBox.Ok:
                self.start_time = time.time()
                if zone_name == "Zone1-Inner":
                    if self.Zone1_Inner_imp >= 1:
                        excel_logger.reset_sheet(f"Zone1-Inner_{freq_suffix}_Imp_scan")
                        self.stop_increment =  True
                    self.Zone1_Inner_imp += 1

                elif zone_name == "Zone2-Mid_Inner":
                    if self.Zone2_Mid_Inner_imp >= 1:
                        excel_logger.reset_sheet(f"Zone2-Mid_Inner_{freq_suffix}_Imp_scan")
                        self.stop_increment = True
                    self.Zone2_Mid_Inner_imp += 1

                elif zone_name == "Zone3-Mid_Edge":
                    if self.Zone3_Mid_Edge_imp >= 1:
                        excel_logger.reset_sheet(f"Zone3-Mid_Edge_{freq_suffix}_Imp_scan")
                        self.stop_increment = True
                    self.Zone3_Mid_Edge_imp += 1

                elif zone_name == "Zone4-Edge":
                    if self.Zone4_Edge_imp >= 1:
                        excel_logger.reset_sheet(f"Zone4-Edge_{freq_suffix}_Imp_scan")
                        self.stop_increment = True
                    self.Zone4_Edge_imp += 1

                elif zone_name == "Zone5-Outer":
                    if self.Zone5_Outer_imp >= 1:
                        excel_logger.reset_sheet(f"Zone5-Outer_{freq_suffix}_Imp_scan")
                        self.stop_increment = True
                    self.Zone5_Outer_imp += 1

                else:
                    pass

                if hasattr(self, 'worker') and self.worker:
                    self.worker.stop()
                if selected_freq == "60 MHz":
                    script_path = '/home/robot/Manufacturing_test/aipc_beta/VNA_start_stop_60Mhz.py'
                else:
                    script_path = '/home/robot/Manufacturing_test/aipc_beta/VNA_Final.py'
                self.worker = Worker(
                    self.ssh_handler,
                    script_path,
                    command
                )








                self.worker.output_ready.connect(self.handle_Zone_impedance_output)
                self.worker.error_occurred.connect(self.handle_imp_error)
                self.worker.finished_signal.connect(lambda: self._set_tabs_locked(False))
                # Start the thread
                self.worker.start()
                self._set_tabs_locked(True)
                self.imp_idle_timer.start(self._IDLE_TIMEOUT_MS)
                self._log_Impedance_message(f"Starting measurement for {zone_name}")
            else:
                self._log_Impedance_message(f"Impedance Scan {zone_name} suspended,is_error= True")


        except Exception as e:
            self._log_Impedance_message(f"Failed to start test: {str(e)}", is_error=True)
            self.cleanup_resources()


    def _abort_scan_missing_suffix(self, log_fn, idle_timer):
        """Stop a running scan and show the animated error dialog when the
        assembly suffix is not set.  *log_fn* is the scan-specific log helper
        (e.g. self._log_Impedance_message) and *idle_timer* is the
        corresponding QTimer."""
        idle_timer.stop()
        log_fn("Test stopped: assembly suffix not set.", is_error=True)
        if hasattr(self, 'worker') and self.worker:
            self.worker.stop()

    def _check_assembly_suffix_or_abort(self) -> bool:
        """Return True if the assembly suffix is set; otherwise show the
        animated error dialog, navigate to Unit Setup, and return False."""
        if not getattr(self, 'assembly_suffix', None):
            dlg = AssemblySuffixErrorDialog(self)
            dlg.exec_()
            self.tab_widget.setCurrentIndex(0)
            return False
        return True

    def handle_imp_error(self, error_msg):
        self.imp_idle_timer.stop()
        self._log_Impedance_message(f"ERROR: {error_msg}", is_error=True)
        self.cleanup_resources()

    def handle_config_test(self):
        if hasattr(self, 'worker') and self.worker:
            self.worker.stop()

        # Raspberry Pi SSH details
        #host = "192.168.1.2"  # Replace with your Pi's IP
        host = "10.119.9.225"
        #host =
        username = "robot"
        password = "robot"  # Default, change if needed

        # File paths
        local_file = "C:\\Config\\config.json"  # Windows path
        remote_file = "/home/robot/Manufacturing_test/aipc_beta/config.json"  # Destination on Pi

        # Initialize SSH client
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(host, username=username, password=password)

        # Transfer file using SFTP
        sftp = ssh.open_sftp()
        sftp.put(local_file, remote_file)
        sftp.close()
        ssh.close()


    def file_transer(self,local_file,remote_file):
        if hasattr(self, 'worker') and self.worker:
            self.worker.stop()

        # Raspberry Pi SSH details
        #host = "192.168.1.2"  # Replace with your Pi's IP
        host = "10.119.9.225"
        username = "robot"
        password = "robot"  # Default, change if needed
        self.append_console_message("Transferring the APPOTP file to RPI")
        # File paths
        #local_file = "C:\\Config\\config.json"  # Windows path
        #remote_file = "/home/robot/Manufacturing_test/aipc_beta/config.json"  # Destination on Pi

        # Initialize SSH client
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(host, username=username, password=password)

        # Transfer file using SFTP
        sftp = ssh.open_sftp()
        sftp.put(local_file, remote_file)
        sftp.close()
        ssh.close()



    def process_single_imp_measurement(self, zone_name, measurement_line):
        """
        Process single measurement line from RPi in format: setpoint,resistance,status
        Example: "0,2.5,True"
        """
        try:
            # Remove any whitespace and split the components
            # setpoint, resistance, status = measurement_line.strip().split(',')
            val = measurement_line.strip().split(',')
            if len(val) < 6:
                self._log_Impedance_message(f"Invalid measurement format (expected 6+ fields): {measurement_line}", is_error=True)
                return False
            setpoint = val[1]
            real = val[2]
            img = val[3]
            imped = val[4]
            status = val[5]
            # Find the table for this zone
            table = self._measurement_tables_imp.get(zone_name)
            self.config2 = self.load_config(self.assembly_suffix)

            # If config could not be loaded (e.g. suffix missing), stop the test
            if self.config2 is None:
                self._abort_scan_missing_suffix(self._log_Impedance_message, self.imp_idle_timer)
                return False

            if not table:
                self._log_Impedance_message(f"No table found for zone {zone_name}", is_error=True)
                return False

            # Find the row with matching setpoint

            for row in range(table.rowCount()):
                if table.item(row, 0).text() == setpoint:
                    # Convert status to display format
                    status_text = "PASS" if status.lower() == "true" else "FAIL"
                    measurement_data = {
                        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'zone_title': zone_name,
                        'Frequency': self.freq_combo.currentText(),
                        'setpoint': setpoint,
                        'Real': float(real),
                        'Imag': float(img),
                        'Z': float(imped),
                        'status': "PASS" if status.lower() == "true" else "FAIL",
                        'table_row': row + 1
                    }


                    if status == 'False':
                        self.Impedance_test = 'FAIL'
                        self.over_all_result = 'FAIL'

                    self.step_no = self.step_no + 1
                    freq_suffix = ExcelLogger._freq_to_sheet_suffix(self.freq_combo.currentText())
                    if float(setpoint) != 0.0:
                        limit_per = float(self.config2[setpoint])
                        prod_val = float(self.config2[zone_name][setpoint])
                        higher_limit = prod_val + (prod_val * limit_per / 100)
                        lower_limit = prod_val - (prod_val * limit_per / 100)

                        # Log metadata along with test data
                        step_data_dict = {
                            'step': str(self.step_no),
                            'unit': 'ohm',
                            'low_limit': f'{lower_limit}',
                            'measure': f'{float(imped)}',
                            'high_limit': f'{higher_limit}',
                            'teststep': 'Impedance Test',
                            'testpoints': f'{zone_name}_{setpoint}_{self.freq_combo.currentText()}',
                            'status': "PASS" if status.lower() == "true" else "FAIL"
                        }
                        self.excel_logger.log_summary(step_data=step_data_dict)
                    else:
                        step_data_dict = {
                            'step': str(self.step_no),
                            'unit': 'ohm',
                            'low_limit': '0.0',
                            'measure': f'{float(imped)}',
                            'high_limit': '0.7',
                            'teststep': 'Impedance Test',
                            'testpoints': f'{zone_name}_{setpoint}_{self.freq_combo.currentText()}',
                            'status': "PASS" if status.lower() == "true" else "FAIL"
                        }
                        self.excel_logger.log_summary(step_data=step_data_dict)

                    #if zone_name == "Zone5-Outer" and float(setpoint) == 143.0:
                    """
                    self.excel_logger.log_summary(
                            teststep_data={
                                'teststep': 'Impedance Test',
                                'status': self.Impedance_test
                            })

                    self.excel_logger.log_summary(
                            metadata={
                                'end_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                'overall_result': self.over_all_result
                            }
                        )
                    """
                    self.excel_logger.update_overall_result(self.over_all_result, finalize=False)
                    # Update the table
                    self.update_impedance_measurement(
                        zone_name,
                        row,
                        float(real),
                        float(img),
                        float(imped),
                        status_text
                    )
                    # Log to Excel (frequency-specific combined sheet)
                    if not self.excel_logger.log_Imp_measurement(measurement_data, f'{zone_name}_{freq_suffix}_Imp_scan'):
                        raise Exception("Failed to log to Excel")
                    return True

            self._log_Impedance_message(f"Setpoint {setpoint} not found in {zone_name}", is_error=True)
            return False

        except ValueError:
            self._log_Impedance_message(f"Invalid measurement format: {measurement_line}", is_error=True)
            return False
        except Exception as e:
            self._log_Impedance_message(f"Error processing measurement: {str(e)}", is_error=True)
            return False

    def _log_Impedance_message(self, message, is_error=False):
        if is_error:
            self._log_output_imp.append(f'<span style="color:#f85149; font-weight:bold;">{message}</span>')
        else:
            self._log_output_imp.append(f'<span style="color:#3fb950; font-weight:bold;">{message}</span>')

        self._log_output_imp.verticalScrollBar().setValue(
            self._log_output_imp.verticalScrollBar().maximum()
        )

    def handle_Zone_impedance_output(self, line):
        # Reset the 5-minute idle watchdog on every received line
        self.imp_idle_timer.start(self._IDLE_TIMEOUT_MS)

        # self._log_resistance_message(f"{line}")
        # self._log_Impedance_message(f"{line}")
        # self._log_Impedance_message(self.names1)
        if "pyvisa.errors" in line:
            QMessageBox.critical(
                self,
                "PyVISA Error",
                f"A PyVISA error occurred: {line.strip()}"
            )
            self.imp_idle_timer.stop()
            self.worker.stop()
            return

        if "No data found for frequencies" in line:
            self._log_Impedance_message(line.strip(), is_error=True)
            self.imp_idle_timer.stop()
            self.worker.stop()
            return

        if "mailbox error" in line.lower():
            self._log_Impedance_message(
                "Mailbox Error on Ethercat please check the Ethercat Data or Contact Support Team",
                is_error=True
            )
            self.imp_idle_timer.stop()
            self.worker.stop()
            return

        if self.names1 in line:
            self._log_Impedance_message(f"Starting measurement for {line}")

            # Process the single measurement line
            if line.strip():
                if not self.process_single_imp_measurement(self.names1, line):
                    self._log_Impedance_message("Failed to process measurement", is_error=True)
            else:
                self._log_Impedance_message("No measurement data received", is_error=True)

        if "Error in slave initialization" in line:
            QMessageBox.critical(
                self,
                "Critical Error",
                "Slave initialization failed! Please check the EtherCAT connection and restart the test."
            )
            self.imp_idle_timer.stop()
            self.worker.stop()

        if "no ping" in line:
            self._log_Impedance_message(f"VNA not connected to the Network", is_error= True)
            self.imp_idle_timer.stop()
            self.worker.stop()

        if "Test_done" in line:
            self._log_Impedance_message(f"============{self.names1} test completed =====")
            self.imp_idle_timer.stop()
            self.worker.stop()

    def _on_imp_idle_timeout(self):
        """Called when no output line has been received from the impedance script for 2 minutes."""
        self._log_Impedance_message(
            "=== ERROR: No data from Raspberry Pi — EtherCAT data broken. Please contact support team. ===",
            is_error=True,
        )
        self._show_idle_timeout_error("Impedance Scan")
        self.worker.stop()
        self._set_tabs_locked(False)

    def update_impedance_measurement(self, zone_name, row_index, real, imag, Imp, status):
        """Update a single measurement in the table"""
        if zone_name in self._measurement_tables_imp:
            table = self._measurement_tables_imp[zone_name]
            if 0 <= row_index < table.rowCount():
                # Update resistance value (3 decimal places)
                real_item = QTableWidgetItem(f"{real:.2f}")
                real_item.setFlags(real_item.flags() ^ Qt.ItemIsEditable)
                table.setItem(row_index, 1, real_item)

                image_item = QTableWidgetItem(f"{imag:.2f}")
                image_item.setFlags(image_item.flags() ^ Qt.ItemIsEditable)
                table.setItem(row_index, 2, image_item)

                Imp_item = QTableWidgetItem(f"{Imp:.2f}")
                Imp_item.setFlags(Imp_item.flags() ^ Qt.ItemIsEditable)
                table.setItem(row_index, 3, Imp_item)

                # Update status with appropriate coloring
                status_item = QTableWidgetItem(status)
                status_item.setFlags(status_item.flags() ^ Qt.ItemIsEditable)
                # Bold font so PASS/FAIL is clearly readable on the dark background
                status_item.setFont(QFont('Arial', 9, QFont.Bold))

                if status == "PASS":
                    status_item.setBackground(QColor(10, 30, 10))    # Dark green tint
                    status_item.setForeground(QColor(0, 230, 118))   # Bright green text
                else:
                    status_item.setBackground(QColor(30, 10, 10))    # Dark red tint
                    status_item.setForeground(QColor(255, 82, 82))   # Bright red text

                table.setItem(row_index, 4, status_item)

                # Scroll to show the updated row
                table.scrollToItem(table.item(row_index, 0))

    def _start_resistance_zone_measurement(self, zone_name):
        try:
            # Guard: abort immediately if assembly suffix is not set
            if not self._check_assembly_suffix_or_abort():
                return

            self.names = zone_name
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText(f"Please connect Zone {zone_name} and click OK to continue")
            msg.setWindowTitle("Resistance Scan")
            msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
            retval = msg.exec_()
            if retval == QMessageBox.Ok:
                self.start_time = time.time()
                if zone_name == "Zone1-Inner":
                    if self.Zone1_Inner_res >= 1:
                        excel_logger.reset_sheet("Zone1-Inner_Res_scan")
                    self.Zone1_Inner_res += 1

                elif zone_name == "Zone2-Mid_Inner":
                    if self.Zone2_Mid_Inner_res >= 1:
                        excel_logger.reset_sheet("Zone2-Mid_Inner_Res_scan")
                    self.Zone2_Mid_Inner_res += 1

                elif zone_name == "Zone3-Mid_Edge":
                    if self.Zone3_Mid_Edge_res >= 1:
                        excel_logger.reset_sheet("Zone3-Mid_Edge_Res_scan")
                    self.Zone3_Mid_Edge_res += 1

                elif zone_name == "Zone4-Edge":
                    if self.Zone4_Edge_res >= 1:
                        excel_logger.reset_sheet("Zone4-Edge_Res_scan")
                    self.Zone4_Edge_res += 1

                elif zone_name == "Zone5-Outer":
                    if self.Zone5_Outer_res >= 1:
                        excel_logger.reset_sheet("Zone5-Outer_Res_scan")
                    self.Zone5_Outer_res += 1

                else:
                    pass
                if hasattr(self, 'worker') and self.worker:
                    self.worker.stop()
                self.worker = Worker(
                    self.ssh_handler,
                    '/home/robot/Manufacturing_test/aipc_beta/resistance_test.py',
                    zone_name
                )

                self.worker.output_ready.connect(self.handle_Zone_output)
                self.worker.error_occurred.connect(self.handle_res_error)
                self.worker.finished_signal.connect(lambda: self._set_tabs_locked(False))
                # Start the thread
                self.worker.start()
                self._set_tabs_locked(True)
                self.res_idle_timer.start(self._IDLE_TIMEOUT_MS)
                self._log_resistance_message(f"Starting measurement for {zone_name}")

            else:
                self._log_resistance_message(f"Resistance Scan : {zone_name} suspended", is_error=True)

        except Exception as e:
            self._log_resistance_message(f"Failed to start test: {str(e)}", is_error=True)
            self.cleanup_resources()


    def handle_res_error(self, error_msg):
        self.res_idle_timer.stop()
        self._log_resistance_message(f"ERROR: {error_msg}", is_error=True)
        self.cleanup_resources()


    def handle_Zone_output(self, line):
        # Reset the 5-minute idle watchdog on every received line
        self.res_idle_timer.start(self._IDLE_TIMEOUT_MS)

        # self._log_resistance_message(f"{line}")
        if "pyvisa.errors" in line:
            QMessageBox.critical(
                self,
                "PyVISA Error",
                f"A PyVISA error occurred: {line.strip()}"
            )
            self.res_idle_timer.stop()
            self.worker.stop()
            return

        if "mailbox error" in line.lower():
            self._log_resistance_message(
                "Mailbox Error on Ethercat please check the Ethercat Data or Contact Support Team",
                is_error=True
            )
            self.res_idle_timer.stop()
            self.worker.stop()
            return

        if "no ping" in line:
            self._log_resistance_message(f"Device not connected to the Network", is_error=True)
            self.res_idle_timer.stop()
            self.worker.stop()

        if "Error in slave initialization" in line:
            QMessageBox.critical(
                self,
                "Critical Error",
                "Slave initialization failed! Please check the EtherCAT connection and restart the test."
            )
            self.res_idle_timer.stop()
            self.worker.stop()

        if self.names in line:
            self._log_resistance_message(f"Starting measurement for {line}")

            # Process the single measurement line
            if line.strip():
                if not self.process_single_measurement(self.names, line):
                    self._log_resistance_message("Failed to process measurement", is_error=True)
            else:
                self._log_resistance_message("No measurement data received", is_error=True)
        if "Test_done" in line:
            self._log_resistance_message(f"============{self.names} test completed =====")
            self.res_idle_timer.stop()
            self.worker.stop()

    def _on_res_idle_timeout(self):
        """Called when no output line has been received from the resistance script for 2 minutes."""
        self._log_resistance_message(
            "=== ERROR: No data from Raspberry Pi — EtherCAT data broken. Please contact support team. ===",
            is_error=True,
        )
        self._show_idle_timeout_error("Resistance Scan")
        self.worker.stop()
        self._set_tabs_locked(False)


    def update_resistance_measurement(self, zone_name, row_index, resistance_value, status):
        """Update a single measurement in the table"""
        if zone_name in self._measurement_tables:
            table = self._measurement_tables[zone_name]
            if 0 <= row_index < table.rowCount():
                # Update resistance value (3 decimal places)
                resistance_item = QTableWidgetItem(f"{resistance_value:.3f}")
                resistance_item.setFlags(resistance_item.flags() ^ Qt.ItemIsEditable)
                table.setItem(row_index, 1, resistance_item)

                # Update status with appropriate coloring
                status_item = QTableWidgetItem(status)
                status_item.setFlags(status_item.flags() ^ Qt.ItemIsEditable)
                # Bold font so PASS/FAIL is clearly readable on the dark background
                status_item.setFont(QFont('Arial', 9, QFont.Bold))

                if status == "PASS":
                    status_item.setBackground(QColor(10, 30, 10))    # Dark green tint
                    status_item.setForeground(QColor(0, 230, 118))   # Bright green text
                else:
                    status_item.setBackground(QColor(30, 10, 10))    # Dark red tint
                    status_item.setForeground(QColor(255, 82, 82))   # Bright red text

                table.setItem(row_index, 2, status_item)

                # Scroll to show the updated row
                table.scrollToItem(table.item(row_index, 0))


    def _log_resistance_message(self, message, is_error=False):
        if is_error:
            self._log_output.append(f'<span style="color:#f85149; font-weight:bold;">{message}</span>')
        else:
            self._log_output.append(f'<span style="color:#3fb950; font-weight:bold;">{message}</span>')
        self._log_output.verticalScrollBar().setValue(
            self._log_output.verticalScrollBar().maximum()
        )


    # ================================================================== #
    #  Shared design-system constants & helpers                           #
    # ================================================================== #

    _DARK_CONSOLE_SS = """
        QTextBrowser {
            background-color: #0d1117;
            color: #c9d1d9;
            border: 1px solid #30363d;
            border-top: none;
            border-bottom-left-radius: 4px;
            border-bottom-right-radius: 4px;
            font-family: 'Courier New', Consolas, monospace;
            font-size: 10pt;
            padding: 6px;
            selection-background-color: #264f78;
        }
    """

    _CONSOLE_HDR_SS = """
        QLabel {
            background-color: #343a40;
            color: #adb5bd;
            font-size: 8pt;
            font-weight: bold;
            padding: 4px 8px;
            border-top-left-radius: 4px;
            border-top-right-radius: 4px;
        }
    """

    _PILL_READY_SS = (
        "QLabel { background-color:#17a2b8; color:white; padding:6px 14px;"
        " border-radius:14px; font-weight:bold; font-size:10pt; }"
    )
    _PILL_RUN_SS = (
        "QLabel { background-color:#fd7e14; color:white; padding:6px 14px;"
        " border-radius:14px; font-weight:bold; font-size:10pt; }"
    )
    _PILL_PASS_SS = (
        "QLabel { background-color:#28a745; color:white; padding:6px 14px;"
        " border-radius:14px; font-weight:bold; font-size:10pt; }"
    )
    _PILL_FAIL_SS = (
        "QLabel { background-color:#dc3545; color:white; padding:6px 14px;"
        " border-radius:14px; font-weight:bold; font-size:10pt; }"
    )
    _PILL_GRAY_SS = (
        "QLabel { background-color:#6c757d; color:white; padding:6px 14px;"
        " border-radius:14px; font-weight:bold; font-size:10pt; }"
    )

    _BTN_GREEN_SS = """
        QPushButton {
            background-color: #28a745; color: white; border: none;
            border-radius: 5px; font-size: 10pt; font-weight: bold; padding: 6px 18px;
        }
        QPushButton:hover    { background-color: #218838; }
        QPushButton:pressed  { background-color: #1e7e34; }
        QPushButton:disabled { background-color: #94d3a2; color: #e9f7ed; }
    """
    _BTN_RED_SS = """
        QPushButton {
            background-color: #dc3545; color: white; border: none;
            border-radius: 5px; font-size: 10pt; font-weight: bold; padding: 6px 18px;
        }
        QPushButton:hover    { background-color: #c82333; }
        QPushButton:pressed  { background-color: #bd2130; }
        QPushButton:disabled { background-color: #e8a5ac; color: #fce0e3; }
    """
    _BTN_TEAL_SS = """
        QPushButton {
            background-color: #17a2b8; color: white; border: none;
            border-radius: 5px; font-size: 10pt; font-weight: bold; padding: 6px 18px;
        }
        QPushButton:hover    { background-color: #138496; }
        QPushButton:pressed  { background-color: #117a8b; }
        QPushButton:disabled { background-color: #8bd4df; color: #d9f3f7; }
    """
    _BTN_GRAY_SS = """
        QPushButton {
            background-color: #6c757d; color: white; border: none;
            border-radius: 5px; font-size: 10pt; font-weight: bold; padding: 6px 18px;
        }
        QPushButton:hover    { background-color: #5a6268; }
        QPushButton:pressed  { background-color: #4e555b; }
        QPushButton:disabled { background-color: #adb5bd; color: #e9ecef; }
    """
    _BTN_PURPLE_SS = """
        QPushButton {
            background-color: #6f42c1; color: white; border: none;
            border-radius: 5px; font-size: 10pt; font-weight: bold; padding: 6px 18px;
        }
        QPushButton:hover    { background-color: #5e35b1; }
        QPushButton:pressed  { background-color: #512da8; }
        QPushButton:disabled { background-color: #c3a8e8; color: #f1ebfc; }
    """

    _PROGRESS_SS = """
        QProgressBar {
            border: 1px solid #adb5bd; border-radius: 5px; background-color: #e9ecef;
            text-align: center; font-size: 8pt; color: #343a40;
        }
        QProgressBar::chunk {
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 #28a745, stop:1 #20c997);
            border-radius: 5px;
        }
    """

    _ZONE_PANEL_SS = """
        QGroupBox {
            font-weight: bold; font-size: 9pt; color: #fff;
            border: none; border-radius: 6px;
            margin-top: 8px; padding-top: 14px;
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 #1565C0, stop:1 #0d47a1);
        }
        QGroupBox::title {
            subcontrol-origin: margin; subcontrol-position: top center;
            padding: 0 8px; color: white;
        }
    """

    _STATE_IDLE_SS = """
        QLabel {
            background-color: #e9ecef; color: #6c757d;
            border: 2px solid #ced4da; border-radius: 6px;
            font-size: 11pt; font-weight: bold; padding: 10px 20px;
        }
    """
    _STATE_OPEN_SS = """
        QLabel {
            background-color: #dc3545; color: white;
            border: 2px solid #c82333; border-radius: 6px;
            font-size: 11pt; font-weight: bold; padding: 10px 20px;
        }
    """
    _STATE_CLOSED_SS = """
        QLabel {
            background-color: #28a745; color: white;
            border: 2px solid #218838; border-radius: 6px;
            font-size: 11pt; font-weight: bold; padding: 10px 20px;
        }
    """

    def _make_tab_header(self, title, subtitle, color1="#1565C0", color2="#0288D1"):
        """Return a styled gradient header QFrame for any tab."""
        header = QFrame()
        header.setMinimumHeight(52)
        header.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {color1}, stop:1 {color2});
                border-radius: 6px;
            }}
        """)
        hdr_l = QVBoxLayout(header)
        hdr_l.setContentsMargins(16, 8, 16, 8)
        hdr_l.setSpacing(2)
        t = QLabel(title)
        t.setWordWrap(True)
        t.setStyleSheet(
            "color: white; font-size: 14pt; font-weight: bold; background: transparent;"
        )
        hdr_l.addWidget(t)
        return header

    def _make_status_pill(self, text="● Test: Ready"):
        """Return a teal 'Ready' status pill label."""
        lbl = QLabel(text)
        lbl.setAlignment(Qt.AlignCenter)
        lbl.setMinimumWidth(170)
        lbl.setStyleSheet(self._PILL_READY_SS)
        return lbl

    def _make_action_button(self, text, style_ss, min_height=38, min_width=150):
        """Return a consistently styled action button."""
        btn = QPushButton(text)
        btn.setMinimumHeight(min_height)
        btn.setMinimumWidth(min_width)
        btn.setCursor(Qt.PointingHandCursor)
        btn.setStyleSheet(style_ss)
        return btn

    def _make_controls_row(self, pill, *buttons):
        """Return an QHBoxLayout: pill | stretch | buttons..."""
        row = QHBoxLayout()
        row.setSpacing(12)
        row.addWidget(pill)
        row.addStretch()
        for btn in buttons:
            row.addWidget(btn)
        return row

    def _make_styled_progress(self, fmt="%p%", rng=(0, 100)):
        """Return a gradient styled QProgressBar."""
        pb = QProgressBar()
        pb.setRange(*rng)
        pb.setValue(0)
        pb.setFormat(fmt)
        pb.setMinimumHeight(22)
        pb.setStyleSheet(self._PROGRESS_SS)
        return pb

    def _make_console_header(self, text="  Test Output"):
        """Return the dark bar label placed above the dark console."""
        lbl = QLabel(text)
        lbl.setStyleSheet(self._CONSOLE_HDR_SS)
        return lbl

    def _make_dark_console(self, min_height=450, max_height=1500):
        """Return a styled dark QTextBrowser terminal widget."""
        c = QTextBrowser()
        c.setMinimumHeight(min_height)
        c.setMaximumHeight(max_height)
        c.setStyleSheet(self._DARK_CONSOLE_SS)
        return c

    # ================================================================== #

    def create_test_tab(self, title, show_progress=False):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)

        test_section = QGroupBox()
        test_section.setStyleSheet("""
                QGroupBox {
                    background-color: white;
                    padding: 15px;
                    border-radius: 5px;
                }
                QGroupBox::title {
                    font-weight: bold;
                    font-size: 16px;
                    subcontrol-origin: margin;
                    left: 3px;
                    padding: 0 3px;
                }
            """)
        test_layout = QVBoxLayout(test_section)
        controls = QWidget()
        controls_layout = QHBoxLayout(controls)
        controls_layout.setContentsMargins(0, 0, 0, 0)

        if title == "Interlock System Check":
            # ── Header banner ──────────────────────────────────────────────
            test_layout.addWidget(self._make_tab_header(
                "Interlock System Check",
                "Verifies fan interlock and switch interlock state transitions (Open / Closed).",
                "#b71c1c", "#e53935"
            ))

            # ── Controls row ───────────────────────────────────────────────
            self.test_status_label = self._make_status_pill("● Test: Ready")
            self.interlock_start_button = self._make_action_button(
                "▶  Start Test", self._BTN_GREEN_SS
            )
            self.interlock_start_button.clicked.connect(self.start_interlock_test)
            self.interlock_end_button = self._make_action_button(
                "■  End Test", self._BTN_RED_SS
            )
            self.interlock_end_button.clicked.connect(self.end_interlock_test)
            self.interlock_end_button.setEnabled(False)
            test_layout.addLayout(
                self._make_controls_row(
                    self.test_status_label,
                    self.interlock_start_button,
                    self.interlock_end_button,
                )
            )

            # ── Interlock state indicators ─────────────────────────────────
            interlock_state_group = QGroupBox("Switch State")
            interlock_state_group.setStyleSheet("""
                QGroupBox {
                    font-weight: bold; font-size: 9pt; color: #333;
                    border: 1px solid #ced4da; border-radius: 6px;
                    margin-top: 8px; padding-top: 10px;
                    background-color: #f8f9fa;
                }
                QGroupBox::title {
                    subcontrol-origin: margin; subcontrol-position: top left;
                    left: 12px; padding: 0 6px;
                }
            """)
            state_outer = QHBoxLayout(interlock_state_group)
            state_outer.setSpacing(12)
            state_outer.setContentsMargins(12, 14, 12, 10)

            self.interlock_status_layout = QHBoxLayout()
            self.interlock_open_label = QLabel("🔓  OPEN")
            self.interlock_open_label.setAlignment(Qt.AlignCenter)
            self.interlock_open_label.setMinimumHeight(54)
            self.interlock_open_label.setStyleSheet(self._STATE_IDLE_SS)

            self.interlock_closed_label = QLabel("🔒  CLOSED")
            self.interlock_closed_label.setAlignment(Qt.AlignCenter)
            self.interlock_closed_label.setMinimumHeight(54)
            self.interlock_closed_label.setStyleSheet(self._STATE_IDLE_SS)

            state_outer.addWidget(self.interlock_open_label)
            state_outer.addWidget(self.interlock_closed_label)
            self.interlock_status_layout.addWidget(interlock_state_group)
            test_layout.addLayout(self.interlock_status_layout)

            # ── Console ────────────────────────────────────────────────────
            test_layout.addWidget(self._make_console_header("  Interlock Test Output"))
            self.interlock_console = self._make_dark_console(min_height=600)
            test_layout.addWidget(self.interlock_console)

        elif title == "System Self Test":
            # ── Header banner ──────────────────────────────────────────────
            test_layout.addWidget(self._make_tab_header(
                "System Self Test",
                "Executes a full end-to-end EtherCAT self-diagnostics pass on the DUT.",
                "#1B5E20", "#388E3C"
            ))

            # ── Controls row ───────────────────────────────────────────────
            self.test_status_label_start = self._make_status_pill("● Test: Ready")
            self.self_start_button = self._make_action_button(
                "▶  Start Test", self._BTN_GREEN_SS
            )
            self.self_start_button.clicked.connect(self.start_self_test)
            test_layout.addLayout(
                self._make_controls_row(self.test_status_label_start, self.self_start_button)
            )

            # ── Console ────────────────────────────────────────────────────
            test_layout.addWidget(self._make_console_header("  Self Test Output"))
            self.selftest_console = self._make_dark_console()
            test_layout.addWidget(self.selftest_console)

        elif title == "Impedance Scan":
            # ── Header banner ──────────────────────────────────────────────
            impedance_widget = QWidget()
            impedance_layout = QVBoxLayout(impedance_widget)
            impedance_layout.setContentsMargins(0, 0, 0, 0)
            impedance_layout.setSpacing(8)

            impedance_layout.addWidget(self._make_tab_header(
                "Impedance Scan",
                "Measures complex impedance (R + jX) at selected frequency across all 5 RF zones.",
                "#004D40", "#00897B"
            ))

            # ── Control panel ──────────────────────────────────────────────
            control_panel = QGroupBox("Test Parameters")
            control_panel.setStyleSheet("""
                QGroupBox {
                    font-weight: bold; font-size: 9pt; color: #333;
                    border: 1px solid #ced4da; border-radius: 6px;
                    margin-top: 8px; padding-top: 10px; background-color: #f8f9fa;
                }
                QGroupBox::title {
                    subcontrol-origin: margin; subcontrol-position: top left;
                    left: 12px; padding: 0 6px;
                }
                QLabel { font-weight: bold; color: #333; }
                QComboBox {
                    background-color: white; border: 1px solid #ced4da;
                    padding: 4px 8px; font-weight: bold; color: #222; min-width: 120px;
                    border-radius: 4px;
                }
                QComboBox:hover { border-color: #0288D1; }
            """)
            control_layout = QHBoxLayout(control_panel)
            control_layout.setContentsMargins(12, 8, 12, 8)

            self.frequencies = ["362.3 KHz", "400 KHz", "500 KHz", "50 MHz", "60 MHz", "70 MHz"]
            freq_label = QLabel("Test Frequency:")
            freq_label.setFont(QFont('Arial', 10))
            self.freq_combo = QComboBox()
            self.freq_combo.addItems(self.frequencies)
            self.freq_combo.setCurrentIndex(0)
            self.freq_combo.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            control_layout.addWidget(freq_label)
            control_layout.addWidget(self.freq_combo)
            control_layout.addStretch()
            impedance_layout.addWidget(control_panel)

            # ── Zone panels ────────────────────────────────────────────────
            self._zone_names_imp = ["Zone1-Inner", "Zone2-Mid_Inner", "Zone3-Mid_Edge", "Zone4-Edge", "Zone5-Outer"]
            self._relay_values_imp = [0, 1, 2, 4, 8, 16, 32, 64, 127, 128, 135, 141, 142, 143]
            self._measurement_tables_imp = {}

            zones_scroll = QScrollArea()
            zones_scroll.setWidgetResizable(True)
            zones_container = QWidget()
            zones_layout = QHBoxLayout(zones_container)
            zones_layout.setContentsMargins(5, 5, 5, 5)
            zones_layout.setSpacing(10)
            for zone in self._zone_names_imp:
                zone_panel = self._create_Impedance_zone_panel(zone)
                zone_panel.setMinimumWidth(250)
                zones_layout.addWidget(zone_panel)
            zones_scroll.setWidget(zones_container)
            impedance_layout.addWidget(zones_scroll, stretch=1)

            # ── Measurement log ────────────────────────────────────────────
            impedance_layout.addWidget(self._make_console_header("  Measurement Log"))
            self._log_output_imp = self._make_dark_console(min_height=160, max_height=400)
            impedance_layout.addWidget(self._log_output_imp)

            clear_btn_imp = self._make_action_button("Clear Log", self._BTN_GRAY_SS,
                                                     min_height=30, min_width=100)
            clear_btn_imp.clicked.connect(self._clear_impedance_log_display)
            clear_btn_imp.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            clear_row = QHBoxLayout()
            clear_row.addStretch()
            clear_row.addWidget(clear_btn_imp)
            impedance_layout.addLayout(clear_row)

            test_layout.addWidget(impedance_widget)

        elif title == "Resistance Test":
            # ── Header banner ──────────────────────────────────────────────
            resistance_widget = QWidget()
            resistance_layout = QVBoxLayout(resistance_widget)
            resistance_layout.setContentsMargins(0, 0, 0, 0)
            resistance_layout.setSpacing(8)

            resistance_layout.addWidget(self._make_tab_header(
                "Resistance Test",
                "Measures DC resistance (Ω) across relay setpoints for each of the 5 RF zones.",
                "#311B92", "#4527A0"
            ))

            # ── Zone panels ────────────────────────────────────────────────
            self._zone_names = ["Zone1-Inner", "Zone2-Mid_Inner", "Zone3-Mid_Edge", "Zone4-Edge", "Zone5-Outer"]
            self._relay_values = [0, 1, 2, 4, 8, 16, 32, 64, 127]
            self._measurement_tables = {}

            zones_scroll = QScrollArea()
            zones_scroll.setWidgetResizable(True)
            zones_container = QWidget()
            zones_layout = QHBoxLayout(zones_container)
            zones_layout.setContentsMargins(5, 5, 5, 5)
            zones_layout.setSpacing(10)
            for zone in self._zone_names:
                zone_panel = self._create_resistance_zone_panel(zone)
                zone_panel.setMinimumWidth(250)
                zones_layout.addWidget(zone_panel)
            zones_scroll.setWidget(zones_container)
            resistance_layout.addWidget(zones_scroll, stretch=1)

            # ── Measurement log ────────────────────────────────────────────
            resistance_layout.addWidget(self._make_console_header("  Measurement Log"))
            self._log_output = self._make_dark_console(min_height=160, max_height=400)
            resistance_layout.addWidget(self._log_output)

            clear_btn_res = self._make_action_button("Clear Log", self._BTN_GRAY_SS,
                                                     min_height=30, min_width=100)
            clear_btn_res.clicked.connect(self._clear_resistance_log_display)
            clear_btn_res.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            clear_row = QHBoxLayout()
            clear_row.addStretch()
            clear_row.addWidget(clear_btn_res)
            resistance_layout.addLayout(clear_row)

            test_layout.addWidget(resistance_widget)

        elif title == "VNA Calibration":
            # ── Header banner ──────────────────────────────────────────────
            test_layout.addWidget(self._make_tab_header(
                "VNA Calibration",
                "Performs electronic calibration of the vector network analyser (E-Cal module required).",
                "#880E4F", "#AD1457"
            ))

            # ── Controls row ───────────────────────────────────────────────
            self.VNA_status_label_start = self._make_status_pill("● Test: Ready")
            self.VNA_start_button = self._make_action_button(
                "▶  Start Calibration", self._BTN_GREEN_SS
            )
            self.VNA_start_button.clicked.connect(self.VNA_cal_test)
            test_layout.addLayout(
                self._make_controls_row(self.VNA_status_label_start, self.VNA_start_button)
            )

            # ── Progress bar ───────────────────────────────────────────────
            self.vna_progress = self._make_styled_progress("%p% complete")
            test_layout.addWidget(self.vna_progress)

            # ── Console ────────────────────────────────────────────────────
            test_layout.addWidget(self._make_console_header("  Calibration Output"))
            self.VNAtest_console = self._make_dark_console()
            test_layout.addWidget(self.VNAtest_console)

        elif title == "Verify BNC Port":
            # ================================================================
            # BNC Port Verification – UI layout
            # ================================================================

            # ── Header banner ──────────────────────────────────────────────
            header = QFrame()
            header.setMinimumHeight(52)
            header.setStyleSheet("""
                QFrame {
                    background: qlineargradient(
                        x1:0, y1:0, x2:1, y2:0,
                        stop:0 #1565C0, stop:1 #0288D1
                    );
                    border-radius: 6px;
                }
            """)
            header_layout = QVBoxLayout(header)
            header_layout.setContentsMargins(16, 8, 16, 8)
            header_layout.setSpacing(2)

            header_title = QLabel("BNC Port Verification")
            header_title.setWordWrap(True)
            header_title.setStyleSheet(
                "color: white; font-size: 14pt; font-weight: bold; background: transparent;"
            )
            header_layout.addWidget(header_title)
            test_layout.addWidget(header)

            # ── Controls row (status pill + Start button) ──────────────────
            controls_row = QHBoxLayout()
            controls_row.setSpacing(12)

            self.BNC_status_label_start = QLabel("● Test: Ready")
            self.BNC_status_label_start.setAlignment(Qt.AlignCenter)
            self.BNC_status_label_start.setMinimumWidth(160)
            self.BNC_status_label_start.setStyleSheet("""
                QLabel {
                    background-color: #17a2b8;
                    color: white;
                    padding: 6px 14px;
                    border-radius: 14px;
                    font-weight: bold;
                    font-size: 10pt;
                }
            """)

            self.BNC_start_button = QPushButton("▶  Start Test")
            self.BNC_start_button.setMinimumHeight(38)
            self.BNC_start_button.setMinimumWidth(150)
            self.BNC_start_button.setCursor(Qt.PointingHandCursor)
            self.BNC_start_button.setStyleSheet("""
                QPushButton {
                    background-color: #28a745;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    font-size: 10pt;
                    font-weight: bold;
                    padding: 6px 18px;
                }
                QPushButton:hover    { background-color: #218838; }
                QPushButton:pressed  { background-color: #1e7e34; }
                QPushButton:disabled { background-color: #94d3a2; color: #e9f7ed; }
            """)
            self.BNC_start_button.clicked.connect(self.BNC_test)

            controls_row.addWidget(self.BNC_status_label_start)
            controls_row.addStretch()
            controls_row.addWidget(self.BNC_start_button)
            test_layout.addLayout(controls_row)

            # ── Zone status panel ──────────────────────────────────────────
            zones_group = QGroupBox("Zone Status")
            zones_group.setStyleSheet("""
                QGroupBox {
                    font-weight: bold;
                    font-size: 9pt;
                    color: #333;
                    border: 1px solid #ced4da;
                    border-radius: 6px;
                    margin-top: 8px;
                    padding-top: 10px;
                    background-color: #f8f9fa;
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    subcontrol-position: top left;
                    left: 12px;
                    padding: 0 6px;
                }
            """)
            zones_outer = QVBoxLayout(zones_group)
            zones_outer.setSpacing(8)
            zones_outer.setContentsMargins(10, 12, 10, 10)

            # Four zone indicator boxes — subtitles from the class-level mapping
            zone_row = QHBoxLayout()
            zone_row.setSpacing(10)
            ZONE_DEFS = [
                (znum, f"Zone {znum}", subtitle)
                for znum, subtitle in self._BNC_ZONE_SUBTITLES.items()
            ]
            self.BNC_zone_labels = {}
            for znum, zname, zdesc in ZONE_DEFS:
                lbl = QLabel(f"⬜  {zname}\n{zdesc}")
                lbl.setAlignment(Qt.AlignCenter)
                lbl.setMinimumHeight(56)
                lbl.setStyleSheet("""
                    QLabel {
                        background-color: #e9ecef;
                        color: #6c757d;
                        border: 2px solid #ced4da;
                        border-radius: 6px;
                        font-size: 9pt;
                        font-weight: bold;
                        padding: 4px 6px;
                    }
                """)
                self.BNC_zone_labels[znum] = lbl
                zone_row.addWidget(lbl)
            zones_outer.addLayout(zone_row)

            # Progress bar inside the zone panel
            self.bnc_progress_bar = QProgressBar()
            self.bnc_progress_bar.setRange(0, 4)
            self.bnc_progress_bar.setValue(0)
            self.bnc_progress_bar.setFormat("%v / 4 zones complete")
            self.bnc_progress_bar.setMinimumHeight(22)
            self.bnc_progress_bar.setStyleSheet("""
                QProgressBar {
                    border: 1px solid #adb5bd;
                    border-radius: 5px;
                    background-color: #e9ecef;
                    text-align: center;
                    font-size: 8pt;
                    color: #343a40;
                }
                QProgressBar::chunk {
                    background: qlineargradient(
                        x1:0, y1:0, x2:1, y2:0,
                        stop:0 #28a745, stop:1 #20c997
                    );
                    border-radius: 5px;
                }
            """)
            zones_outer.addWidget(self.bnc_progress_bar)
            test_layout.addWidget(zones_group)

            # ── Console header label ───────────────────────────────────────
            console_header = QLabel("  Test Output")
            console_header.setStyleSheet("""
                QLabel {
                    background-color: #343a40;
                    color: #adb5bd;
                    font-size: 8pt;
                    font-weight: bold;
                    padding: 4px 8px;
                    border-top-left-radius: 4px;
                    border-top-right-radius: 4px;
                }
            """)
            test_layout.addWidget(console_header)

            # ── Dark terminal console ──────────────────────────────────────
            # Minimum height is 400 px (reduced from 500) to give room to the
            # new zone status panel above while still showing ample output.
            self.BNCtest_console = QTextBrowser()
            self.BNCtest_console.setMinimumHeight(400)
            self.BNCtest_console.setMaximumHeight(1500)
            self.BNCtest_console.setStyleSheet("""
                QTextBrowser {
                    background-color: #0d1117;
                    color: #c9d1d9;
                    border: 1px solid #30363d;
                    border-top: none;
                    border-bottom-left-radius: 4px;
                    border-bottom-right-radius: 4px;
                    font-family: 'Courier New', Consolas, monospace;
                    font-size: 10pt;
                    padding: 6px;
                    selection-background-color: #264f78;
                }
            """)
            test_layout.addWidget(self.BNCtest_console)


        else:
            pass

        layout.addWidget(test_section)
        return tab


    def append_interlock_message(self, message, is_error=False):
        """Helper method to append colored messages to interlock console"""
        if is_error:
            self.interlock_console.append(f'<span style="color:#f85149; font-weight:bold;">{message}</span>')
        else:
            self.interlock_console.append(f'<span style="color:#3fb950; font-weight:bold;">{message}</span>')
        # Auto-scroll to bottom
        self.interlock_console.verticalScrollBar().setValue(
            self.interlock_console.verticalScrollBar().maximum()
        )


    def append_self_message(self, message, is_error=False):
        """Helper method to append colored messages to interlock console"""
        if is_error:
            self.selftest_console.append(f'<span style="color:#f85149; font-weight:bold;">{message}</span>')
        else:
            self.selftest_console.append(f'<span style="color:#3fb950; font-weight:bold;">{message}</span>')
        # Auto-scroll to bottom
        self.selftest_console.verticalScrollBar().setValue(
            self.selftest_console.verticalScrollBar().maximum()
        )


    def start_self_test(self):
        try:
            if self.self_t >= 1:
                excel_logger.reset_sheet("Self Test")
            self.self_t += 1
            self.selftest_console.clear()
            self._self_test_lines = []

            # Clean up any previous self-test worker
            if self._self_test_worker is not None:
                try:
                    self._self_test_worker.output_ready.disconnect(self._on_selftest_line)
                    self._self_test_worker.finished_signal.disconnect(self._on_selftest_finished)
                    self._self_test_worker.error_occurred.disconnect(self._on_selftest_error)
                except (TypeError, RuntimeError):
                    pass
                self._self_test_worker.stop()
                self._self_test_worker = None

            # Create a Worker that streams each output line in real time.
            # The Worker will Connect_RPI internally and disconnect on cleanup.
            self._self_test_worker = Worker(
                self.ssh_handler,
                '/home/robot/Manufacturing_test/aipc_beta/test.py',
                'ecat selftest',
                timeout=300,
            )
            self._self_test_worker.output_ready.connect(self._on_selftest_line)
            self._self_test_worker.finished_signal.connect(self._on_selftest_finished)
            self._self_test_worker.error_occurred.connect(self._on_selftest_error)

            self.self_start_button.setEnabled(False)
            self._set_tabs_locked(True)
            self.test_status_label_start.setText("● Running…")
            self.test_status_label_start.setStyleSheet(self._PILL_RUN_SS)
            self.append_self_message("\n==================Self Test Started=======================\n")
            self.append_self_message("\nWait Test in process..........\n")

            self._self_test_worker.start()

        except Exception as e:
            self.logger.error(f"Error in self_test : {str(e)}", exc_info=True,
                              extra={'func_name': 'start_self_test'})
            QMessageBox.critical(self, "Error", f"SELF TEST FAIL: {str(e)}")
            self.self_start_button.setEnabled(True)

    def _on_selftest_line(self, line):
        """Slot: receives each output line streamed from the self-test Worker."""
        if line:
            self._self_test_lines.append(line)
            self.append_self_message(line)

    def _on_selftest_finished(self):
        """Slot: called when the self-test Worker thread finishes."""
        stdout = '\n'.join(self._self_test_lines)
        stderr = ""  # stderr is captured inside the Worker and emitted via error_occurred
        self.handle_self_test_output(stdout, stderr)
        self.self_start_button.setEnabled(True)
        self._set_tabs_locked(False)
        self._self_test_worker = None

    def _on_selftest_error(self, error_msg):
        """Slot: called when the self-test Worker emits an error."""
        self.append_self_message(f"ERROR: {error_msg}", is_error=True)
        self.logger.error(error_msg, extra={'func_name': 'start_self_test'})
        self.test_status_label_start.setText("● Completed — FAIL")
        self.test_status_label_start.setStyleSheet(self._PILL_FAIL_SS)
        self.self_start_button.setEnabled(True)
        self._set_tabs_locked(False)
        self._self_test_worker = None


    def start_interlock_test(self):
        try:
            # Clean up any existing test
            if self.impedance_scan >= 1:
                excel_logger.reset_sheet("Interlock Test")
            self.impedance_scan += 1
            self.fan_interlock = 30

            if hasattr(self, 'worker') and self.worker:
                try:
                    self.worker.output_ready.disconnect(self.handle_interlock_output)
                    self.worker.finished_signal.disconnect(self.on_interlock_test_finished)
                    self.worker.error_occurred.disconnect(self.handle_interlock_error)
                except (TypeError, RuntimeError):
                    pass
                self.worker.stop()

            # Reset UI state
            self.interlock_console.clear()
            self.append_interlock_message("\n1. Fan Interlock Test Started\n")
            self.open_count = 0
            self.closed_count = 0
            # self.update_interlock_counters()

            # Create new worker
            self.worker = Worker(
                self.ssh_handler,
                '/home/robot/Manufacturing_test/aipc_beta/Interlock.py',
                'ecat test_interlock'
            )

            # Connect signals
            self.worker.output_ready.connect(self.handle_interlock_output)
            self.worker.finished_signal.connect(self.on_interlock_test_finished)
            self.worker.error_occurred.connect(self.handle_interlock_error)

            # Update UI
            self.interlock_start_button.setEnabled(False)
            self.interlock_end_button.setEnabled(True)
            self.test_status_label.setText("● Running…")
            self.test_status_label.setStyleSheet(self._PILL_RUN_SS)

            # Start the thread
            self.worker.start()
            self._set_tabs_locked(True)
            self.interlock_idle_timer.start(self._IDLE_TIMEOUT_MS)

        except Exception as e:
            self.append_interlock_message(f"Failed to start test: {str(e)}", is_error=True)
            self.cleanup_resources()


    def handle_interlock_error(self, error_msg):
        self.interlock_idle_timer.stop()
        self.append_interlock_message(f"ERROR: {error_msg}", is_error=True)
        self.cleanup_resources()


    def handle_interlock_output(self, line):
        """Handle output from the interlock test"""
        # Reset the 5-minute idle watchdog on every received line
        self.interlock_idle_timer.start(self._IDLE_TIMEOUT_MS)

        if "Error in slave initialization" in line:
            QMessageBox.critical(
                self,
                "Critical Error",
                "Slave initialization failed! Please check the EtherCAT connection and restart the test."
            )
            self.interlock_idle_timer.stop()
            self.interlock_start_button.setEnabled(True)
            self.worker.stop()

        if "mailbox error" in line.lower():
            self.append_interlock_message(
                "Mailbox Error on Ethercat please check the Ethercat Data or Contact Support Team",
                is_error=True
            )
            self.interlock_idle_timer.stop()
            self.interlock_start_button.setEnabled(True)
            self.worker.stop()
            return

        if "Cooling Fan Working" in line:
            self.fan_interlock = True
            self.append_interlock_message("✔ Fan Interlock Pass \n\n")
            time.sleep(1)
            self.append_interlock_message("\n2. Switch Interlock Test Started\n\n")
            self.append_interlock_message("\nPress the Interlock Switch.......\n")
            time.sleep(1)

        if "Cooling Fan Warning" in line:
            self.fan_interlock = False
            self.over_all_result = "FAIL"
            self.append_interlock_message("✖ Fan Interlock Fail \n\n", is_error=True)
            # time.sleep(1)
            self.append_interlock_message("\n2. Switch Interlock Test Started\n\n")
            self.append_interlock_message("\nPress the Interlock Switch.......\n")
            # time.sleep(1)
            # self.check = True

        if "Interlock Open" in line:
            self.open_count += 1
            self.interlock_open_label.setStyleSheet(self._STATE_OPEN_SS)
            self.interlock_open_label.setText(f"OPEN")
            if self.open_count == 1:
                self.check_true += 1
                # self.append_interlock_message("Interlock Open detected")
        elif "Interlock Closed" in line:
            self.closed_count += 1
            self.interlock_closed_label.setStyleSheet(self._STATE_CLOSED_SS)
            self.interlock_closed_label.setText(f"CLOSED")
            if self.closed_count == 1:
                self.check_true += 1
                self.append_interlock_message("Interlock Closed detected")
        if self.check_true == 1:
            # self.append_interlock_message("Press the Interlock Switch.......")
            # QMessageBox.information("Press the Interlock switch and ok buttom")
            self.check_true += 1

    def _on_interlock_idle_timeout(self):
        """Called when no output line has been received from the interlock script for 2 minutes."""
        self.append_interlock_message(
            "=== ERROR: No data from Raspberry Pi — EtherCAT data broken. Please contact support team. ===",
            is_error=True,
        )
        self._show_idle_timeout_error("Interlock Test")
        self.worker.stop()
        self.interlock_start_button.setEnabled(True)
        self.interlock_end_button.setEnabled(False)
        self.test_status_label.setText("● Failed")
        self.test_status_label.setStyleSheet(self._PILL_FAIL_SS)
        self._set_tabs_locked(False)

    def end_interlock_test(self):
        try:
            self.interlock_idle_timer.stop()
            if self.fan_interlock != 30:
                open_count = False
                close_count = False
                if hasattr(self, 'worker') and self.worker:
                    self.worker.stop()

                if self.closed_count >= 1:
                    close_count = True

                if self.open_count >= 1:
                    open_count = True

                test_passed = self.closed_count >= 1 and self.open_count >= 1

                if self.fan_interlock:
                    result_msg_i = "TEST PASSED - Fan Interlock detected properly"
                else:
                    result_msg_i = "TEST FAILED -  Fan Interlock test Fail"

                if test_passed:
                    result_msg = "TEST PASSED - Interlock switch detected properly"
                    self.append_interlock_message(result_msg)
                    self.test_status_label.setText("● Completed — PASS")
                    self.test_status_label.setStyleSheet(self._PILL_PASS_SS)
                    count = True
                else:
                    result_msg = f"TEST FAILED -  Interlock test Fail"
                    self.over_all_result = 'FAIL'
                    self.append_interlock_message(result_msg, is_error=True)
                    self.test_status_label.setText("● Completed — FAIL")
                    self.test_status_label.setStyleSheet(self._PILL_FAIL_SS)
                    # Log to Excel

                self.excel_logger.log_interlock_test(
                    test_name='FAN Interlock',
                    open_count='NA',
                    closed_count='NA',
                    test_passed=self.fan_interlock,
                    notes=result_msg_i
                )

                self.excel_logger.log_summary(
                    teststep_data={
                        'teststep': 'FAN Test',
                        'status': "PASS" if self.fan_interlock else "FAIL"  # Updates the existing row
                    }
                )
                self.excel_logger.log_interlock_test(
                    test_name='Switch Interlock',
                    open_count=open_count,
                    closed_count=close_count,
                    test_passed=test_passed,
                    notes=result_msg
                )
                self.excel_logger.log_summary(
                    teststep_data={
                        'teststep': 'Interlock Test',
                        'status':  "PASS" if test_passed else "FAIL"
                    }
                )

                self.excel_logger.log_summary(
                    metadata={
                        'end_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        'overall_result': self.over_all_result
                    }
                )
                self.excel_logger.update_overall_result(self.over_all_result)



            else:
                result_msg = f"TEST aborted without any test"
                self.append_interlock_message(result_msg, is_error=True)

        except Exception as e:
            self.append_interlock_message(f"Error ending test: {str(e)}", is_error=True)
            self.worker.stop()
        finally:
            self.interlock_end_button.setEnabled(False)
            self.interlock_start_button.setEnabled(True)
            self._set_tabs_locked(False)


    # ------------------------------------------------------------------ #
    # Tab-locking helpers                                                  #
    # ------------------------------------------------------------------ #

    def _set_tabs_locked(self, locked):
        """Disable/enable tabs while a test is in progress.

        When *locked* is ``True`` every tab is disabled except the tab that
        is currently visible and the RPI Console tab (so the operator can
        still monitor the SSH console).  When *locked* is ``False`` all tabs
        are re-enabled.
        """
        if not hasattr(self, 'tab_widget'):
            return
        # Resolve the RPI Console index robustly: prefer the stored attribute,
        # fall back to a text-based search so the method stays correct if tab
        # order ever changes.
        rpi_idx = getattr(self, '_RPI_CONSOLE_TAB_INDEX', None)
        if rpi_idx is None:
            for i in range(self.tab_widget.count()):
                if "RPI" in self.tab_widget.tabText(i) or "Console" in self.tab_widget.tabText(i):
                    rpi_idx = i
                    break
            else:
                rpi_idx = self.tab_widget.count() - 1
        current = self.tab_widget.currentIndex()
        for i in range(self.tab_widget.count()):
            self.tab_widget.setTabEnabled(i, not locked or i == current or i == rpi_idx)

    def cleanup_resources(self):
        try:
            if hasattr(self, 'worker') and self.worker:
                self.worker.stop()

            if hasattr(self, '_ssh_console_worker') and self._ssh_console_worker:
                self._ssh_console_worker.stop()

            if hasattr(self, 'ssh_handler') and self.ssh_handler.is_connect:
                self.ssh_handler.SSH_disconnect()

        except Exception as e:
            self.logger.error(f"Cleanup error: {str(e)}", exc_info=True, extra={'func_name': 'cleanup_resources'})
        finally:
            self._set_tabs_locked(False)

    def _show_idle_timeout_error(self, test_name: str = "Test") -> None:
        """Show a prominent, styled error dialog when the idle watchdog fires.

        The dialog is modal, red-themed and carries a clear action message so the
        operator knows exactly what happened and what to do next.
        The timeout duration is derived from :attr:`_IDLE_TIMEOUT_MS` so there is
        a single source of truth.
        """
        timeout_minutes = self._IDLE_TIMEOUT_MS // 60_000
        msg = QMessageBox(self)
        msg.setWindowTitle(f"⚠  {test_name} — Communication Timeout")
        msg.setIcon(QMessageBox.Critical)
        msg.setText(
            "<span style='font-size:13pt; font-weight:bold; color:#c0392b;'>"
            "No data from Raspberry Pi"
            "</span>"
        )
        msg.setInformativeText(
            f"<span style='font-size:10pt;'>"
            f"EtherCAT data stream has been interrupted for {timeout_minutes} minute(s).<br><br>"
            f"<b>Please contact the support team.</b>"
            f"</span>"
        )
        msg.setDetailedText(
            f"The test worker produced no output for {timeout_minutes} consecutive minute(s).\n"
            "This usually means the EtherCAT connection to the Raspberry Pi has\n"
            "been lost or the remote script has crashed.\n\n"
            "Suggested actions:\n"
            "  1. Check the EtherCAT / network cable to the Raspberry Pi.\n"
            "  2. Verify the Raspberry Pi is powered and the script can run.\n"
            "  3. Restart the test from the beginning.\n"
            "  4. If the problem persists, contact the support team."
        )
        msg.setStandardButtons(QMessageBox.Ok)
        msg.setDefaultButton(QMessageBox.Ok)
        # Apply a red-accented stylesheet so the dialog is impossible to miss
        msg.setStyleSheet("""
            QMessageBox {
                background-color: #2b2b2b;
            }
            QLabel {
                color: #f0f0f0;
                font-family: 'Segoe UI', Arial, sans-serif;
            }
            QPushButton {
                background-color: #c0392b;
                color: white;
                border-radius: 6px;
                padding: 6px 22px;
                font-weight: bold;
                font-size: 10pt;
            }
            QPushButton:hover {
                background-color: #e74c3c;
            }
        """)
        msg.exec_()


    def on_interlock_test_finished(self):
        """Clean up after test completion"""
        self.interlock_idle_timer.stop()
        self.ssh_handler.SSH_disconnect()
        self._set_tabs_locked(False)
        # self.ssh_status_label.setText("SSH: Disconnected")
        # self.ssh_status_label.setStyleSheet("background-color: #dc3545; color: white;")


    def reset_interlock_test(self):
        """Reset the test state"""
        self.interlock_start_button.setEnabled(True)
        self.interlock_end_button.setEnabled(False)
        self.test_status_label.setText("Test: Ready")
        self.test_status_label.setStyleSheet(self._PILL_READY_SS)


    def create_unit_setup_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(8)

        # ── Header banner ──────────────────────────────────────────────────
        layout.addWidget(self._make_tab_header(
            "Unit Setup",
            "Enter PCB / assembly information, connect to the Raspberry Pi, and program the OTP.",
            "#1a237e", "#283593"
        ))

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_content = QWidget()
        scroll_layout = QHBoxLayout(scroll_content)
        scroll_layout.setContentsMargins(0, 0, 5, 0)
        scroll_layout.setSpacing(10)

        details_section = QGroupBox()
        details_section.setStyleSheet("""
                QGroupBox {
                    background-color: #f0f0f0;
                    padding: 12px;
                    border-radius: 5px;
                    font-size: 16px;
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    left: 3px;
                    padding: 0 3px;
                }
            """)
        details_layout = QVBoxLayout(details_section)

        # PCB Information Group - Three fields in one line
        pcb_info_group = QWidget()
        pcb_info_layout = QVBoxLayout(pcb_info_group)
        pcb_info_layout.setContentsMargins(0, 0, 0, 0)

        pcb_info_label = QLabel("PCB Information *")
        pcb_info_label.setStyleSheet("color: red;")
        pcb_info_layout.addWidget(pcb_info_label)

        # Horizontal container for the three fields
        pcb_fields_container = QWidget()
        pcb_fields_layout = QHBoxLayout(pcb_fields_container)
        pcb_fields_layout.setContentsMargins(0, 0, 0, 0)
        pcb_fields_layout.setSpacing(10)

        # PCB Part Number (format: XXX-AXXXXX-XXX)
        pcb_pn_group = QWidget()
        pcb_pn_layout = QVBoxLayout(pcb_pn_group)
        pcb_pn_layout.setContentsMargins(0, 0, 0, 0)
        pcb_pn_label = QLabel("Control Part Number")
        pcb_pn_layout.addWidget(pcb_pn_label)

        self.pcb_pn_input = QLineEdit()
        self.pcb_pn_input.setPlaceholderText("Format: 123-A45678-901")
        self.pcb_pn_input.setMinimumHeight(38)
        pcb_pn_layout.addWidget(self.pcb_pn_input)
        pcb_fields_layout.addWidget(pcb_pn_group)

        # PCB Revision (3 characters)
        pcb_rev_group = QWidget()
        pcb_rev_layout = QVBoxLayout(pcb_rev_group)
        pcb_rev_layout.setContentsMargins(0, 0, 0, 0)
        pcb_rev_label = QLabel("Control Revision")
        pcb_rev_layout.addWidget(pcb_rev_label)

        self.pcb_rev_input = QLineEdit()
        self.pcb_rev_input.setPlaceholderText("Format: A")
        self.pcb_rev_input.setMaxLength(3)
        self.pcb_rev_input.setMinimumHeight(38)
        pcb_rev_layout.addWidget(self.pcb_rev_input)
        pcb_fields_layout.addWidget(pcb_rev_group)

        # PCB Serial Number (12 characters)
        pcb_sn_group = QWidget()
        pcb_sn_layout = QVBoxLayout(pcb_sn_group)
        pcb_sn_layout.setContentsMargins(0, 0, 0, 0)
        pcb_sn_label = QLabel("Control Serial Number")
        pcb_sn_layout.addWidget(pcb_sn_label)

        self.pcb_sn_input = QLineEdit()
        # self.pcb_sn_input.setPlaceholderText()
        self.pcb_sn_input.setMaxLength(12)
        self.pcb_sn_input.setMinimumHeight(38)
        pcb_sn_layout.addWidget(self.pcb_sn_input)
        pcb_fields_layout.addWidget(pcb_sn_group)

        pcb_info_layout.addWidget(pcb_fields_container)
        details_layout.addWidget(pcb_info_group)

        # Assembly Information Group - Three fields in one line
        assembly_info_group = QWidget()
        assembly_info_layout = QVBoxLayout(assembly_info_group)
        assembly_info_layout.setContentsMargins(0, 0, 0, 0)

        assembly_info_label = QLabel("Assembly Information *")
        assembly_info_label.setStyleSheet("color: red;")
        assembly_info_layout.addWidget(assembly_info_label)

        # Horizontal container for the three fields
        assembly_fields_container = QWidget()
        assembly_fields_layout = QHBoxLayout(assembly_fields_container)
        assembly_fields_layout.setContentsMargins(0, 0, 0, 0)
        assembly_fields_layout.setSpacing(10)

        # Assembly Part Number (format: XXX-AXXXXX-XXX)
        assembly_pn_group = QWidget()
        assembly_pn_layout = QVBoxLayout(assembly_pn_group)
        assembly_pn_layout.setContentsMargins(0, 0, 0, 0)
        assembly_pn_label = QLabel("Part Number")
        assembly_pn_layout.addWidget(assembly_pn_label)

        self.assembly_pn_input = QLineEdit()
        self.assembly_pn_input.setPlaceholderText("Format: 123-A45678-901")
        self.assembly_pn_input.setMinimumHeight(38)
        assembly_pn_layout.addWidget(self.assembly_pn_input)
        assembly_fields_layout.addWidget(assembly_pn_group)

        # Assembly Revision (3 characters)
        assembly_rev_group = QWidget()
        assembly_rev_layout = QVBoxLayout(assembly_rev_group)
        assembly_rev_layout.setContentsMargins(0, 0, 0, 0)
        assembly_rev_label = QLabel("Revision")
        assembly_rev_layout.addWidget(assembly_rev_label)

        self.assembly_rev_input = QLineEdit()
        self.assembly_rev_input.setPlaceholderText("Format: A")
        self.assembly_rev_input.setMaxLength(3)
        self.assembly_rev_input.setMinimumHeight(38)
        assembly_rev_layout.addWidget(self.assembly_rev_input)
        assembly_fields_layout.addWidget(assembly_rev_group)

        # Assembly Serial Number (12 characters)
        assembly_sn_group = QWidget()
        assembly_sn_layout = QVBoxLayout(assembly_sn_group)
        assembly_sn_layout.setContentsMargins(0, 0, 0, 0)
        assembly_sn_label = QLabel("Serial Number")
        assembly_sn_layout.addWidget(assembly_sn_label)

        self.assembly_sn_input = QLineEdit()
        # self.assembly_sn_input.setPlaceholderText()
        self.assembly_sn_input.setMaxLength(12)
        self.assembly_sn_input.setMinimumHeight(38)
        assembly_sn_layout.addWidget(self.assembly_sn_input)
        assembly_fields_layout.addWidget(assembly_sn_group)

        assembly_info_layout.addWidget(assembly_fields_container)
        details_layout.addWidget(assembly_info_group)

        VN_FN_info_group = QWidget()
        VN_FN_info_layout = QVBoxLayout(VN_FN_info_group)
        VN_FN_info_layout.setContentsMargins(0, 0, 0, 0)

        # Horizontal container for the two fields
        VN_FN_fields_container = QWidget()
        VN_FN_fields_layout = QHBoxLayout(VN_FN_fields_container)
        VN_FN_fields_layout.setContentsMargins(0, 0, 0, 0)
        VN_FN_fields_layout.setSpacing(10)

        # Vendor Name field
        VN_group = QWidget()
        VN_layout = QVBoxLayout(VN_group)  # Fixed: Create new layout for VN_group
        VN_layout.setContentsMargins(0, 0, 0, 0)
        VN_label = QLabel("Vendor Name")
        VN_layout.addWidget(VN_label)

        self.Vendor_name = QLineEdit()
        self.Vendor_name.setMinimumHeight(38)
        VN_layout.addWidget(self.Vendor_name)
        VN_FN_fields_layout.addWidget(VN_group)

        # Fixture Number field
        FN_group = QWidget()
        FN_layout = QVBoxLayout(FN_group)
        FN_layout.setContentsMargins(0, 0, 0, 0)
        FN_label = QLabel("Fixture Number")
        FN_layout.addWidget(FN_label)

        self.Fixture = QLineEdit()
        self.Fixture.setMinimumHeight(38)
        FN_layout.addWidget(self.Fixture)
        VN_FN_fields_layout.addWidget(FN_group)

        VN_FN_info_layout.addWidget(VN_FN_fields_container)
        details_layout.addWidget(VN_FN_info_group)

        # self.Vendor_name = self.create_form_row("Vendor Name", QLineEdit(), details_layout)

        # Rest of the fields
        # self.Test_date = self.create_form_row("Test Date", QLineEdit(), details_layout)
        self.Test_Date = QDateEdit()
        self.Test_Date.setCalendarPopup(True)
        self.Test_Date.setDate(QDate.currentDate())
        self.Test_Date.setDisplayFormat("yyyy-MM-dd")
        self.create_form_row("Test Date", self.Test_Date, details_layout)

        self.Test_name = self.create_form_row("Test Operator Name", QLineEdit(), details_layout)
        # self.VNA_calibration = self.create_form_row("VNA Calibration Date", QLineEdit(), details_layout)
        self.VNA_calibration = QDateEdit()
        self.VNA_calibration.setCalendarPopup(True)
        self.VNA_calibration.setDate(QDate.currentDate())
        self.VNA_calibration.setDisplayFormat("yyyy-MM-dd")
        self.create_form_row("VNA calibration", self.VNA_calibration, details_layout)

        self.VNA_SN = self.create_form_row("VNA SN", QLineEdit(), details_layout)
        self.Ecal_SN = self.create_form_row("Ecal SN", QLineEdit(), details_layout)

        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setFrameShadow(QFrame.Sunken)
        details_layout.addWidget(separator)

        self.product_id = self.create_form_row("Product ID", QLineEdit(), details_layout, True)
        self.esi_revision = self.create_form_row("ESI Revision", QLineEdit(), details_layout, True)
        self.configuration_id = self.create_form_row("Configuration ID", QLineEdit(), details_layout, True)
        self.ethercat_address = self.create_form_row("EtherCAT Address", QLineEdit(), details_layout, True)
        self.firmware_version = self.create_form_row("Firmware Version", QLineEdit(), details_layout, True)

        scroll_layout.addWidget(details_section)

        # Configuration section
        config_section = QGroupBox()
        config_section.setStyleSheet(details_section.styleSheet())
        config_layout = QVBoxLayout(config_section)
        config_layout.setSpacing(15)

        def add_config_widget(label, widget, readonly=False):
            container = QWidget()
            container_layout = QVBoxLayout(container)
            container_layout.setContentsMargins(0, 0, 0, 0)
            container_layout.setSpacing(5)

            lbl = QLabel(label)
            container_layout.addWidget(lbl)

            if isinstance(widget, QLineEdit):
                widget.setMinimumHeight(38)
                if readonly:
                    widget.setReadOnly(True)
                    widget.setStyleSheet("""
                            background-color: #e9ecef;
                            color: #495057;
                            border: 1px solid #ced4da;
                            font-weight: bold;
                        """)

            container_layout.addWidget(widget)
            container_layout.addSpacing(10)
            config_layout.addWidget(container)
            return widget

        self.test_purpose = add_config_widget("Test Purpose", QComboBox())
        self.test_purpose.addItem("High Volume Manufacturing Test")

        otp_program_container = QWidget()
        otp_program_layout = QVBoxLayout(otp_program_container)
        otp_program_layout.setContentsMargins(0, 0, 0, 0)
        otp_program_layout.setSpacing(5)

        self.otp_program_btn = QPushButton("OTP Program")
        self.otp_program_btn.setStyleSheet("""
                QPushButton {
                    background-color: #28a745;
                    color: white;
                    padding: 10px;
                    font-size: 30px;
                    font-weight: bold;
                    min-height: 45px;
                }
                QPushButton:hover {
                    background-color: #218838;
                }
                QPushButton:disabled {
                    background-color: #cccccc;
                    color: #666666;
                }
            """)
        self.otp_program_btn.clicked.connect(self.program_otp)  # Connect to your OTP programming function
        otp_program_layout.addWidget(self.otp_program_btn)
        otp_program_layout.addSpacing(15)
        config_layout.addWidget(otp_program_container)

        # self.pcb_pn = add_config_widget("PCB Part Number", QLineEdit(), True)
        # self.pcb_sn = add_config_widget("PCB Serial Number", QLineEdit(), True)
        # self.pcb_revision = add_config_widget("PCB Revision", QLineEdit(), True)
        # self.assembly_pn = add_config_widget("Assembly Part Number", QLineEdit(), True)
        # self.assembly_sn = add_config_widget("Assembly Serial Number", QLineEdit(), True)
        # self.assembly_revision = add_config_widget("Assembly Revision", QLineEdit(), True)

        auto_load_container = QWidget()
        auto_load_layout = QVBoxLayout(auto_load_container)
        auto_load_layout.setContentsMargins(0, 0, 0, 0)
        auto_load_layout.setSpacing(5)

        self.auto_load_btn = QPushButton("Auto Load and Connect")
        self.auto_load_btn.setStyleSheet("""
                QPushButton {
                    background-color: #007bff; 
                    color: white;
                    padding: 10px;
                    font-size: 20 px;
                    font-weight: bold;
                    min-height: 45px;
                }
                QPushButton:hover {
                    background-color: #0069d9;
                }
                QPushButton:disabled {
                    background-color: #cccccc;
                    color: #666666;
                }
            """)
        self.auto_load_btn.clicked.connect(self.auto_load_connect)
        auto_load_layout.addWidget(self.auto_load_btn)
        auto_load_layout.addSpacing(15)
        config_layout.addWidget(auto_load_container)

        console_group = QGroupBox("Console Output")
        console_group.setStyleSheet("""
                QGroupBox {
                    background-color: white;
                    padding: 10px;
                    border-radius: 5px;
                    font-size: 20px;
                    font-weight: bold;
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    left: 5px;
                    padding: 0 3px;
                }
            """)
        console_layout = QVBoxLayout(console_group)
        console_layout.setContentsMargins(5, 15, 5, 5)
        console_layout.setSpacing(5)

        console_buttons = QHBoxLayout()
        clear_btn = self._make_action_button("⌫  Clear Console", self._BTN_GRAY_SS,
                                             min_height=30, min_width=120)
        clear_btn.clicked.connect(lambda: self.console_output.clear() if hasattr(self, 'console_output') else None)
        console_buttons.addWidget(clear_btn)
        console_buttons.addStretch()
        console_layout.addLayout(console_buttons)

        self.console_output = QTextEdit()
        self.console_output.setReadOnly(True)
        self.console_output.setAcceptRichText(True)
        self.console_output.setStyleSheet("""
            QTextEdit {
                background-color: #0d1117;
                color: #c9d1d9;
                border: 1px solid #30363d;
                border-radius: 4px;
                font-family: 'Courier New', Consolas, monospace;
                font-size: 11pt;
                padding: 6px;
                min-height: 600px;
                selection-background-color: #264f78;
            }
        """)
        console_layout.addWidget(self.console_output)
        config_layout.addWidget(console_group)
        config_layout.addStretch(1)
        # Increase maximum block count for larger output

        scroll_layout.addWidget(config_section)
        scroll.setWidget(scroll_content)
        layout.addWidget(scroll)
        return tab


    def create_ssh_console_tab(self):
        """Create a PuTTY/MobaXterm-style interactive SSH console tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(6)

        # ── Header banner ──────────────────────────────────────────────────
        layout.addWidget(self._make_tab_header(
            "RPI Console",
            "Interactive SSH terminal connected to the Raspberry Pi manufacturing controller.",
            "#212121", "#37474F"
        ))

        # Terminal widget (created first so buttons can reference it directly)
        self._ssh_console_output = TerminalWidget()

        # ── Status / control bar ──────────────────────────────────────────
        status_bar = QHBoxLayout()
        status_bar.setSpacing(8)
        rpi_host = self.ssh_handler.host
        self._ssh_console_status = QLabel(f"⬤  Disconnected  |  {rpi_host}")
        self._ssh_console_status.setStyleSheet(self._PILL_GRAY_SS)
        status_bar.addWidget(self._ssh_console_status)
        status_bar.addStretch()

        self._ssh_connect_btn = self._make_action_button("⚡  Connect", self._BTN_GREEN_SS,
                                                         min_height=32, min_width=110)
        self._ssh_connect_btn.clicked.connect(self._ssh_console_connect)
        status_bar.addWidget(self._ssh_connect_btn)

        self._ssh_disconnect_btn = self._make_action_button("✕  Disconnect", self._BTN_RED_SS,
                                                            min_height=32, min_width=110)
        self._ssh_disconnect_btn.setEnabled(False)
        self._ssh_disconnect_btn.clicked.connect(self._ssh_console_disconnect)
        status_bar.addWidget(self._ssh_disconnect_btn)

        clear_btn = self._make_action_button("⌫  Clear", self._BTN_GRAY_SS,
                                             min_height=32, min_width=80)
        clear_btn.clicked.connect(self._ssh_console_output.clear)
        status_bar.addWidget(clear_btn)

        layout.addLayout(status_bar)
        layout.addWidget(self._ssh_console_output, stretch=1)

        # ── SCP file transfer bar ─────────────────────────────────────────
        scp_bar = QHBoxLayout()
        scp_bar.setSpacing(8)

        self._scp_upload_btn = self._make_action_button("⬆  Upload to RPI", self._BTN_TEAL_SS,
                                                        min_height=32, min_width=140)
        self._scp_upload_btn.setEnabled(False)
        self._scp_upload_btn.clicked.connect(self._scp_upload)
        scp_bar.addWidget(self._scp_upload_btn)

        self._scp_download_btn = self._make_action_button("⬇  Download from RPI", self._BTN_PURPLE_SS,
                                                          min_height=32, min_width=160)
        self._scp_download_btn.setEnabled(False)
        self._scp_download_btn.clicked.connect(self._scp_download)
        scp_bar.addWidget(self._scp_download_btn)

        scp_bar.addStretch()
        layout.addLayout(scp_bar)

        # Hint bar
        hint = QLabel(
            "Click terminal then type directly  |  "
            "↑↓ history  ·  Tab completion  ·  Ctrl+C interrupt  ·  Ctrl+D EOF  ·  Right-click → Paste"
        )
        hint.setStyleSheet("color: #888; font-size: 8pt; padding: 2px 4px;")
        layout.addWidget(hint)

        self._ssh_console_worker = None
        self._scp_worker = None
        return tab

    def _ssh_console_connect(self):
        """Start the SSH console worker and connect to the Raspberry Pi."""
        if self._ssh_console_worker and self._ssh_console_worker.isRunning():
            return
        h = self.ssh_handler
        self._ssh_console_worker = SshConsoleWorker(
            h.host, h.port, h.username, h.password
        )
        self._ssh_console_worker.output_ready.connect(self._ssh_console_append)
        self._ssh_console_worker.connected.connect(self._ssh_console_on_connected)
        self._ssh_console_worker.disconnected.connect(self._ssh_console_on_disconnected)
        self._ssh_console_worker.error_occurred.connect(self._ssh_console_on_error)
        self._ssh_console_worker.start()
        self._ssh_connect_btn.setEnabled(False)
        self._ssh_console_status.setText(f"⬤  Connecting…  |  {h.host}")
        self._ssh_console_status.setStyleSheet(self._PILL_RUN_SS)

    def _ssh_console_disconnect(self):
        """Stop the SSH console worker."""
        if self._ssh_console_worker:
            self._ssh_console_worker.stop()

    def _ssh_console_append(self, text):
        """Forward received SSH output to the terminal widget."""
        self._ssh_console_output.write(text)

    def _ssh_console_on_connected(self):
        h = self.ssh_handler
        self._ssh_console_status.setText(f"⬤  Connected  |  {h.host}")
        self._ssh_console_status.setStyleSheet(self._PILL_PASS_SS)
        self._ssh_connect_btn.setEnabled(False)
        self._ssh_disconnect_btn.setEnabled(True)
        self._scp_upload_btn.setEnabled(True)
        self._scp_download_btn.setEnabled(True)
        self._ssh_console_output.set_send_fn(self._ssh_console_worker.send_command)
        self._ssh_console_output.setFocus()

    def _ssh_console_on_disconnected(self):
        h = self.ssh_handler
        self._ssh_console_status.setText(f"⬤  Disconnected  |  {h.host}")
        self._ssh_console_status.setStyleSheet(self._PILL_GRAY_SS)
        self._ssh_connect_btn.setEnabled(True)
        self._ssh_disconnect_btn.setEnabled(False)
        self._scp_upload_btn.setEnabled(False)
        self._scp_download_btn.setEnabled(False)
        self._ssh_console_output.set_send_fn(None)

    def _ssh_console_on_error(self, msg):
        self._ssh_console_output.write(f'\n[ERROR] {msg}\n')
        self._ssh_console_on_disconnected()

    # ------------------------------------------------------------------
    # SCP file transfer
    # ------------------------------------------------------------------
    def _scp_upload(self):
        """Let the user pick a local file then browse the RPI to choose destination."""
        import posixpath

        # 1. Choose local file with a native file picker
        local_path, _ = QFileDialog.getOpenFileName(
            self, "Select file to upload to RPI"
        )
        if not local_path:
            return

        # 2. Browse the RPI filesystem to select the destination directory
        h = self.ssh_handler
        browser = RemoteFileBrowserDialog(
            h.host, h.port, h.username, h.password,
            mode='dir',
            start_path=f'/home/{h.username}',
            parent=self
        )
        if browser.exec_() != QDialog.Accepted:
            return
        remote_path = posixpath.join(
            browser.selected_path, os.path.basename(local_path)
        )
        self._run_scp_worker('upload', local_path, remote_path)

    def _scp_download(self):
        """Browse the RPI filesystem to pick a file, then choose local save location."""
        h = self.ssh_handler

        # 1. Browse the RPI filesystem to select the remote file
        browser = RemoteFileBrowserDialog(
            h.host, h.port, h.username, h.password,
            mode='file',
            start_path=f'/home/{h.username}',
            parent=self
        )
        if browser.exec_() != QDialog.Accepted:
            return
        remote_path = browser.selected_path

        # 2. Choose local save path with a native save dialog
        local_path, _ = QFileDialog.getSaveFileName(
            self, "Save downloaded file as",
            os.path.basename(remote_path)
        )
        if not local_path:
            return
        self._run_scp_worker('download', local_path, remote_path)

    def _run_scp_worker(self, direction, local_path, remote_path):
        """Start a ScpWorker for the given direction."""
        if self._scp_worker and self._scp_worker.isRunning():
            QMessageBox.information(self, "SCP Busy",
                                    "A file transfer is already in progress.")
            return
        h = self.ssh_handler
        self._scp_worker = ScpWorker(
            h.host, h.port, h.username, h.password,
            direction, local_path, remote_path
        )
        self._scp_worker.progress.connect(self._ssh_console_output.write)
        self._scp_worker.finished.connect(self._on_scp_finished)
        self._scp_upload_btn.setEnabled(False)
        self._scp_download_btn.setEnabled(False)
        self._scp_worker.start()

    def _on_scp_finished(self, success, message):
        self._ssh_console_output.write('\n' + message + '\n')
        # Re-enable SCP buttons only when still connected
        if self._ssh_console_worker and self._ssh_console_worker.isRunning():
            self._scp_upload_btn.setEnabled(True)
            self._scp_download_btn.setEnabled(True)

    # ------------------------------------------------------------------
    # RPI Console tab password gate
    # ------------------------------------------------------------------
    _RPI_CONSOLE_PASSWORD = "lam@rpi"

    def _on_tab_changed(self, index):
        """Guard the RPI Console tab with a password prompt."""
        if index != self._RPI_CONSOLE_TAB_INDEX:
            # Remember this as the last accessible tab
            self._last_tab_index = index
            return
        if self._rpi_console_unlocked:
            return
        pwd, ok = QInputDialog.getText(
            self, "RPI Console – Access Required",
            "Enter password to open the RPI Console:",
            QLineEdit.Password
        )
        if ok and pwd == self._RPI_CONSOLE_PASSWORD:
            self._rpi_console_unlocked = True
        else:
            if ok:  # wrong password was entered
                QMessageBox.warning(self, "Access Denied", "Incorrect password.")
            # Switch back to the previous tab without re-triggering the guard
            self.tab_widget.blockSignals(True)
            self.tab_widget.setCurrentIndex(self._last_tab_index)
            self.tab_widget.blockSignals(False)

    def init_ui(self):
        """Initialize the user interface with responsive layouts"""
        # Create a main scroll area
        main_scroll = QScrollArea()
        main_scroll.setWidgetResizable(True)

        # Create the central widget that will hold everything
        central_widget = QWidget()
        self.main_layout = QVBoxLayout(central_widget)

        # Create tab widget
        self.tab_widget = QTabWidget()
        self.main_layout.addWidget(self.tab_widget)

        # Add tabs
        self.tab_widget.addTab(self.create_unit_setup_tab(), "Unit Setup")
        self.tab_widget.addTab(self.create_test_tab("Interlock System Check"), "Interlock")
        self.tab_widget.addTab(self.create_test_tab("Verify BNC Port"), "BNC Port Verification")
        self.tab_widget.addTab(self.create_test_tab("Impedance Scan"), "Impedance Scan")
        self.tab_widget.addTab(self.create_test_tab("Resistance Test"), "Resistance")
        self.tab_widget.addTab(self.create_test_tab("System Self Test"), "Self Test")
        self.tab_widget.addTab(self.create_test_tab("VNA Calibration"), "VNA Cal")
        self.tab_widget.addTab(self.create_ssh_console_tab(), "RPI Console")

        # RPI Console tab is the last tab; guard it with a password
        self._RPI_CONSOLE_TAB_INDEX = self.tab_widget.count() - 1
        self._rpi_console_unlocked = False
        self._last_tab_index = 0
        self.tab_widget.currentChanged.connect(self._on_tab_changed)

        # Set the central widget
        main_scroll.setWidget(central_widget)
        self.setCentralWidget(main_scroll)


if __name__ == "__main__":
    try:
        app = QApplication(sys.argv)
        app.setStyle('Fusion')
        font = QFont()
        font.setPointSize(10)
        app.setFont(font)
        window = TestStationInterface()
        window.show()
        sys.exit(app.exec_())
    except Exception as e:
        logger.critical(f"Application crashed: {str(e)}", exc_info=True)
        QMessageBox.critical(None, "Fatal Error", f"The application encountered a fatal error:\n{str(e)}")
        sys.exit(1)
